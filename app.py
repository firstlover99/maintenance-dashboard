import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import re
import requests
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="보전팀 통합 분석 시스템",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    .main-title {font-size:26px; font-weight:700; color:#1e3a5f; margin-bottom:2px;}
    .sub-title  {font-size:12px; color:#888; margin-bottom:16px;}
    .kpi-box    {background:#f0f4fa; border-radius:10px; padding:14px 18px;
                 text-align:center; height:90px;}
    .kpi-val    {font-size:28px; font-weight:700; color:#1e3a5f;}
    .kpi-unit   {font-size:14px; color:#888;}
    .kpi-label  {font-size:11px; color:#666; margin-top:4px;}
    .warn-box   {background:#fff3cd; border-left:4px solid #ffc107;
                 padding:10px 14px; border-radius:4px; margin:8px 0; font-size:13px;}
    .ok-box     {background:#d4edda; border-left:4px solid #28a745;
                 padding:10px 14px; border-radius:4px; margin:8px 0; font-size:13px;}
    .err-box    {background:#f8d7da; border-left:4px solid #dc3545;
                 padding:10px 14px; border-radius:4px; margin:8px 0; font-size:13px;}
    .card-red   {border-left:5px solid #e74c3c; background:#fff5f5;
                 padding:12px 16px; border-radius:0 8px 8px 0; margin-bottom:8px;}
    .card-org   {border-left:5px solid #e67e22; background:#fff9f0;
                 padding:12px 16px; border-radius:0 8px 8px 0; margin-bottom:8px;}
    .card-grn   {border-left:5px solid #27ae60; background:#f0fff4;
                 padding:12px 16px; border-radius:0 8px 8px 0; margin-bottom:8px;}
    .stTabs [data-baseweb="tab"] {font-size:13px; font-weight:600;}
    .code-badge {display:inline-block; background:#e8f4fd; color:#1a5276;
                 border-radius:4px; padding:2px 7px; font-size:11px;
                 font-weight:600; margin:2px;}
    /* ── 예방정비 호버 툴팁 ─────────────────────── */
    .pm-card { position:relative; cursor:default; }
    .pm-tooltip {
        display:none; position:absolute; z-index:9999;
        background:#fff; border:1px solid #d0d0d0;
        border-radius:8px; padding:12px 16px;
        min-width:340px; max-width:480px;
        left:0; top:105%;
        box-shadow:0 6px 20px rgba(0,0,0,.15);
        font-size:12px; line-height:2.0;
        pointer-events:none;
    }
    .pm-card:hover .pm-tooltip { display:block; }
    /* ── 인쇄/PDF 잘림 방지 ─────────────────────── */
    @media print {
        section[data-testid="stSidebar"] { display:none !important; }
        .block-container { padding:0 !important; max-width:100% !important; }
        .stApp, .main, [data-testid="stAppViewContainer"]
            { overflow:visible !important; height:auto !important; }
        .card-red,.card-org,.card-grn { page-break-inside:avoid; }
        .stPlotlyChart { page-break-inside:avoid; }
        @page { size:A4; margin:15mm; }
    }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# 상수 및 정규화 맵
# ══════════════════════════════════════════════════════
VALID_YEAR_MIN, VALID_YEAR_MAX = 2010, 2030

EQUIP_NORM = {
    '로보트':'로봇','로보트 ':'로봇',' 로보트':'로봇','로봇 ':'로봇',' 로봇':'로봇',
    '블래킹':'블랭킹','플랭킹':'블랭킹','B/K':'블랭킹',
    '파일러1':'파일러','파일러 ':'파일러',
    '지하컨베어':'컨베어','텔레스코픽컨베어':'텔레스코프',
    '1500T':'프레스-1500T','1200T':'프레스-1200T',
    '800T':'프레스-800T','600T':'프레스-600T',
}

# ── 라인명 정규화 맵 (입력 오류 통합) ──────────────────
import re as _re

def norm_line(v):
    """라인명 정규화 — 대소문자/오타/단위표기 통합"""
    if v is None or (isinstance(v, float) and __import__('numpy').isnan(v)):
        return v
    s = str(v).strip()
    if not s or s in ('nan','None',''): return s
    # 대문자 통일
    su = s.upper()
    # 프레스 톤수 정규화: 숫자+t/톤/T → 숫자+T
    su = _re.sub(r'(\d+)(톤|t\b)', lambda m: m.group(1)+'T', su)
    # ASSY 통합: ASS'Y, Assy, ASSY → ASSY
    su = _re.sub(r"ASS'?Y", 'ASSY', su, flags=_re.IGNORECASE)
    # 연속 공백 정리
    su = _re.sub(r'\s+', ' ', su).strip()
    return su

# ── 표준 코드 정의 ──────────────────────────────────
# 고장계통코드
FAULT_SYSTEM_MAP = {
    '전기':  ['케이블','단선','전기','전원','퓨즈','eocr','인버터','누전','차단기','접촉불량','배선'],
    '제어':  ['plc','프로그램','통신','로직','파라미터','설정값','티칭','teaching','신호','출력'],
    '유압':  ['오일','유압','누유','펌프','밸브','압력','실린더','호스','피팅'],
    '공압':  ['air','에어','공압','컴프레셔','압축','솔레노이드','에어밸브','니플','튜브'],
    '기계':  ['베어링','볼트','마모','균열','파손','벨트','체인','기어','변형','용접','스프링','핀'],
    '센서':  ['센서','l/s','ls','리미트','근접','광전','포토','encoder','엔코더','감지'],
    '안전':  ['안전','emergency','비상','light curtain','도어','인터락','guard'],
    '윤활':  ['그리스','윤활','오일팬','급유','grease'],
}

# 원인코드
CAUSE_CODE_MAP = {
    '마모':      ['마모','닳','wear','worn','파손','변형','균열','크랙'],
    '단선':      ['단선','단락','끊','케이블','배선','접촉','접점'],
    '오염':      ['오염','먼지','이물','슬러지','찌꺼기','막힘','clogging'],
    '간섭':      ['간섭','충돌','걸림','jamming','jam','부딪'],
    '설정값이탈':['설정','파라미터','티칭','teaching','캘리브','calibr','원점'],
    '품질기인':  ['품질','이종','버','burr','불량','가공','제품'],
    '작업자기인':['작업자','조작','인위','human','사람','실수'],
    '자연열화':  ['열화','노후','aging','수명','life','교체주기'],
}

# 조치코드
ACTION_CODE_MAP = {
    '교체': ['교체','교환','신품','spare','부품','교환','신규','대체'],
    '수리': ['수리','용접','패치','보수','fix','repair'],
    '조정': ['조정','셋팅','설정','세팅','재조임','조임','adjust','calibr'],
    '티칭': ['티칭','teaching','원점','원위치','좌표'],
    '청소': ['청소','세척','cleaning','clean','제거','blow','에어블로'],
    '리셋': ['리셋','reset','재기동','재가동','원복','복구','재설정'],
    '개선': ['개선','방지','대책','영구','恒久','항구','modification'],
    '예방': ['예방','pm','p/m','계획','정기'],
}

def assign_fault_system(현상, 원인, 고장부위=''):
    """고장계통코드 자동 분류"""
    text = (str(현상 or '') + str(원인 or '') + str(고장부위 or '')).lower()
    for code, kws in FAULT_SYSTEM_MAP.items():
        if any(k in text for k in kws):
            return code
    return '기타'

def assign_cause_code(원인, 현상=''):
    """원인코드 자동 분류"""
    text = (str(원인 or '') + str(현상 or '')).lower()
    for code, kws in CAUSE_CODE_MAP.items():
        if any(k in text for k in kws):
            return code
    return '기타'

def assign_action_code(조치):
    """조치코드 자동 분류"""
    text = str(조치 or '').lower()
    for code, kws in ACTION_CODE_MAP.items():
        if any(k in text for k in kws):
            return code
    return '기타'

# ══════════════════════════════════════════════════════
# 유틸 함수
# ══════════════════════════════════════════════════════
def parse_dt(val):
    if val is None: return None
    if isinstance(val, float):
        if np.isnan(val): return None
        return None
    if isinstance(val, datetime): return val
    if hasattr(val, 'to_pydatetime'):
        try: return val.to_pydatetime()
        except: return None
    if isinstance(val, str):
        val = val.strip()
        if val in ('','00:00:00','None','0','NaT','nan','NaN'): return None
        for fmt in ['%Y-%m-%d %H:%M:%S','%Y-%m-%d %H:%M','%Y-%m-%d',
                    '%Y/%m/%d %H:%M:%S','%Y/%m/%d %H:%M','%Y/%m/%d',
                    '%Y.%m.%d %H:%M:%S','%Y.%m.%d %H:%M','%Y.%m.%d']:
            try: return datetime.strptime(val, fmt)
            except: pass
    return None

def sanitize_dt(dt):
    if dt is None: return None
    try: return dt if VALID_YEAR_MIN <= dt.year <= VALID_YEAR_MAX else None
    except: return None

def to_float_safe(v):
    try:
        f = float(v)
        return f if 0 < f <= 1440 else None
    except: return None

def norm_equip(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return None
    s = str(v).strip() if not isinstance(v, str) else v.strip()
    if s in ('','nan','None','0'): return None
    return EQUIP_NORM.get(s, s)

def classify_fault(현상, 원인):
    """기존 호환 유지"""
    text = (str(현상 or '') + str(원인 or '')).lower()
    if any(k in text for k in ['케이블','단선','전기','전원','퓨즈','eocr','인버터','센서','스위치','접촉','누전']): return '전기/제어'
    if any(k in text for k in ['오일','유압','누유','펌프','실린더','밸브','압력','윤활','그리스']): return '유압/윤활'
    if any(k in text for k in ['베어링','볼트','마모','균열','파손','벨트','체인','기어','용접','변형']): return '기계적결함'
    if any(k in text for k in ['티칭','teaching','프로그램','통신','plc','로직','파라미터','설정값']): return '제어/프로그램'
    if any(k in text for k in ['작업자','이종','투입','조작','인위','사람']): return '작업자과실'
    if any(k in text for k in ['예방','pm','계획','정기점검']): return '예방보전'
    return '기타/불명'

def classify_action(조치):
    text = str(조치 or '').lower()
    if any(k in text for k in ['교체','교환','신품','부품','spare']): return '부품교체'
    if any(k in text for k in ['티칭','teaching','티이칭']): return '티칭수정'
    if any(k in text for k in ['예방','pm ','p/m','계획','정기']): return '예방보전'
    if any(k in text for k in ['조정','셋팅','설정','세팅','재조임','조임']): return '조정/설정'
    if any(k in text for k in ['리셋','reset','재기동','재가동','원복','복구']): return '긴급복구'
    return '기타'

def classify_bm_pm(row):
    """BM/PM 분류"""
    if row.get('조치유형') == '예방보전': return 'PM(계획)'
    if row.get('고장분류') == '예방보전': return 'PM(계획)'
    txt = str(row.get('현상') or '') + str(row.get('원인') or '') + str(row.get('조치내역') or '')
    if any(k in txt for k in ['예방','PM','계획정비','정기점검','P/M']): return 'PM(계획)'
    return 'BM(돌발)'

# ══════════════════════════════════════════════════════
# 실제 근무 인원 기준 리스트 (유사도 매칭 기준)
# ══════════════════════════════════════════════════════
from difflib import SequenceMatcher

VALID_WORKERS = [
    # ── 차체팀 (기존 17명) ──────────────────────────
    '정한식','이준호','박진만','이태진','최병화','송치원',
    '김상진','최문석','강지용','최성진','노현우','임찬영',
    '한은수','이민수','최수한','지훈태','원태양',
    # ── 광주공장 (추가 18명) ────────────────────────
    '황동건','김재훈','박한빈','이기상','최영조','김동희',
    '주태훈','신철','박리건','장성용','황장현','이창용',
    '임대웅','이현수','심지형','박성수','김태규','차윤환',
]
WORKER_MATCH_THRESHOLD = 0.65  # 유사도 임계값 (0.65 = 오타 최대 복원)

def match_worker_name(name: str) -> str | None:
    """
    입력 이름을 VALID_WORKERS 와 유사도 비교.
    - 완전일치 → 즉시 반환
    - 유사도 >= WORKER_MATCH_THRESHOLD → 가장 유사한 기준 이름 반환
    - 미달 → None (삭제 대상)
    """
    name = name.strip()
    # 완전일치 우선
    if name in VALID_WORKERS:
        return name
    # 유사도 비교
    best_name  = None
    best_score = 0.0
    for valid in VALID_WORKERS:
        score = SequenceMatcher(None, name, valid).ratio()
        if score > best_score:
            best_score = score
            best_name  = valid
    if best_score >= WORKER_MATCH_THRESHOLD:
        return best_name
    return None  # 매칭 실패 → 삭제


def parse_workers(조치자_val):
    NOISE = {'야간','주간','주간조','야간조','주야간','업체','가동중','조치','기타'}
    if not 조치자_val or not isinstance(조치자_val, str): return []
    workers = [w.strip() for w in re.split(r'[,/.\+]+', 조치자_val) if w.strip()]
    # 공백 단독 구분 케이스: 분리 후 한글 2자 이상 토큰만 추출
    expanded = []
    for w in workers:
        sub = [s.strip() for s in w.split() if s.strip()]
        # 한글 포함 2자 이상 토큰이 2개 이상이면 공백이 구분자인 것으로 판단
        korean_tokens = [s for s in sub if len(s) >= 2 and re.search(r'[가-힣]', s)]
        if len(korean_tokens) >= 2:
            expanded.extend(korean_tokens)
        else:
            expanded.append(w)
    workers = [w for w in expanded if len(w) >= 2]
    workers = [w for w in workers if re.search(r'[가-힣]', w)]
    workers = [w for w in workers if w not in NOISE]
    # ── 유사도 기반 이름 정규화 ─────────────────────────
    # 기준 인원과 매칭 성공 → 표준 이름으로 교체
    # 매칭 실패(유사도 < 0.7) → 삭제
    matched = []
    for w in workers:
        std_name = match_worker_name(w)
        if std_name:
            matched.append(std_name)
        # else: 매칭 실패 → 리스트에 추가하지 않음(삭제)
    # 중복 제거 (같은 건에 동일인 중복 파싱 방지)
    seen = []
    for w in matched:
        if w not in seen:
            seen.append(w)
    return seen


def parse_workers_with_type(조치자_val):
    """
    조치자 파싱 + 출동유형(단독/협업) + 인원수 반환
    반환: (workers_list, 출동유형, 인원수)
    """
    workers = parse_workers(조치자_val)
    cnt = len(workers)
    if cnt == 0:   return [], '미상', 0
    if cnt == 1:   return workers, '단독', 1
    return workers, '협업', cnt

def get_세부분류(row):
    유형 = str(row.get('설비유형') or '').strip()
    설비 = str(row.get('고장설비') or '').strip().upper()
    if '로봇' in 유형 or '로보트' in 유형:
        m = re.match(r'R(\d+)', 설비)
        return f'로봇-{설비}' if m else '로봇-기타'
    if '지그' in 유형:
        if 설비.startswith('A'): return '지그-A계열(조립)'
        if 설비.startswith('S'): return '지그-S계열(스터드)'
        if 'PLT' in 설비:        return '지그-PLT(팔레트)'
        if 설비.startswith('CS'): return '지그-CS계열'
        if 설비.startswith('FA'): return '지그-FA계열'
        if 설비.startswith('M'):  return '지그-M계열'
        return '지그-기타'
    return 유형 if 유형 else '기타'

def get_고장부위_그룹(v):
    v = str(v or '').strip()
    if not v or v == 'nan': return '기타'
    if '일시정지' in v: return '일시정지'
    if '에러' in v:     return '에러'
    if 'L/S' in v or 'LS' in v: return 'L/S이상'
    if '스터드' in v:   return '스터드'
    if 'T/C' in v or 'TC' in v: return 'T/C이상'
    if '센서' in v:     return '센서'
    if '실러' in v:     return '실러'
    if '품질' in v:     return '품질'
    if '냉각' in v:     return '냉각수'
    if '그리퍼' in v:   return '그리퍼'
    if '티칭' in v:     return '티칭수정'
    if 'AIR' in v:      return 'AIR'
    if '파트' in v:     return '파트이상'
    return '기타'

# ── 재발 판정 함수 ──────────────────────────────────
def calc_recurrence(df, window_days=90):
    """
    동일 설비_KEY + 동일 고장부위 기준으로
    완료시각 → 다음 건 출동시각 간격(분) 기준 재발 판정.
    완료/출동시각 없으면 발생일시 fallback.
    반환: 재발여부(bool) Series
    """
    df = df.sort_values('발생일시').copy()
    df['재발여부_계산'] = False
    df['재발KEY'] = df['설비_KEY'].astype(str) + '||' + df['고장부위'].fillna('').astype(str)
    for key, grp in df.groupby('재발KEY'):
        idx = grp.index.tolist()
        for i in range(1, len(idx)):
            # 이전 건 완료시각, 다음 건 출동시각 우선
            prev_end = grp.loc[idx[i-1], '완료시각'] if '완료시각' in grp.columns else None
            next_start = grp.loc[idx[i], '출동시각'] if '출동시각' in grp.columns else None
            # fallback: 발생일시 사용
            if prev_end is None or pd.isna(prev_end):
                prev_end = grp.loc[idx[i-1], '발생일시']
            if next_start is None or pd.isna(next_start):
                next_start = grp.loc[idx[i], '발생일시']
            if prev_end and next_start:
                try:
                    gap_days = (next_start - prev_end).total_seconds() / 86400
                    if 0 <= gap_days <= window_days:
                        df.loc[idx[i], '재발여부_계산'] = True
                except:
                    pass
    return df['재발여부_계산']


# ── 고장 클러스터링 함수 ─────────────────────────────
def cluster_faults(grp, cluster_min=60):
    """
    동일 설비_KEY 그룹 내에서
    이전 건 완료시각 ~ 다음 건 출동시각 간격이 cluster_min 이내이면
    동일 고장으로 통합. BM(돌발) 건만 대상.
    반환: 클러스터링된 DataFrame (건수 감소, 수리시간 합산)
    """
    grp = grp.sort_values('발생일시').copy()
    clusters = []
    current = None
    for _, row in grp.iterrows():
        if current is None:
            current = row.copy()
            current['_cluster_count'] = 1
            continue
        # 이전 완료 ~ 현재 출동 간격 계산 (분)
        prev_end   = current.get('완료시각')
        next_start = row.get('출동시각')
        if prev_end is None or pd.isna(prev_end):
            prev_end = current.get('발생일시')
        if next_start is None or pd.isna(next_start):
            next_start = row.get('발생일시')
        gap_min = None
        if prev_end is not None and next_start is not None:
            try:
                gap_min = (next_start - prev_end).total_seconds() / 60
            except:
                gap_min = None
        if gap_min is not None and 0 <= gap_min <= cluster_min:
            # 동일 고장 — 수리시간 합산, 완료시각 갱신
            cur_mtr = current.get('소요시간') or 0
            row_mtr = row.get('소요시간') or 0
            current['소요시간'] = cur_mtr + row_mtr
            if row.get('완료시각') and not pd.isna(row.get('완료시각')):
                current['완료시각'] = row['완료시각']
            current['_cluster_count'] = current.get('_cluster_count', 1) + 1
        else:
            clusters.append(current)
            current = row.copy()
            current['_cluster_count'] = 1
    if current is not None:
        clusters.append(current)
    result = pd.DataFrame(clusters)
    if '_cluster_count' not in result.columns:
        result['_cluster_count'] = 1
    return result


# ── 분석기간 내 월~토 가동시간 계산 (방법 B) ───────────
def calc_worktime_hours(start_dt, end_dt, daily_hours=15.0):
    """
    start_dt ~ end_dt 기간 중 월~토(일요일 제외) 일수 × daily_hours
    반환: 근무가동시간(시간, float)
    """
    if start_dt is None or end_dt is None:
        return 0.0
    total_days = (end_dt.date() - start_dt.date()).days + 1
    # 일요일(weekday==6) 제외
    work_days = sum(
        1 for i in range(total_days)
        if (start_dt.date() + pd.Timedelta(days=i)).weekday() != 6
    )
    return work_days * daily_hours

# ══════════════════════════════════════════════════════
# 파일 로드
# ══════════════════════════════════════════════════════
def _enrich_codes(df):
    """
    표준 코드 컬럼 자동 생성 (기존 컬럼에 없을 경우만)
    - 고장계통코드, 원인코드, 조치코드, 보전구분(BM/PM)
    """
    if '고장계통코드' not in df.columns:
        df['고장계통코드'] = df.apply(
            lambda r: assign_fault_system(r.get('현상'), r.get('원인'), r.get('고장부위')), axis=1)
    if '원인코드' not in df.columns:
        df['원인코드'] = df.apply(
            lambda r: assign_cause_code(r.get('원인'), r.get('현상')), axis=1)
    if '조치코드' not in df.columns:
        df['조치코드'] = df.apply(
            lambda r: assign_action_code(r.get('조치내역')), axis=1)
    if '보전구분' not in df.columns:
        df['보전구분'] = df.apply(classify_bm_pm, axis=1)
    # 예비품사용여부
    if '예비품사용여부' not in df.columns:
        df['예비품사용여부'] = df['조치코드'].apply(
            lambda c: 'Y' if c == '교체' else 'N')
    # 계획정비전환대상 (임시 — 재발3회이상 or MTTR>=60 기준)
    if '계획정비전환대상' not in df.columns:
        df['계획정비전환대상'] = '-'   # merge 후 별도 계산
    return df

def load_press(file_obj):
    xf = pd.ExcelFile(file_obj)
    sheet = '설비보전현황_통합' if '설비보전현황_통합' in xf.sheet_names else xf.sheet_names[0]
    file_obj.seek(0)
    df = pd.read_excel(file_obj, sheet_name=sheet, header=0)
    df.columns = ['년','월','일','주','라인','정지시각','출동시각','완료시각',
                  '소요시간','설비유형','고장설비','고장부위','현상','원인','조치내역','조치자','비고']
    for col in ['라인','설비유형','고장설비','고장부위','현상','원인','조치내역','조치자','비고']:
        df[col] = df[col].apply(
            lambda v: None if (v is None or (isinstance(v,float) and pd.isna(v)))
            else str(v).strip() if str(v).strip() not in ('nan','None','') else None)
    for col in ['정지시각','출동시각','완료시각']:
        df[col] = df[col].apply(parse_dt).apply(sanitize_dt)
    def make_dt(row):
        if pd.notna(row['정지시각']): return row['정지시각']
        if pd.notna(row['출동시각']): return row['출동시각']
        if pd.notna(row['완료시각']): return row['완료시각']
        try:
            y,m,d = int(row['년']),int(row['월']),int(row['일'])
            if 1<=m<=12 and 1<=d<=31: return datetime(y,m,d)
        except: pass
        return None
    df['발생일시'] = df.apply(make_dt, axis=1).apply(sanitize_dt)
    def fix_dur(row):
        v = to_float_safe(row['소요시간'])
        if v: return v
        if pd.notna(row['완료시각']) and pd.notna(row['출동시각']):
            try:
                d = (row['완료시각']-row['출동시각']).total_seconds()/60
                return round(d,1) if d>0 else None
            except: return None
        return None
    df['소요시간'] = df.apply(fix_dur, axis=1)
    df['설비유형'] = df['설비유형'].apply(norm_equip)
    df['고장분류'] = df.apply(lambda r: classify_fault(r['현상'],r['원인']), axis=1)
    df['조치유형'] = df['조치내역'].apply(classify_action)
    df['파일출처'] = '프레스'
    df['차종'] = None
    df['조치'] = df['조치내역']
    df['라인_차종'] = df['라인'].astype(str).str.strip()
    df['설비_KEY'] = df['라인_차종'] + ' | ' + df['고장설비'].astype(str).str.strip()
    return _enrich_codes(df)

def load_robot(file_obj):
    df = pd.read_excel(file_obj, sheet_name='Sheet1', header=0)
    df.columns = ['발생일시_raw','월','일','주','라인','라인_KEY','차종','설비유형',
                  '고장설비','고장부위','고장부위_STD','현상','원인','조치',
                  '소요시간','정지시각','출동시각','완료시각','조치자','비고','NO']
    for col in ['정지시각','출동시각','완료시각']:
        df[col] = df[col].apply(parse_dt).apply(sanitize_dt)
    def make_dt(row):
        dt = sanitize_dt(parse_dt(row['발생일시_raw']))
        if dt: return dt
        if pd.notna(row['출동시각']): return row['출동시각']
        if pd.notna(row['정지시각']): return row['정지시각']
        try: return sanitize_dt(datetime(2025,int(row['월']),int(row['일'])))
        except: return None
    df['발생일시'] = df.apply(make_dt, axis=1)
    def fix_dur(row):
        v = to_float_safe(row['소요시간'])
        if v: return v
        if row['완료시각'] and row['출동시각']:
            try:
                d=(row['완료시각']-row['출동시각']).total_seconds()/60
                return round(d,1) if d>0 else None
            except: return None
        return None
    df['소요시간'] = df.apply(fix_dur, axis=1)
    df['설비유형'] = df['설비유형'].apply(norm_equip)
    df['고장분류'] = df.apply(lambda r: classify_fault(r['현상'],r['원인']), axis=1)
    df['조치유형'] = df['조치'].apply(classify_action)
    df['파일출처'] = '로봇/지그'
    df['조치내역'] = df['조치']
    df['년'] = df['발생일시'].apply(lambda x: x.year if x else None)
    df['차종_clean'] = df['차종'].fillna('').astype(str).str.strip().replace({'nan':'','NaN':'','None':''})
    def make_차종라인(r):
        라인 = str(r['라인']).strip() if pd.notna(r['라인']) and str(r['라인']).strip() not in ('','nan','NaN') else ''
        차종 = r['차종_clean']
        if 차종 and 라인: return f"{차종} / {라인}"
        if 라인: return 라인
        if 차종: return f"{차종} / (미상)"
        return '미상'
    df['라인_차종'] = df.apply(make_차종라인, axis=1)
    df['설비_KEY'] = df['라인_차종'] + ' | ' + df['고장설비'].astype(str).str.strip()
    return _enrich_codes(df)

def detect_and_load(file_obj, fname=''):
    try:
        xf = pd.ExcelFile(file_obj)
        sheets = xf.sheet_names
        file_obj.seek(0)
        if '설비보전현황_통합' in sheets:
            return load_press(file_obj), 'press'
        target = 'Sheet1' if 'Sheet1' in sheets else sheets[0]
        df_h = pd.read_excel(file_obj, sheet_name=target, nrows=2)
        file_obj.seek(0)
        cols = df_h.columns.tolist()
        if any(c in cols for c in ['라인_KEY','차종','고장부위_STD']):
            return load_robot(file_obj), 'robot'
        if all(c in cols for c in ['년','월','라인','설비유형']):
            return load_press(file_obj), 'press'
        if len(cols) >= 20: return load_robot(file_obj), 'robot'
        if len(cols) >= 15: return load_press(file_obj), 'press'
        return None, f'인식불가 (시트: {target}, 컬럼수: {len(cols)})'
    except Exception as e:
        return None, f'오류: {e}'

def merge_dfs(press_df, robot_df):
    COMMON = ['발생일시','년','월','일','라인','차종','라인_차종','설비유형','고장설비',
              '고장부위','설비_KEY','현상','원인','조치내역','조치자','소요시간',
              '정지시각','출동시각','완료시각','비고','조치유형','고장분류','파일출처',
              '고장계통코드','원인코드','조치코드','보전구분','예비품사용여부','계획정비전환대상']
    frames = []
    if press_df is not None:
        frames.append(press_df[[c for c in COMMON if c in press_df.columns]])
    if robot_df is not None:
        frames.append(robot_df[[c for c in COMMON if c in robot_df.columns]])
    if not frames: return None
    merged = pd.concat(frames, ignore_index=True)
    merged = merged[merged['발생일시'].notna()].copy()
    merged['년'] = merged['발생일시'].dt.year
    merged['월'] = merged['발생일시'].dt.month
    if '라인_차종' not in merged.columns:
        merged['라인_차종'] = merged['라인'].astype(str).str.strip()
    else:
        merged['라인_차종'] = merged['라인_차종'].fillna(merged['라인'].astype(str).str.strip())
    mask_no_car = (merged['라인_차종'].astype(str).str.endswith(' / ') |
                   merged['라인_차종'].astype(str).str.contains(r'/ $', regex=True))
    merged.loc[mask_no_car,'라인_차종'] = merged.loc[mask_no_car,'라인'].astype(str).str.strip()
    # 라인명 정규화 (입력 오류 통합: 1500t→1500T, ASS'Y→ASSY 등)
    if '라인' in merged.columns:
        merged['라인'] = merged['라인'].apply(
            lambda v: norm_line(v) if pd.notna(v) else v)
    if '라인_차종' in merged.columns:
        merged['라인_차종'] = merged['라인_차종'].apply(
            lambda v: norm_line(v) if pd.notna(v) else v)
    if '설비_KEY' in merged.columns:
        merged['설비_KEY'] = merged['라인_차종'].astype(str) + ' | ' + \
            merged['고장설비'].fillna('').astype(str).str.strip()
    merged['세부분류'] = merged.apply(get_세부분류, axis=1)
    merged['부위그룹'] = merged['고장부위'].apply(get_고장부위_그룹) if '고장부위' in merged.columns else '기타'
    # ── 재발여부 계산 (90일 기준) ──
    merged = merged.sort_values('발생일시').reset_index(drop=True)
    merged['재발여부'] = calc_recurrence(merged, 90)
    # ── 계획정비전환대상 계산 ──
    # 재발3회 이상이거나 MTTR 60분 이상이면 전환대상
    cnt_map = merged.groupby('설비_KEY').size()
    merged['_건수'] = merged['설비_KEY'].map(cnt_map)
    mttr_map = merged.groupby('설비_KEY')['소요시간'].mean()
    merged['_mttr'] = merged['설비_KEY'].map(mttr_map)
    merged['계획정비전환대상'] = merged.apply(
        lambda r: 'Y' if (r['_건수'] >= 3 and r['재발여부']) or r.get('_mttr', 0) >= 60
        else 'N', axis=1)
    merged.drop(columns=['_건수','_mttr'], inplace=True, errors='ignore')
    merged.sort_values('발생일시', inplace=True)
    merged.reset_index(drop=True, inplace=True)
    return merged

def load_from_gdrive(url):
    try:
        url = url.strip()
        file_id = None
        for p in [r'spreadsheets/d/([a-zA-Z0-9_-]+)',r'/d/([a-zA-Z0-9_-]+)',r'id=([a-zA-Z0-9_-]+)']:
            m = re.search(p, url)
            if m: file_id = m.group(1); break
        if not file_id: return None, '파일 ID를 찾을 수 없습니다.'
        if 'spreadsheets' in url or 'docs.google.com' in url:
            dl_url = 'https://docs.google.com/spreadsheets/d/' + file_id + '/export?format=xlsx'
        else:
            dl_url = 'https://drive.google.com/uc?export=download&id=' + file_id
        headers = {'User-Agent': 'Mozilla/5.0'}
        resp = requests.get(dl_url, timeout=30, headers=headers)
        if resp.status_code == 200 and b'confirm=' in resp.content[:2000]:
            cm = re.search(rb'confirm=([0-9A-Za-z_-]+)', resp.content)
            if cm:
                resp = requests.get(dl_url + '&confirm=' + cm.group(1).decode(), timeout=30, headers=headers)
        if resp.status_code != 200: return None, f'다운로드 실패 (HTTP {resp.status_code})'
        if len(resp.content) < 500: return None, '파일이 너무 작습니다 — 공유 권한을 확인해주세요.'
        return io.BytesIO(resp.content), None
    except Exception as e:
        return None, str(e)

# ══════════════════════════════════════════════════════
# 분석 함수
# ══════════════════════════════════════════════════════
def calc_mttr_mtbf(df, cluster_min=60):
    """
    근사 MTBF 계산 (방법 B):
      - BM(돌발) 건만 MTBF 계산 대상
      - 클러스터링: 완료시각~출동시각 gap <= cluster_min 분이면 동일 고장 통합
      - 근사 MTBF = (분석기간 월~토 가동시간h - 설비별 총수리시간h) ÷ 클러스터 건수
      - 0h 건(출동=완료 또는 수리시간 0) 분리 처리
    """
    results = []
    quality_issues = []  # 데이터 품질 문제 건

    df2 = df[df['소요시간'].notna()].copy()

    # 분석기간 계산 (전체 데이터 기준)
    valid_dt = df2['발생일시'].dropna()
    if valid_dt.empty:
        return pd.DataFrame(results), pd.DataFrame(quality_issues)
    period_start = valid_dt.min()
    period_end   = valid_dt.max()
    total_work_h = calc_worktime_hours(period_start, period_end, daily_hours=15.0)

    for key, grp in df2.groupby('설비_KEY'):
        grp = grp.sort_values('발생일시').copy()
        parts = key.split(' | ')
        라인 = parts[0] if len(parts) > 0 else ''
        설비 = parts[1] if len(parts) > 1 else ''
        유형 = grp['설비유형'].mode()[0] if not grp['설비유형'].isna().all() else ''

        # ── 데이터 품질 체크 ──
        zero_dur = grp[grp['소요시간'] <= 0]
        no_time  = grp[grp['출동시각'].isna() | grp['완료시각'].isna()] if '출동시각' in grp.columns else pd.DataFrame()
        if not zero_dur.empty:
            quality_issues.append({
                '설비_KEY': key, '설비유형': 유형,
                '문제유형': '수리시간 0 이하',
                '건수': len(zero_dur),
                '비고': '출동시각=완료시각 또는 입력 오류 의심'
            })
        if not no_time.empty:
            quality_issues.append({
                '설비_KEY': key, '설비유형': 유형,
                '문제유형': '출동/완료시각 없음',
                '건수': len(no_time),
                '비고': 'MTBF 계산 정확도 저하 가능'
            })

        # ── 전체 건 집계 (PM+BM 합산) ──
        total_cnt  = len(grp)
        total_stop = grp['소요시간'].sum()
        total_mttr = grp['소요시간'].mean()

        # ── BM(돌발) 건만 MTBF 계산 ──
        if '보전구분' in grp.columns:
            bm_grp = grp[grp['보전구분'] == 'BM(돌발)'].copy()
        else:
            bm_grp = grp.copy()

        mtbf_근사 = None
        cluster_cnt = None

        if len(bm_grp) >= 2:
            # 클러스터링 적용
            clustered = cluster_faults(bm_grp, cluster_min=cluster_min)
            cluster_cnt = len(clustered)
            # 설비별 총수리시간(시간)
            equip_repair_h = clustered['소요시간'].sum() / 60.0
            # 근사 MTBF = (전체가동시간 - 수리시간) / 클러스터건수
            avail_h = max(total_work_h - equip_repair_h, 0)
            if cluster_cnt > 0:
                mtbf_근사 = round(avail_h / cluster_cnt, 1)
        elif len(bm_grp) == 1:
            cluster_cnt = 1
            # 건수 1건: 분석기간 전체를 1건으로 나눔 (상한으로 표시)
            equip_repair_h = bm_grp['소요시간'].sum() / 60.0
            avail_h = max(total_work_h - equip_repair_h, 0)
            mtbf_근사 = round(avail_h, 1)  # 건수 1이므로 / 1

        results.append({
            '라인': 라인,
            '고장설비': 설비,
            '설비_KEY': key,
            '설비유형': 유형,
            '전체건수': total_cnt,
            'BM건수': len(bm_grp),
            '클러스터건수(BM)': cluster_cnt,
            'MTTR(분)': round(total_mttr, 1),
            'MTBF_근사(시간)': mtbf_근사,
            '총정지시간(분)': round(total_stop, 1),
            '분석기간_가동시간(h)': round(total_work_h, 1),
        })

    result_df = pd.DataFrame(results).sort_values('총정지시간(분)', ascending=False)
    quality_df = pd.DataFrame(quality_issues)
    return result_df, quality_df

def get_worker_df(df):
    rows = []
    for _, r in df.iterrows():
        workers, 출동유형, 인원수 = parse_workers_with_type(r.get('조치자'))
        for w in workers:
            rows.append({
                '조치자':    w,
                '소요시간':  r['소요시간'] if pd.notna(r.get('소요시간')) else 0,
                '발생일시':  r['발생일시'],
                '라인':      r.get('라인'),
                '라인_차종': r.get('라인_차종', r.get('라인', '')),
                '설비유형':  r.get('설비유형'),
                '고장설비':  r.get('고장설비'),
                '출동유형':  출동유형,   # 단독 / 협업 / 미상
                '협업인원수': 인원수,
                '조치자_원본': str(r.get('조치자') or ''),
                '설비_KEY':  r.get('설비_KEY', ''),
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame()

def calc_response_time(df):
    mask = df['정지시각'].notna() & df['출동시각'].notna()
    sub = df[mask].copy()
    if sub.empty: return None
    sub['응답시간_분'] = sub.apply(
        lambda r: (r['출동시각']-r['정지시각']).total_seconds()/60
        if r['출동시각']>r['정지시각'] else None, axis=1)
    return sub[sub['응답시간_분'].notna() & (sub['응답시간_분']>0) & (sub['응답시간_분']<240)]

def to_excel(df_dict):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as w:
        for name, df in df_dict.items():
            df.to_excel(w, sheet_name=name[:31], index=False)
    return buf.getvalue()

def kpi_card_html(label, actual, target, unit, higher_is_better=False,
                  prev_val=None, compare_label='전기 대비'):
    if higher_is_better:
        rate = min(actual/target*100, 200) if target else 0
        ok = actual >= target
    else:
        rate = min(target/actual*100, 200) if actual else 0
        ok = actual <= target
    color = '#27ae60' if ok else '#e74c3c'
    icon  = '✅' if ok else '❌'
    # 전기 대비 delta
    delta_html = ''
    if prev_val is not None and prev_val > 0:
        diff = actual - prev_val
        pct  = diff / prev_val * 100
        bad  = (diff > 0 and not higher_is_better) or (diff < 0 and higher_is_better)
        d_color = '#e74c3c' if bad else '#27ae60'
        arrow = '▲' if diff >= 0 else '▼'
        delta_html = (f'<div style="font-size:11px;color:{d_color};margin-top:3px;">'
                      f'{arrow} {compare_label} {diff:+.1f}{unit} ({pct:+.1f}%)</div>')
    return (f'<div style="background:#f8f9fa;border-radius:10px;padding:14px 16px;'
            f'border-left:5px solid {color};margin-bottom:8px;">'
            f'<div style="font-size:12px;color:#666;">{label}</div>'
            f'<div style="font-size:24px;font-weight:700;color:{color};line-height:1.2;">'
            f'{actual:.1f} <span style="font-size:13px">{unit}</span></div>'
            f'<div style="font-size:11px;color:#888;">목표: {target} {unit} &nbsp;|&nbsp; '
            f'달성률: <b>{rate:.0f}%</b> {icon}</div>'
            f'{delta_html}</div>')

# ══════════════════════════════════════════════════════
# 세션 상태
# ══════════════════════════════════════════════════════
for k in ['press_df','robot_df','merged_df']:
    if k not in st.session_state:
        st.session_state[k] = None

# 데이터 로드 완료 후 자동 rerun 플래그
if '_just_loaded' not in st.session_state:
    st.session_state['_just_loaded'] = False

# KPI 목표값 기본값 (세션 유지)
_KPI_DEFAULTS = {
    'kpi_target_cnt': 100, 'kpi_target_mttr': 30,
    'kpi_target_stop': 3000, 'kpi_target_mtbf': 200,
}
for k, v in _KPI_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# 전역 필터 초기값
_GF_DEFAULTS = {
    'gf_mode':        '📅 연도선택',
    'gf_years':       [],
    'gf_year_single': None,
    'gf_months':      [],
    'gf_start':       None,
    'gf_end':         None,
    'gf_label':       '전체',
}
for k, v in _GF_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── 자동 요약 텍스트 생성 함수 ───────────────────────
def auto_summary(df, period_label='이번 기간'):
    """선택 기간 데이터로 한국어 요약 문장 자동 생성"""
    if df is None or df.empty:
        return "데이터가 없습니다."
    lines = []
    total = len(df)
    df_v = df[df['소요시간'].notna()]
    total_stop = df_v['소요시간'].sum()
    avg_mttr   = df_v['소요시간'].mean() if not df_v.empty else 0

    lines.append(f"📋 **{period_label} 고장 총 {total:,}건**, 총 정지시간 {total_stop/60:.1f}시간, 평균 MTTR {avg_mttr:.0f}분.")

    # 설비유형별 Pareto
    if '설비유형' in df.columns:
        top_eq = df['설비유형'].value_counts().head(1)
        if not top_eq.empty:
            eq_name = top_eq.index[0]
            eq_cnt  = top_eq.iloc[0]
            eq_pct  = eq_cnt / total * 100
            lines.append(f"📌 **{eq_name}** 고장이 {eq_cnt}건({eq_pct:.0f}%)으로 가장 많습니다.")

    # 재발 현황
    if '재발여부' in df.columns:
        recur = df['재발여부'].sum()
        recur_pct = recur / total * 100 if total else 0
        if recur_pct >= 20:
            lines.append(f"⚠️ 재발 고장 {recur}건({recur_pct:.0f}%) — 근본원인 미해결 설비 점검 필요.")
        else:
            lines.append(f"✅ 재발 고장 {recur}건({recur_pct:.0f}%) — 관리 수준 양호.")

    # 야간 돌발
    if '발생일시' in df.columns and '보전구분' in df.columns:
        bm = df[df['보전구분']=='BM(돌발)']
        if not bm.empty:
            night = bm[bm['발생일시'].dt.hour.apply(lambda h: h>=22 or h<6)]
            night_pct = len(night)/len(bm)*100
            if night_pct >= 25:
                lines.append(f"🌙 야간 돌발 고장 비율 {night_pct:.0f}% — 야간 단독 작업 위험도 높음.")

    # 고장계통 TOP
    if '고장계통코드' in df.columns:
        top_sys = df['고장계통코드'].value_counts().head(1)
        if not top_sys.empty:
            lines.append(f"🔍 고장 계통 1위: **{top_sys.index[0]}** ({top_sys.iloc[0]}건).")

    return "\n\n".join(lines)

# ── 전월 대비 증감 계산 ───────────────────────────────
def calc_mom_delta(df, year, month):
    """전월 대비 당월 고장건수·정지시간 증감 반환"""
    if df is None: return None, None, None, None
    df = df[df['발생일시'].notna()].copy()
    df['년'] = df['발생일시'].dt.year
    df['월'] = df['발생일시'].dt.month
    cur  = df[(df['년']==year) & (df['월']==month)]
    # 전월 처리 (1월이면 전년 12월)
    prev_yr  = year if month > 1 else year - 1
    prev_mo  = month - 1 if month > 1 else 12
    prev = df[(df['년']==prev_yr) & (df['월']==prev_mo)]
    cur_cnt  = len(cur);   prev_cnt  = len(prev)
    cur_stop = cur['소요시간'].sum() if not cur.empty else 0
    prev_stop= prev['소요시간'].sum() if not prev.empty else 0
    return cur_cnt, prev_cnt, cur_stop, prev_stop

# ══════════════════════════════════════════════════════
# 헤더
# ══════════════════════════════════════════════════════
st.markdown('<div class="main-title">🔧 보전팀 통합 분석 시스템</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">호원오토 평택공장 | 보전관리 2팀</div>', unsafe_allow_html=True)

# ── 🚨 상단 경고 배너 (데이터 로드 후 자동 표시) ──────
_mdf = st.session_state.merged_df
if _mdf is not None:
    _warn_items = []
    # MTTR 60분 이상 설비
    if '소요시간' in _mdf.columns and '설비_KEY' in _mdf.columns:
        _mttr_bad = (_mdf.groupby('설비_KEY')['소요시간'].mean()
                     .reset_index(name='평균MTTR'))
        _mttr_cnt = (_mttr_bad['평균MTTR'] >= 60).sum()
        if _mttr_cnt > 0:
            _warn_items.append(f"🔴 MTTR 60분 이상 설비 **{_mttr_cnt}개**")
    # 재발률 50% 이상 설비
    if '재발여부' in _mdf.columns:
        _eq_r = (_mdf.groupby('설비_KEY')
                 .agg(전체=('재발여부','count'), 재발=('재발여부','sum'))
                 .reset_index())
        _eq_r = _eq_r[_eq_r['전체'] >= 3]
        if not _eq_r.empty:
            _eq_r['재발률'] = _eq_r['재발'] / _eq_r['전체'] * 100
            _danger_cnt = (_eq_r['재발률'] >= 50).sum()
            if _danger_cnt > 0:
                _warn_items.append(f"🟠 재발률 50% 이상 설비 **{_danger_cnt}개**")
    # BM 비율 경고
    if '보전구분' in _mdf.columns:
        _bm_pct = (_mdf['보전구분'] == 'BM(돌발)').mean() * 100
        if _bm_pct >= 80:
            _warn_items.append(f"⚠️ BM(돌발) 비율 **{_bm_pct:.0f}%** — PM 전환 검토 필요")
    if _warn_items:
        _banner = " &nbsp;|&nbsp; ".join(_warn_items)
        st.markdown(
            f'<div style="background:#fff3cd;border-left:5px solid #ffc107;'
            f'padding:10px 18px;border-radius:6px;margin-bottom:10px;font-size:14px;">'
            f'⚠️ <b>즉시 확인 필요</b> &nbsp;→&nbsp; {_banner}</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            '<div style="background:#d4edda;border-left:5px solid #28a745;'
            'padding:8px 18px;border-radius:6px;margin-bottom:10px;font-size:13px;">'
            '✅ 긴급 경고 없음 — 정상 관리 중</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# 전역 기간 필터 — 모든 탭 공통 적용
# ══════════════════════════════════════════════════════
def apply_global_filter(df):
    if df is None: return None
    mode = st.session_state.get('gf_mode', '📅 연도선택')
    if mode == '📅 연도선택':
        yrs = st.session_state.get('gf_years', [])
        return df[df['년'].isin(yrs)].copy() if yrs else df.copy()
    elif mode == '📆 월선택':
        yr  = st.session_state.get('gf_year_single')
        mos = st.session_state.get('gf_months', [])
        if yr and mos:
            return df[(df['년'] == yr) & (df['월'].isin(mos))].copy()
        return df.copy()
    else:
        s = st.session_state.get('gf_start')
        e = st.session_state.get('gf_end')
        if s and e:
            return df[
                (df['발생일시'].dt.date >= s) &
                (df['발생일시'].dt.date <= e)
            ].copy()
        return df.copy()

_mdf_raw = st.session_state.merged_df

# ── 통합 완료 직후 성공 메시지 표시 (rerun 후 1회) ──────
if st.session_state.get('_just_loaded') and _mdf_raw is not None:
    _yr_min = int(_mdf_raw['년'].min())
    _yr_max = int(_mdf_raw['년'].max())
    _재발수 = int(_mdf_raw['재발여부'].sum()) if '재발여부' in _mdf_raw.columns else 0
    st.success(
        f"✅ 데이터 통합 완료 — {len(_mdf_raw):,}건 "
        f"({_yr_min}~{_yr_max}년) | 재발 판정: {_재발수:,}건 "
        f"| 전역 기간 필터가 활성화되었습니다.")
    st.session_state['_just_loaded'] = False  # 플래그 초기화 (1회만 표시)

# ── 전역 필터 초기값 세팅 (데이터 있을 때) ─────────────
if _mdf_raw is not None:
    _all_yrs = sorted(_mdf_raw['년'].dropna().unique().astype(int))
    _dt_min  = _mdf_raw['발생일시'].dropna().min().date()
    _dt_max  = _mdf_raw['발생일시'].dropna().max().date()
    if not st.session_state['gf_years']:
        st.session_state['gf_years'] = _all_yrs
    if st.session_state['gf_year_single'] is None:
        st.session_state['gf_year_single'] = _all_yrs[-1] if _all_yrs else None
    if not st.session_state['gf_months']:
        st.session_state['gf_months'] = list(range(1, 13))
    if st.session_state['gf_start'] is None:
        st.session_state['gf_start'] = _dt_min
    if st.session_state['gf_end'] is None:
        st.session_state['gf_end'] = _dt_max

with st.expander(
    "🗓️ 전역 기간 필터 — 모든 탭에 공통 적용",
    expanded=(_mdf_raw is not None)
):
    if _mdf_raw is None:
        st.info("📂 **TAB1 [데이터 불러오기]** 에서 파일을 업로드하고 "
                "**[🔄 데이터 통합 실행]** 버튼을 누르면 "
                "기간 필터가 자동으로 활성화됩니다.")
    else:
        # ── 한 줄 3컬럼: [기간모드] | [선택위젯] | [현재필터 설명] ──
        _gc1, _gc2, _gc3 = st.columns([2, 4, 3])

        with _gc1:
            _gf_mode = st.radio(
                "기간 모드",
                ["📅 연도선택", "📆 월선택", "🗓️ 날짜범위"],
                index=["📅 연도선택","📆 월선택","🗓️ 날짜범위"].index(
                    st.session_state['gf_mode']),
                horizontal=True,
                key='_gf_mode_radio',
            )
            st.session_state['gf_mode'] = _gf_mode

        with _gc2:
            if _gf_mode == "📅 연도선택":
                _sel_yrs = st.multiselect(
                    "연도 선택", _all_yrs,
                    default=st.session_state['gf_years'],
                    key='_gf_years_ms',
                )
                st.session_state['gf_years'] = _sel_yrs
                st.session_state['gf_label'] = (
                    '+'.join(str(y) for y in sorted(_sel_yrs))+'년'
                    if _sel_yrs else '전체')

            elif _gf_mode == "📆 월선택":
                _mc1, _mc2 = st.columns(2)
                with _mc1:
                    _sel_yr_s = st.selectbox(
                        "연도", _all_yrs,
                        index=_all_yrs.index(st.session_state['gf_year_single'])
                              if st.session_state['gf_year_single'] in _all_yrs
                              else len(_all_yrs)-1,
                        key='_gf_yr_single',
                    )
                    st.session_state['gf_year_single'] = _sel_yr_s
                with _mc2:
                    _avail_mo = sorted(
                        _mdf_raw[_mdf_raw['년']==_sel_yr_s]['월']
                        .dropna().unique().astype(int))
                    _cur_mos = [m for m in st.session_state['gf_months']
                                if m in _avail_mo] or _avail_mo
                    _sel_mos = st.multiselect(
                        "월 선택", _avail_mo,
                        default=_cur_mos,
                        key='_gf_months_ms',
                    )
                    st.session_state['gf_months'] = _sel_mos
                _mos_str = ','.join(str(m)+'월' for m in sorted(_sel_mos)) if _sel_mos else '전체'
                st.session_state['gf_label'] = f"{_sel_yr_s}년 {_mos_str}"

            else:  # 날짜범위
                _dc1, _dc2 = st.columns(2)
                with _dc1:
                    _sel_start = st.date_input(
                        "시작일", value=st.session_state['gf_start'],
                        min_value=_dt_min, max_value=_dt_max,
                        key='_gf_start_di',
                    )
                    st.session_state['gf_start'] = _sel_start
                with _dc2:
                    _sel_end = st.date_input(
                        "종료일", value=st.session_state['gf_end'],
                        min_value=_dt_min, max_value=_dt_max,
                        key='_gf_end_di',
                    )
                    st.session_state['gf_end'] = _sel_end
                st.session_state['gf_label'] = f"{_sel_start} ~ {_sel_end}"

        with _gc3:
            _gf_df  = apply_global_filter(_mdf_raw)
            _gf_cnt = len(_gf_df) if _gf_df is not None else 0
            _gf_pct = _gf_cnt / len(_mdf_raw) * 100 if len(_mdf_raw) > 0 else 0
            # 컬럼 상단 여백 맞추기용 빈 레이블
            st.markdown("&nbsp;", unsafe_allow_html=True)
            st.markdown(
                f"📌 **{st.session_state['gf_label']}**  \n"
                f"적용: **{_gf_cnt:,}건** "
                f"({_gf_pct:.1f}% / 전체 {len(_mdf_raw):,}건)")

(tab1, tab2, tab3, tab4, tab5, tab6, tab7,
 tab8, tab9, tab10, tab11, tab12, tab13, tab14, tab15) = st.tabs([
    "📂 데이터 불러오기",
    "📊 고장현황 (Pareto)",
    "⚙️ 설비분석 (MTTR/MTBF)",
    "👷 인원분석",
    "🏆 설비 위험도",
    "⏱️ 유실시간 분석",
    "🔧 예방정비 추천",
    "📈 월별 트렌드",
    "🎯 KPI 목표관리",
    "🔄 BM/PM 분석",
    "🔁 재발 고장 전용",
    "🏷️ 표준코드 분석",
    "📝 월보·주보 자동작성",
    "📥 출력",
    "🔄 POP양식 변환",
])

# ══════════════════════════════════════════════════════
# TAB 1 — 데이터 불러오기
# ══════════════════════════════════════════════════════
with tab1:
    st.subheader("데이터 불러오기")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("#### ① 구글 드라이브 공유 링크")
        gdrive_url = st.text_input("링크 붙여넣기", placeholder="https://drive.google.com/file/d/...")
        if st.button("🔗 드라이브에서 불러오기", use_container_width=True):
            if gdrive_url:
                with st.spinner("다운로드 중..."):
                    fobj, err = load_from_gdrive(gdrive_url)
                    if err:
                        st.error(f"오류: {err}")
                    else:
                        result, ftype = detect_and_load(fobj)
                        if ftype == 'press':
                            st.session_state.press_df = result
                            st.success(f"✅ 프레스 파일 로드 — {len(result):,}건")
                        elif ftype == 'robot':
                            st.session_state.robot_df = result
                            st.success(f"✅ 로봇/지그 파일 로드 — {len(result):,}건")
                        else:
                            st.error(f"파일 형식 인식 실패: {ftype}")
            else:
                st.warning("링크를 입력해주세요.")
    with c2:
        st.markdown("#### ② 직접 파일 업로드 (.xlsx / .csv)")
        uploaded = st.file_uploader("파일 선택 (여러 파일 동시 가능)",
                                    type=['xlsx','csv'], accept_multiple_files=True)
        if uploaded:
            for uf in uploaded:
                with st.spinner(f"{uf.name} 처리 중..."):
                    try:
                        if uf.name.endswith('.csv'):
                            st.info(f"CSV '{uf.name}' — xlsx 권장")
                        else:
                            result, ftype = detect_and_load(uf, uf.name)
                            if ftype == 'press':
                                st.session_state.press_df = result
                                st.success(f"✅ 프레스 '{uf.name}' — {len(result):,}건")
                            elif ftype == 'robot':
                                st.session_state.robot_df = result
                                st.success(f"✅ 로봇/지그 '{uf.name}' — {len(result):,}건")
                            else:
                                st.warning(f"'{uf.name}' 형식 인식 실패: {ftype}")
                    except Exception as e:
                        st.error(f"'{uf.name}' 처리 오류: {e}")

    st.divider()

    # ── 데이터 품질 진단 (로드 직후) ──
    with st.expander("🩺 데이터 품질 진단 (로드된 파일 기준)", expanded=False):
        for label, sdf in [('프레스', st.session_state.press_df),
                           ('로봇/지그', st.session_state.robot_df)]:
            if sdf is None: continue
            st.markdown(f"**{label} 파일**")
            total = len(sdf)
            no_dt   = sdf['발생일시'].isna().sum() if '발생일시' in sdf.columns else 0
            no_time = sdf['소요시간'].isna().sum() if '소요시간' in sdf.columns else 0
            no_equip= sdf['설비유형'].isna().sum() if '설비유형' in sdf.columns else 0
            dup     = sdf.duplicated(subset=['발생일시','설비_KEY','소요시간'], keep=False).sum() if '설비_KEY' in sdf.columns else 0
            d1,d2,d3,d4 = st.columns(4)
            d1.metric("전체 행수", f"{total:,}")
            d2.metric("발생일시 누락", f"{no_dt}", delta=f"-{no_dt/total*100:.1f}%" if total else "")
            d3.metric("소요시간 누락", f"{no_time}", delta=f"-{no_time/total*100:.1f}%" if total else "")
            d4.metric("중복 의심행", f"{dup}")
            if no_equip > 0:
                st.warning(f"설비유형 미분류: {no_equip}건 — 원본 확인 필요")

    if st.button("🔄 데이터 통합 실행", type="primary", use_container_width=True):
        merged = merge_dfs(st.session_state.press_df, st.session_state.robot_df)
        if merged is not None:
            st.session_state.merged_df    = merged
            # 전역 필터 초기값 강제 리셋 (새 데이터 기준으로 재설정)
            st.session_state['gf_years']       = []
            st.session_state['gf_year_single'] = None
            st.session_state['gf_months']      = []
            st.session_state['gf_start']       = None
            st.session_state['gf_end']         = None
            st.session_state['gf_label']       = '전체'
            st.session_state['_just_loaded']   = True
            st.rerun()   # 즉시 재실행 → 전역 필터 활성화
        else:
            st.warning("불러온 데이터가 없습니다.")

    s1, s2, s3 = st.columns(3)
    with s1:
        if st.session_state.press_df is not None:
            st.markdown(f'<div class="ok-box">✅ 프레스 파일<br>{len(st.session_state.press_df):,}건 로드됨</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="warn-box">⬜ 프레스 파일 미로드</div>', unsafe_allow_html=True)
    with s2:
        if st.session_state.robot_df is not None:
            st.markdown(f'<div class="ok-box">✅ 로봇/지그 파일<br>{len(st.session_state.robot_df):,}건 로드됨</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="warn-box">⬜ 로봇/지그 파일 미로드</div>', unsafe_allow_html=True)
    with s3:
        if st.session_state.merged_df is not None:
            df_m = st.session_state.merged_df
            st.markdown(f'<div class="ok-box">✅ 통합 완료<br>{len(df_m):,}건 ({int(df_m["년"].min())}~{int(df_m["년"].max())}년)</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="warn-box">⬜ 통합 미실행</div>', unsafe_allow_html=True)

    if st.session_state.merged_df is not None:
        with st.expander("📋 통합 데이터 미리보기 (상위 200건)"):
            preview_cols = ['발생일시','라인','설비_KEY','설비유형','고장설비','고장부위',
                            '현상','원인','조치내역','조치자','소요시간','고장분류',
                            '고장계통코드','원인코드','조치코드','보전구분','재발여부','파일출처']
            st.dataframe(st.session_state.merged_df[
                [c for c in preview_cols if c in st.session_state.merged_df.columns]
            ].head(200), use_container_width=True)


# ══════════════════════════════════════════════════════
# TAB 2 — 고장현황 (Pareto)
# ══════════════════════════════════════════════════════
with tab2:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.subheader("고장현황 분석")
        fc1,fc2,fc3,fc4,fc5 = st.columns([2,2,2,2,1])
        with fc1:
            equips = ['전체'] + sorted(df['설비유형'].dropna().unique().tolist(), key=str)
            sel_eq = st.selectbox("설비유형", equips, key='t2e')
        with fc2:
            lines = ['전체'] + sorted(df['라인'].dropna().unique().tolist(), key=str)
            sel_ln = st.selectbox("라인", lines, key='t2l')
        with fc3:
            cars = ['전체'] + sorted(df['차종'].dropna().unique().tolist(), key=str) if '차종' in df.columns else ['전체']
            sel_car = st.selectbox("차종", cars, key='t2car')
        with fc4:
            top_n = st.slider("Top N", 5, 30, 20, key='t2n')

        fdf = apply_global_filter(df)
        if fdf is None or fdf.empty:
            st.warning(f"선택한 기간({st.session_state.get('gf_label','')})에 데이터가 없습니다.")
        if sel_eq  != '전체': fdf = fdf[fdf['설비유형'] == sel_eq]
        if sel_ln  != '전체': fdf = fdf[fdf['라인'] == sel_ln]
        if sel_car != '전체' and '차종' in fdf.columns: fdf = fdf[fdf['차종'] == sel_car]

        total_cnt  = len(fdf)
        total_stop = fdf['소요시간'].sum()
        avg_mttr   = fdf['소요시간'].mean()
        resp_df    = calc_response_time(fdf)
        avg_resp   = resp_df['응답시간_분'].mean() if resp_df is not None and not resp_df.empty else 0

        k1,k2,k3,k4 = st.columns(4)
        for col,label,val,unit in [
            (k1,'총 고장건수',f'{total_cnt:,}','건'),
            (k2,'총 정지시간',f'{total_stop/60:.0f}','시간'),
            (k3,'평균 MTTR',f'{avg_mttr:.0f}','분'),
            (k4,'평균 응답시간',f'{avg_resp:.0f}','분')]:
            col.markdown(f'<div class="kpi-box"><div class="kpi-val">{val}'
                         f'<span class="kpi-unit"> {unit}</span></div>'
                         f'<div class="kpi-label">{label}</div></div>', unsafe_allow_html=True)
        st.divider()

        # ── hover helper: 그룹 기준 ─────────────────────
        def _tf_grp(group_col, group_val, src_df):
            sub = src_df[src_df[group_col] == group_val]
            cols = [c for c in ['고장부위','현상'] if c in sub.columns]
            if not cols: return '-'
            if len(cols) == 2:
                fk = sub['고장부위'].fillna('').str.strip() + ' / ' + sub['현상'].fillna('').str.strip()
                fk = fk[(sub['고장부위'].notna() & (sub['고장부위'].str.strip()!='')) |
                        (sub['현상'].notna() & (sub['현상'].str.strip()!=''))]
                fk = fk.str.strip(' /').str.strip()
            else:
                fk = sub[cols[0]].dropna()
                fk = fk[fk.str.strip() != '']
            if fk.empty: return '-'
            top = fk.value_counts().head(3)
            return '<br>'.join([f'  {i+1}. {str(k)[:30]} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

        def _te_grp(group_col, group_val, src_df):
            sub = src_df[src_df[group_col] == group_val]
            if '고장설비' not in sub.columns: return '-'
            col = sub['고장설비'].dropna()
            col = col[col.str.strip() != '']
            if col.empty: return '-'
            top = col.value_counts().head(3)
            return '<br>'.join([f'  {i+1}. {k} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

        p1,p2 = st.columns(2)
        with p1:
            st.markdown("##### 설비유형별 Pareto — 건수")
            grp = (fdf.groupby('설비유형')
                   .agg(건수=('소요시간','count'),총정지시간=('소요시간','sum'))
                   .reset_index().sort_values('건수',ascending=False).head(top_n))
            grp['누적%'] = (grp['건수'].cumsum()/grp['건수'].sum()*100).round(1)
            grp['평균MTTR'] = (grp['총정지시간']/grp['건수']).round(1)
            grp['_tf'] = grp['설비유형'].apply(lambda v: _tf_grp('설비유형', v, fdf))
            grp['_te'] = grp['설비유형'].apply(lambda v: _te_grp('설비유형', v, fdf))
            fig = make_subplots(specs=[[{"secondary_y":True}]])
            fig.add_trace(go.Bar(x=grp['설비유형'],y=grp['건수'],name='건수',
                                 marker_color='#1e3a5f',
                                 customdata=np.stack([grp['총정지시간'],grp['평균MTTR'],grp['누적%'],grp['_tf'],grp['_te']],axis=-1),
                                 hovertemplate='<b>%{x}</b><br>건수: %{y:,}건<br>총정지: %{customdata[0]:.0f}분<br>평균MTTR: %{customdata[1]:.1f}분<br>누적: %{customdata[2]:.1f}%<extra></extra>'),secondary_y=False)
            fig.add_trace(go.Scatter(x=grp['설비유형'],y=grp['누적%'],name='누적%',
                                     line=dict(color='#e74c3c',width=2),mode='lines+markers',
                                     hovertemplate='누적: %{y:.1f}%<extra></extra>'),secondary_y=True)
            fig.add_hline(y=80,line_dash='dash',line_color='orange',annotation_text='80%',secondary_y=True)
            fig.update_yaxes(title_text='건수',secondary_y=False)
            fig.update_yaxes(title_text='누적 %',range=[0,105],secondary_y=True)
            fig.update_layout(height=320,margin=dict(t=10,b=40,l=10,r=60))
            st.plotly_chart(fig,use_container_width=True)
        with p2:
            st.markdown("##### 설비유형별 Pareto — 정지시간")
            grp2 = (fdf.groupby('설비유형')
                    .agg(건수=('소요시간','count'),총정지시간=('소요시간','sum'))
                    .reset_index().sort_values('총정지시간',ascending=False).head(top_n))
            grp2['누적%'] = (grp2['총정지시간'].cumsum()/grp2['총정지시간'].sum()*100).round(1)
            grp2['평균MTTR'] = (grp2['총정지시간']/grp2['건수']).round(1)
            grp2['_tf'] = grp2['설비유형'].apply(lambda v: _tf_grp('설비유형', v, fdf))
            grp2['_te'] = grp2['설비유형'].apply(lambda v: _te_grp('설비유형', v, fdf))
            fig2 = make_subplots(specs=[[{"secondary_y":True}]])
            fig2.add_trace(go.Bar(x=grp2['설비유형'],y=grp2['총정지시간'],name='정지시간(분)',
                                  marker_color='#c0392b',
                                  customdata=np.stack([grp2['건수'],grp2['평균MTTR'],grp2['누적%'],grp2['_tf'],grp2['_te']],axis=-1),
                                  hovertemplate='<b>%{x}</b><br>정지: %{y:,.0f}분<br>건수: %{customdata[0]}건<br>평균MTTR: %{customdata[1]:.1f}분<br>누적: %{customdata[2]:.1f}%<extra></extra>'),secondary_y=False)
            fig2.add_trace(go.Scatter(x=grp2['설비유형'],y=grp2['누적%'],name='누적%',
                                      line=dict(color='#e67e22',width=2),mode='lines+markers'),secondary_y=True)
            fig2.add_hline(y=80,line_dash='dash',line_color='orange',annotation_text='80%',secondary_y=True)
            fig2.update_yaxes(title_text='정지시간(분)',secondary_y=False)
            fig2.update_yaxes(title_text='누적 %',range=[0,105],secondary_y=True)
            fig2.update_layout(height=320,margin=dict(t=10,b=40,l=10,r=60))
            st.plotly_chart(fig2,use_container_width=True)

        st.markdown("##### 라인별 고장건수 Top N")
        line_grp = (fdf.groupby('라인_차종')
                    .agg(건수=('소요시간','count'),총정지시간=('소요시간','sum'))
                    .reset_index().sort_values('건수',ascending=False).head(top_n))
        line_grp['평균MTTR'] = (line_grp['총정지시간']/line_grp['건수']).round(1)

        def _top_faults_str(라인명):
            sub = fdf[fdf['라인_차종'] == 라인명]
            if '고장부위' in sub.columns and '현상' in sub.columns:
                fault_key = sub['고장부위'].fillna('') + ' / ' + sub['현상'].fillna('')
            elif '고장부위' in sub.columns:
                fault_key = sub['고장부위'].fillna('기타')
            else:
                fault_key = sub['현상'].fillna('기타')
            top = fault_key.value_counts().head(3)
            items = [f"  {i+1}. {str(k)[:30]} ({v}건)" for i, (k, v) in enumerate(top.items())]
            return '<br>'.join(items) if items else '-'

        def _top_equip_str(라인명):
            sub = fdf[fdf['라인_차종'] == 라인명]
            if '고장설비' not in sub.columns:
                return '-'
            top = sub['고장설비'].fillna('미상').value_counts().head(3)
            return '<br>'.join(
                [f"  {i+1}. {k} ({v}건)" for i, (k, v) in enumerate(top.items())]) or '-'

        line_grp['_top_faults'] = line_grp['라인_차종'].apply(_top_faults_str)
        line_grp['_top_equip']  = line_grp['라인_차종'].apply(_top_equip_str)
        if '재발여부' in fdf.columns:
            recur_map = fdf.groupby('라인_차종')['재발여부'].sum().astype(int)
            line_grp['_recur'] = line_grp['라인_차종'].map(recur_map).fillna(0).astype(int)
        else:
            line_grp['_recur'] = 0

        fig_ln = px.bar(line_grp, x='건수', y='라인_차종', orientation='h',
                        color='총정지시간', color_continuous_scale='Blues',
                        custom_data=['총정지시간', '평균MTTR', '_recur', '_top_equip', '_top_faults'])
        fig_ln.update_traces(
            hovertemplate=(
                '<b>%{y}</b><br>'
                '건수: %{x}건 &nbsp;|&nbsp; 총정지: %{customdata[0]:,.0f}분 &nbsp;|&nbsp; 평균MTTR: %{customdata[1]:.1f}분<br>'
                '재발: %{customdata[2]}건<br>'
                '<b>▶ 반복 고장 Top 3 (부위/현상)</b><br>%{customdata[4]}<br>'
                '<b>▶ 고장 다발 설비 Top 3</b><br>%{customdata[3]}'
                '<extra></extra>'
            )
        )
        fig_ln.update_layout(height=max(300, len(line_grp)*26),
                              margin=dict(t=10, b=20, l=10, r=60), yaxis_title='')
        st.plotly_chart(fig_ln, use_container_width=True)

        with st.expander("📋 고장현황 상세 데이터"):
            disp_cols = ['발생일시','라인','설비유형','고장설비','고장부위',
                         '현상','원인','소요시간','고장계통코드','원인코드','조치코드','보전구분','재발여부']
            st.dataframe(fdf[[c for c in disp_cols if c in fdf.columns]],use_container_width=True)


# ══════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════
# TAB 3 — 설비분석 (MTTR/MTBF)
# ══════════════════════════════════════════════════════
with tab3:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.subheader("설비 MTTR / MTBF 분석")
        st.caption("⚠️ MTBF는 근사값입니다 — 월~토 일15시간 가동 적용")

        # 전역 기간 필터 적용
        mf1, mf2, mf3 = st.columns([3, 2, 1])

        mdf = apply_global_filter(df)
        if mdf is None: mdf = pd.DataFrame()

        # ── 필터 행 2: 설비유형 + TopN + 클러스터링 ───────
        ff1, ff2, ff3 = st.columns([2, 1, 2])
        with ff1:
            eq_m = ['전체'] + sorted(df['설비유형'].dropna().unique().tolist(), key=str)
            sel_eq_m = st.selectbox("설비유형", eq_m, key='t3e')
        with ff2:
            top_m = st.slider("Top N", 10, 50, 20, key='t3n')
        with ff3:
            cluster_min_val = st.slider(
                "클러스터링 임계값(분) — N분 이내 재발 시 동일 고장 통합",
                min_value=30, max_value=240, value=60, step=10, key='t3_cluster')

        if sel_eq_m != '전체':
            mdf = mdf[mdf['설비유형'] == sel_eq_m]

        if mdf.empty:
            st.warning("선택한 기간/조건에 해당하는 데이터가 없습니다.")
        else:
            mttr_df, quality_df = calc_mttr_mtbf(mdf, cluster_min=cluster_min_val)

            if mttr_df.empty:
                st.warning("분석 가능한 데이터가 없습니다.")
            else:
                # ── KPI 지표 4개 ──────────────────────────
                mk1, mk2, mk3, mk4 = st.columns(4)
                mk1.metric("분석 설비수", f"{len(mttr_df):,}개")
                mk2.metric("평균 MTTR", f"{mttr_df['MTTR(분)'].mean():.1f}분",
                           help="출동시각~완료시각 기준 평균")
                valid_mtbf = mttr_df['MTBF_근사(시간)'].dropna()
                mk3.metric("평균 MTBF(근사) ⚠️",
                           f"{valid_mtbf.mean():.1f}시간" if not valid_mtbf.empty else "N/A",
                           help="월~토 일 15h 가동 근사값. 실가동시간 미반영으로 과대평가 가능")
                mk4.metric("총 정지시간",
                           f"{mttr_df['총정지시간(분)'].sum():,.0f}분",
                           help="출동시각~완료시각 합산")
                st.divider()

                # ── hover helper: 설비_KEY 기준 ────────────────────
                def _tf_key(key, src_df):
                    sub = src_df[src_df['설비_KEY'] == key] if '설비_KEY' in src_df.columns else src_df.iloc[0:0]
                    cols = [c for c in ['고장부위','현상'] if c in sub.columns]
                    if not cols: return '-'
                    if len(cols) == 2:
                        fk = sub['고장부위'].fillna('').str.strip() + ' / ' + sub['현상'].fillna('').str.strip()
                        fk = fk[(sub['고장부위'].notna() & (sub['고장부위'].str.strip()!='')) |
                                (sub['현상'].notna() & (sub['현상'].str.strip()!=''))]
                        fk = fk.str.strip(' /').str.strip()
                    else:
                        fk = sub[cols[0]].dropna()
                        fk = fk[fk.str.strip() != '']
                    if fk.empty: return '-'
                    top = fk.value_counts().head(3)
                    return '<br>'.join([f'  {i+1}. {str(k)[:30]} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

                def _te_key(key, src_df):
                    sub = src_df[src_df['설비_KEY'] == key] if '설비_KEY' in src_df.columns else src_df.iloc[0:0]
                    if '조치자' not in sub.columns: return '-'
                    col = sub['조치자'].dropna()
                    col = col[col.astype(str).str.strip() != '']
                    if col.empty: return '-'
                    top = col.value_counts().head(3)
                    return '<br>'.join([f'  {i+1}. {str(k)} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

                # ── MTTR 상위 차트 ────────────────────────
                m_top = mttr_df.head(top_m).copy()
                m_top['_tf'] = m_top['설비_KEY'].apply(lambda k: _tf_key(k, mdf))
                m_top['_te'] = m_top['설비_KEY'].apply(lambda k: _te_key(k, mdf))
                st.markdown(f"##### MTTR 상위 {top_m}개 (평균수리시간 긴 순)")
                fig_mttr = px.bar(
                    m_top.sort_values('MTTR(분)'),
                    x='MTTR(분)', y='설비_KEY', orientation='h', color='라인',
                    custom_data=['라인', '고장설비', '전체건수', 'BM건수',
                                 'MTBF_근사(시간)', '총정지시간(분)', '설비유형', '_tf', '_te'])
                fig_mttr.update_traces(
                    texttemplate='%{x:.0f}분', textposition='outside',
                    hovertemplate=(
                        '<b>%{y}</b><br>유형: %{customdata[6]}<br>'
                        'MTTR: %{x:.1f}분<br>MTBF(근사): %{customdata[4]}시간<br>'
                        '전체: %{customdata[2]}건 / BM: %{customdata[3]}건<br>'
                        '총정지: %{customdata[5]:.0f}분<br>'
                        '<b>▶ 반복 고장 Top 3 (부위/현상)</b><br>%{customdata[7]}<br>'
                        '<b>▶ 주요 조치자 Top 3</b><br>%{customdata[8]}'
                        '<extra></extra>'))
                fig_mttr.update_layout(
                    height=max(420, len(m_top)*28),
                    margin=dict(t=30, b=20), yaxis_title='')
                st.plotly_chart(fig_mttr, use_container_width=True)

                # ── MTBF 하위 차트 (색상 구분) ────────────
                mtbf_data = mttr_df[mttr_df['MTBF_근사(시간)'].notna()].nsmallest(
                    top_m, 'MTBF_근사(시간)')
                if not mtbf_data.empty:
                    st.markdown("##### MTBF(근사) 하위 — 짧을수록 잦은 고장")
                    st.caption(
                        "🔴 1시간 미만: 데이터 확인 필요 | "
                        "🟠 1~10시간: 위험 | 🟢 10시간 초과: 양호")

                    def _mtbf_color(h):
                        if h is None: return '데이터없음'
                        if h < 1:  return '🔴 데이터확인필요'
                        if h < 10: return '🟠 위험'
                        return '🟢 양호'

                    mtbf_data = mtbf_data.copy()
                    mtbf_data['MTBF등급'] = mtbf_data['MTBF_근사(시간)'].apply(_mtbf_color)
                    mtbf_data['_tf'] = mtbf_data['설비_KEY'].apply(lambda k: _tf_key(k, mdf))
                    mtbf_data['_te'] = mtbf_data['설비_KEY'].apply(lambda k: _te_key(k, mdf))
                    color_map_mtbf = {
                        '🔴 데이터확인필요': '#e74c3c',
                        '🟠 위험':           '#e67e22',
                        '🟢 양호':           '#27ae60',
                        '데이터없음':         '#aaaaaa',
                    }
                    fig_mtbf = px.bar(
                        mtbf_data, x='MTBF_근사(시간)', y='설비_KEY',
                        orientation='h', color='MTBF등급',
                        color_discrete_map=color_map_mtbf,
                        custom_data=['라인', '고장설비', 'BM건수',
                                     '클러스터건수(BM)', 'MTTR(분)', '설비유형', 'MTBF등급', '_tf', '_te'])
                    fig_mtbf.update_traces(
                        texttemplate='%{x:.1f}h', textposition='outside',
                        hovertemplate=(
                            '<b>%{y}</b><br>MTBF(근사): %{x:.1f}시간<br>'
                            '등급: %{customdata[6]}<br>'
                            'BM건수: %{customdata[2]}건 → 클러스터: %{customdata[3]}건<br>'
                            'MTTR: %{customdata[4]:.1f}분<br>'
                                        '<b>▶ 반복 고장 Top 3 (부위/현상)</b><br>%{customdata[7]}<br>'
                                        '<b>▶ 주요 조치자 Top 3</b><br>%{customdata[8]}'
                                        '<extra></extra>'))
                    fig_mtbf.update_layout(
                        height=max(420, len(mtbf_data)*28),
                        margin=dict(t=30, b=20), yaxis_title='',
                        legend=dict(orientation='h',y=1.0,x=0,yanchor='bottom',xanchor='left'))
                    st.plotly_chart(fig_mtbf, use_container_width=True)

                    # 1시간 미만 설비 경고 배지
                    danger_equip = mtbf_data[mtbf_data['MTBF_근사(시간)'] < 1]
                    if not danger_equip.empty:
                        names = ', '.join(danger_equip['설비_KEY'].tolist()[:5])
                        st.markdown(
                            f'<div class="err-box">🔴 <b>MTBF 1시간 미만 설비 '
                            f'{len(danger_equip)}개</b> — 데이터 품질 확인 또는 '
                            f'클러스터링 임계값 조정 권장<br>'
                            f'해당 설비: {names}{"..." if len(danger_equip)>5 else ""}'
                            f'</div>', unsafe_allow_html=True)

                # ── MTTR vs MTBF 산점도 ───────────────────
                st.markdown("##### MTTR vs MTBF(근사) 산점도")
                scatter_d = mttr_df[mttr_df['MTBF_근사(시간)'].notna()].copy()
                if not scatter_d.empty:
                    scatter_d['_tf'] = scatter_d['설비_KEY'].apply(lambda k: _tf_key(k, mdf))
                    scatter_d['_te'] = scatter_d['설비_KEY'].apply(lambda k: _te_key(k, mdf))
                    fig_sc = px.scatter(
                        scatter_d, x='MTBF_근사(시간)', y='MTTR(분)',
                        color='라인', size='전체건수', hover_name='설비_KEY',
                        custom_data=['라인', '고장설비', '전체건수',
                                     '총정지시간(분)', '설비유형', '_tf', '_te'],
                        labels={
                            'MTBF_근사(시간)': 'MTBF 근사(시간, 길수록 안전)',
                            'MTTR(분)': 'MTTR(분, 낮을수록 좋음)'})
                    mtbf_med = scatter_d['MTBF_근사(시간)'].median()
                    mttr_med = scatter_d['MTTR(분)'].median()
                    fig_sc.add_vline(x=mtbf_med, line_dash='dash', line_color='gray',
                                     annotation_text=f'MTBF 중앙값 {mtbf_med:.0f}h')
                    fig_sc.add_hline(y=mttr_med, line_dash='dash', line_color='gray',
                                     annotation_text=f'MTTR 중앙값 {mttr_med:.0f}분')
                    fig_sc.update_traces(
                        hovertemplate=(
                            '<b>%{hovertext}</b><br>'
                            'MTBF: %{x:.1f}시간 | MTTR: %{y:.1f}분<br>'
                            '전체: %{customdata[2]}건 | 총정지: %{customdata[3]:.0f}분<br>'
                            '<b>▶ 반복 고장 Top 3 (부위/현상)</b><br>%{customdata[5]}<br>'
                            '<b>▶ 주요 조치자 Top 3</b><br>%{customdata[6]}'
                            '<extra></extra>'
                        )
                    )
                    fig_sc.update_layout(height=460, margin=dict(t=30, b=20))
                    st.plotly_chart(fig_sc, use_container_width=True)

                # ── 응답시간 분포 ─────────────────────────
                resp = calc_response_time(mdf)
                if resp is not None and not resp.empty:
                    st.markdown("##### ⏱ 응답시간 분포 (정지→출동)")
                    rc1, rc2 = st.columns(2)
                    with rc1:
                        fig_rh = px.histogram(resp, x='응답시간_분', nbins=30,
                                              color_discrete_sequence=['#1e3a5f'])
                        fig_rh.add_vline(
                            x=resp['응답시간_분'].mean(), line_dash='dash',
                            line_color='red',
                            annotation_text=f"평균 {resp['응답시간_분'].mean():.0f}분")
                        fig_rh.update_layout(height=300, margin=dict(t=20, b=20))
                        st.plotly_chart(fig_rh, use_container_width=True)
                    with rc2:
                        fig_rb = px.box(resp, x='설비유형', y='응답시간_분',
                                        color='설비유형')
                        fig_rb.update_layout(height=300, margin=dict(t=20, b=20),
                                             showlegend=False)
                        st.plotly_chart(fig_rb, use_container_width=True)

                # ── 전체 테이블 ───────────────────────────
                with st.expander("📋 전체 MTTR/MTBF 테이블"):
                    show_cols = ['설비_KEY', '설비유형', '라인', '전체건수',
                                 'BM건수', '클러스터건수(BM)', 'MTTR(분)',
                                 'MTBF_근사(시간)', '총정지시간(분)',
                                 '분석기간_가동시간(h)']
                    show_cols = [c for c in show_cols if c in mttr_df.columns]
                    st.dataframe(mttr_df[show_cols], use_container_width=True)

                # ── 데이터 품질 문제 목록 ─────────────────
                with st.expander("⚠️ 데이터 보완 필요 목록 (MTBF 계산 정확도 영향)"):
                    if quality_df.empty:
                        st.success("데이터 품질 문제 없음")
                    else:
                        # 요약
                        q_summary = quality_df.groupby('문제유형')['건수'].sum().reset_index()
                        for _, qrow in q_summary.iterrows():
                            st.markdown(
                                f'<div class="warn-box">⚠️ <b>{qrow["문제유형"]}</b>: '
                                f'총 {qrow["건수"]}건</div>',
                                unsafe_allow_html=True)
                        st.dataframe(quality_df, use_container_width=True)
                        # 다운로드
                        q_excel = to_excel({'데이터품질이슈': quality_df})
                        st.download_button(
                            "📥 품질이슈 Excel 다운로드", data=q_excel,
                            file_name="데이터품질이슈.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════
# TAB 4 — 인원분석
# ══════════════════════════════════════════════════════
with tab4:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.subheader("인원 업무부하 분석")
        wdf_base = apply_global_filter(df)
        worker_df = get_worker_df(wdf_base)

        if worker_df.empty:
            st.warning("조치자 데이터가 없습니다.")
        else:
            # ── 기본 집계 ─────────────────────────────
            person_agg = (worker_df.groupby('조치자')
                          .agg(출동건수=('소요시간','count'),
                               총소요시간_분=('소요시간','sum'))
                          .reset_index())
            person_agg['평균소요시간_분'] = (
                person_agg['총소요시간_분'] / person_agg['출동건수']).round(1)
            person_agg['총소요시간_시'] = (person_agg['총소요시간_분'] / 60).round(1)

            # ── 단독/협업 분리 집계 ──────────────────
            solo_agg = (worker_df[worker_df['출동유형']=='단독']
                        .groupby('조치자').size().reset_index(name='단독콜'))
            coop_agg = (worker_df[worker_df['출동유형']=='협업']
                        .groupby('조치자').size().reset_index(name='협업콜'))
            person_agg = person_agg.merge(solo_agg, on='조치자', how='left')
            person_agg = person_agg.merge(coop_agg, on='조치자', how='left')
            person_agg['단독콜'] = person_agg['단독콜'].fillna(0).astype(int)
            person_agg['협업콜'] = person_agg['협업콜'].fillna(0).astype(int)
            person_agg['협업비율(%)'] = (
                person_agg['협업콜'] / person_agg['출동건수'] * 100).round(1)
            person_agg = person_agg[person_agg['출동건수'] >= 1].sort_values(
                '출동건수', ascending=False)

            # ── KPI 요약 ──────────────────────────────
            ka1, ka2, ka3, ka4 = st.columns(4)
            ka1.metric("분석 인원", f"{len(person_agg)}명")
            ka2.metric("단독 출동", f"{person_agg['단독콜'].sum():,}건")
            ka3.metric("협업 출동", f"{person_agg['협업콜'].sum():,}건")
            ka4.metric("평균 협업비율",
                       f"{person_agg['협업비율(%)'].mean():.1f}%")
            st.divider()

            # ── 스택바: 단독/협업 분리 차트 ───────────
            st.markdown("##### 인원별 출동건수 (단독 / 협업 구분)")
            st.caption("협업 비율 높음 → 난이도 높은 고장 전문가 또는 추가 교육 필요 대상")
            top20 = person_agg.head(20).copy()

            # ── hover helper: 조치자 기준 ──────────────────────────
            def _tf_worker(worker, src_df):
                sub = src_df[src_df['조치자'] == worker] if '조치자' in src_df.columns else src_df.iloc[0:0]
                cols = [c for c in ['고장부위','현상'] if c in sub.columns]
                if not cols: return '-'
                if len(cols) == 2:
                    fk = sub['고장부위'].fillna('').str.strip() + ' / ' + sub['현상'].fillna('').str.strip()
                    fk = fk[(sub['고장부위'].notna() & (sub['고장부위'].str.strip()!='')) |
                            (sub['현상'].notna() & (sub['현상'].str.strip()!=''))]
                    fk = fk.str.strip(' /').str.strip()
                else:
                    fk = sub[cols[0]].dropna()
                    fk = fk[fk.str.strip() != '']
                if fk.empty: return '-'
                top = fk.value_counts().head(3)
                return '<br>'.join([f'  {i+1}. {str(k)[:30]} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

            def _te_worker(worker, src_df):
                sub = src_df[src_df['조치자'] == worker] if '조치자' in src_df.columns else src_df.iloc[0:0]
                if '설비_KEY' not in sub.columns: return '-'
                col = sub['설비_KEY'].dropna()
                col = col[col.astype(str).str.strip() != '']
                if col.empty: return '-'
                top = col.value_counts().head(3)
                return '<br>'.join([f'  {i+1}. {k} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

            top20['_tf'] = top20['조치자'].apply(lambda w: _tf_worker(w, worker_df))
            top20['_te'] = top20['조치자'].apply(lambda w: _te_worker(w, worker_df))

            fig_stack = go.Figure()
            fig_stack.add_trace(go.Bar(
                name='단독 출동',
                x=top20['조치자'], y=top20['단독콜'],
                marker_color='#1e3a5f',
                text=top20['단독콜'],
                textposition='inside',
                customdata=top20[['단독콜','_tf','_te']].values,
                hovertemplate='<b>%{x}</b><br>단독콜: %{y}건<br><b>▶ 주요 고장유형 Top 3</b><br>%{customdata[1]}<br><b>▶ 담당 설비 Top 3</b><br>%{customdata[2]}<extra></extra>'))
            fig_stack.add_trace(go.Bar(
                name='협업 출동',
                x=top20['조치자'], y=top20['협업콜'],
                marker_color='#e67e22',
                text=top20['협업콜'],
                textposition='inside',
                customdata=top20[['협업콜','_tf','_te']].values,
                hovertemplate='<b>%{x}</b><br>협업콜: %{y}건<br><b>▶ 주요 고장유형 Top 3</b><br>%{customdata[1]}<br><b>▶ 담당 설비 Top 3</b><br>%{customdata[2]}<extra></extra>'))
            fig_stack.update_layout(
                barmode='stack', height=420,
                margin=dict(t=30, b=60),
                xaxis_tickangle=-30,
                legend=dict(orientation='h', y=1.05))
            st.plotly_chart(fig_stack, use_container_width=True)

            # ── 협업비율 상위 경고 ────────────────────
            high_coop = person_agg[person_agg['협업비율(%)'] >= 70].head(5)
            if not high_coop.empty:
                names = ', '.join(
                    f"{r['조치자']}({r['협업비율(%)']}%)"
                    for _, r in high_coop.iterrows())
                st.markdown(
                    f'<div class="warn-box">⚠️ 협업 비율 70% 이상: <b>{names}</b>'
                    f' — 해당 인원이 담당하는 설비 난이도 또는 인력 배치 검토 필요'
                    f'</div>', unsafe_allow_html=True)

            pa1, pa2 = st.columns(2)
            with pa1:
                st.markdown("##### 인원별 총 소요시간")
                fig_p2 = px.bar(
                    top20, x='조치자', y='총소요시간_시',
                    color='총소요시간_시', color_continuous_scale='Reds',
                    custom_data=['출동건수', '평균소요시간_분', '협업비율(%)', '_tf', '_te'])
                fig_p2.update_traces(
                    texttemplate='%{y:.0f}h', textposition='outside',
                    hovertemplate=(
                        '<b>%{x}</b><br>총소요: %{y:.1f}시간<br>'
                        '출동: %{customdata[0]}건<br>'
                        '협업비율: %{customdata[2]:.1f}%<br>'
                        '<b>▶ 주요 고장유형 Top 3</b><br>%{customdata[3]}<br>'
                        '<b>▶ 담당 설비 Top 3</b><br>%{customdata[4]}'
                        '<extra></extra>'))
                fig_p2.update_layout(height=400, margin=dict(t=20, b=60),
                                     showlegend=False, xaxis_tickangle=-30)
                st.plotly_chart(fig_p2, use_container_width=True)

            with pa2:
                st.markdown("##### 협업 비율 순위")
                coop_top = person_agg[person_agg['출동건수'] >= 3].sort_values(
                    '협업비율(%)', ascending=False).head(20)
                coop_top['_tf'] = coop_top['조치자'].apply(lambda w: _tf_worker(w, worker_df))
                coop_top['_te'] = coop_top['조치자'].apply(lambda w: _te_worker(w, worker_df))
                fig_coop = px.bar(
                    coop_top, x='조치자', y='협업비율(%)',
                    color='협업비율(%)', color_continuous_scale='Oranges',
                    custom_data=['출동건수', '단독콜', '협업콜', '_tf', '_te'])
                fig_coop.add_hline(y=50, line_dash='dash', line_color='red',
                                   annotation_text='50% 기준선')
                fig_coop.update_traces(
                    texttemplate='%{y:.0f}%', textposition='outside',
                    hovertemplate=(
                        '<b>%{x}</b><br>협업비율: %{y:.1f}%<br>'
                        '단독: %{customdata[1]}건 / 협업: %{customdata[2]}건<br>'
                        '<b>▶ 주요 고장유형 Top 3</b><br>%{customdata[3]}<br>'
                        '<b>▶ 담당 설비 Top 3</b><br>%{customdata[4]}'
                        '<extra></extra>'))
                fig_coop.update_layout(height=400, margin=dict(t=20, b=60),
                                       showlegend=False, xaxis_tickangle=-30)
                st.plotly_chart(fig_coop, use_container_width=True)

            # ── 협업 필요 설비 Top N ───────────────────
            st.markdown("##### 협업 출동이 많은 설비 Top 15")
            st.caption("항상 2인 이상 출동하는 설비 → 난이도 분류, 안전 기준 수립 참고")
            if '설비_KEY' in worker_df.columns:
                coop_equip = (worker_df[worker_df['출동유형'] == '협업']
                              .groupby('설비_KEY')
                              .agg(협업출동건수=('출동유형','count'),
                                   평균인원수=('협업인원수','mean'))
                              .reset_index()
                              .sort_values('협업출동건수', ascending=False)
                              .head(15))
                if not coop_equip.empty:
                    coop_equip['평균인원수'] = coop_equip['평균인원수'].round(1)
                    fig_ce = px.bar(
                        coop_equip, x='협업출동건수', y='설비_KEY',
                        orientation='h', color='평균인원수',
                        color_continuous_scale='YlOrRd',
                        custom_data=['평균인원수'])
                    fig_ce.update_traces(
                        texttemplate='%{x}건', textposition='outside',
                        hovertemplate=(
                            '<b>%{y}</b><br>협업출동: %{x}건<br>'
                            '평균인원: %{customdata[0]:.1f}명<extra></extra>'))
                    fig_ce.update_layout(
                        height=max(380, len(coop_equip)*28),
                        margin=dict(t=20, b=20), yaxis_title='')
                    st.plotly_chart(fig_ce, use_container_width=True)

            # ── 협업 시 MTTR vs 단독 MTTR 비교 ──────
            st.markdown("##### 단독 vs 협업 출동 시 평균 MTTR 비교")
            st.caption("협업해도 MTTR이 길면 → 부품·기술 문제 가능성")
            mttr_comp = (worker_df[worker_df['출동유형'].isin(['단독','협업'])]
                         .groupby('출동유형')['소요시간']
                         .mean().reset_index())
            mttr_comp.columns = ['출동유형', '평균MTTR(분)']
            mttr_comp['평균MTTR(분)'] = mttr_comp['평균MTTR(분)'].round(1)
            fig_mttr_comp = px.bar(
                mttr_comp, x='출동유형', y='평균MTTR(분)',
                color='출동유형',
                color_discrete_map={'단독': '#1e3a5f', '협업': '#e67e22'},
                text='평균MTTR(분)')
            fig_mttr_comp.update_traces(textposition='outside')
            fig_mttr_comp.update_layout(height=320, showlegend=False,
                                        margin=dict(t=20, b=20))
            st.plotly_chart(fig_mttr_comp, use_container_width=True)

            # ── 주별 업무시간 ──────────────────────────
            st.markdown("##### ⚠️ 주별 업무시간 (보전 소요시간 기준)")
            wdf2 = worker_df[worker_df['발생일시'].notna()].copy()
            wdf2['연도'] = wdf2['발생일시'].dt.year
            wdf2['주차'] = wdf2['발생일시'].dt.isocalendar().week.astype(int)
            weekly = wdf2.groupby(['조치자','연도','주차'])['소요시간'].sum().reset_index()
            weekly['소요시간_시'] = (weekly['소요시간'] / 60).round(1)
            weekly['초과위험'] = weekly['소요시간_시'] > 20
            over_workers = weekly[weekly['초과위험']]['조치자'].unique()
            if len(over_workers) > 0:
                st.markdown(
                    f'<div class="warn-box">⚠️ 주간 보전업무 20시간 초과: '
                    f'<b>{", ".join(over_workers[:10])}</b> ({len(over_workers)}명)</div>',
                    unsafe_allow_html=True)
            top_workers = person_agg['조치자'].tolist()
            sel_w = st.selectbox("인원 선택 (주별 차트)", top_workers, key='t4w')
            w_data = weekly[weekly['조치자'] == sel_w].copy()
            w_data['년주'] = (w_data['연도'].astype(str) + '-W' +
                              w_data['주차'].astype(str).str.zfill(2))
            fig_wk = px.bar(w_data, x='년주', y='소요시간_시',
                            color='초과위험',
                            color_discrete_map={True: '#dc3545', False: '#1e3a5f'})
            fig_wk.add_hline(y=20, line_dash='dash', line_color='orange',
                              annotation_text='경고 20h/주')
            fig_wk.update_layout(
                height=340, margin=dict(t=30, b=50),
                xaxis_tickangle=-45, showlegend=False,
                title=f'{sel_w} — 주별 보전업무 소요시간')
            st.plotly_chart(fig_wk, use_container_width=True)

            # ── 전체 인원 콜수 히트맵 ─────────────────────
            st.divider()
            st.markdown("##### 👷 전체 인원 콜수 히트맵 (요일 × 시간대)")
            st.caption("전체 조치자 출동 건수 합산 기준 — 색이 진할수록 출동 집중 시간대")
            _wdf_all = worker_df[worker_df['발생일시'].notna()].copy()
            if not _wdf_all.empty:
                _wdf_all['요일_en'] = _wdf_all['발생일시'].dt.day_name()
                _wdf_all['시간']    = _wdf_all['발생일시'].dt.hour
                _heat_all = (_wdf_all.groupby(['요일_en','시간'])
                             .size().reset_index(name='콜수'))
                _day_kr4    = {'Monday':'월','Tuesday':'화','Wednesday':'수',
                               'Thursday':'목','Friday':'금','Saturday':'토','Sunday':'일'}
                _day_order4 = ['Monday','Tuesday','Wednesday','Thursday',
                               'Friday','Saturday','Sunday']
                _heat_all['요일'] = pd.Categorical(
                    _heat_all['요일_en'].map(_day_kr4),
                    categories=[_day_kr4[d] for d in _day_order4], ordered=True)
                _pivot_all4 = (_heat_all.pivot(index='요일', columns='시간', values='콜수')
                               .reindex(columns=list(range(24)), fill_value=0)
                               .fillna(0).astype(int))
                _peak_hour4 = int(_pivot_all4.sum().idxmax())
                _peak_day4  = _pivot_all4.sum(axis=1).idxmax()
                _peak_val4  = int(_pivot_all4.values.max())
                st.caption(
                    f"🔴 최대 집중: **{_peak_day4}요일 {_peak_hour4}시** "
                    f"| 최고 콜수: **{_peak_val4}건**")
                _fig_all4 = px.imshow(
                    _pivot_all4,
                    color_continuous_scale='YlOrRd',
                    text_auto=True,
                    labels=dict(x='시간(시)', y='요일', color='콜수'),
                    aspect='auto')
                _fig_all4.update_layout(
                    height=320, margin=dict(t=20, b=20),
                    xaxis=dict(tickmode='linear', tick0=0, dtick=1))
                st.plotly_chart(_fig_all4, use_container_width=True)
                # 주간/야간 KPI
                _h1, _h2, _h3 = st.columns(3)
                _day_tot4   = int(_pivot_all4.loc[:, 6:17].values.sum())
                _night_tot4 = int(_pivot_all4.loc[:, [*range(0,6), *range(18,24)]].values.sum())
                _total4     = int(_pivot_all4.values.sum())
                _h1.metric("전체 콜수", f"{_total4:,}건")
                _h2.metric("주간 콜수 (06~17시)", f"{_day_tot4:,}건",
                           f"{_day_tot4/_total4*100:.1f}%" if _total4 else "0%")
                _h3.metric("야간 콜수 (18~05시)", f"{_night_tot4:,}건",
                           f"{_night_tot4/_total4*100:.1f}%" if _total4 else "0%")
            else:
                st.info("조치자 데이터가 없어 히트맵을 표시할 수 없습니다.")

            # ── 시간대별 출동 히트맵 (선택 인원 기준) ──────
            st.markdown(f"##### 시간대별 출동 히트맵 — {sel_w}")
            st.caption(f"선택 인원 **{sel_w}** 의 출동 시간대 분포")
            wdf3 = worker_df[
                (worker_df['발생일시'].notna()) &
                (worker_df['조치자'] == sel_w)
            ].copy()
            if wdf3.empty:
                st.info(f"{sel_w} 의 출동 데이터가 없습니다.")
            else:
                wdf3['요일_en'] = wdf3['발생일시'].dt.day_name()
                wdf3['시간'] = wdf3['발생일시'].dt.hour
                heat = wdf3.groupby(['요일_en','시간']).size().reset_index(name='건수')
                day_kr = {'Monday':'월','Tuesday':'화','Wednesday':'수','Thursday':'목',
                          'Friday':'금','Saturday':'토','Sunday':'일'}
                day_order = ['Monday','Tuesday','Wednesday','Thursday',
                             'Friday','Saturday','Sunday']
                heat['요일'] = pd.Categorical(
                    heat['요일_en'].map(day_kr),
                    categories=[day_kr[d] for d in day_order], ordered=True)
                pivot_h = (heat.pivot(index='요일', columns='시간', values='건수')
                           .reindex(columns=list(range(24)), fill_value=0)
                           .fillna(0).astype(int))
                fig_heat = px.imshow(
                    pivot_h, color_continuous_scale='YlOrRd',
                    text_auto=True,
                    labels=dict(x='시간(시)', y='요일', color='건수'),
                    aspect='auto')
                fig_heat.update_layout(
                    height=320, margin=dict(t=40, b=20),
                    xaxis=dict(tickmode='linear', tick0=0, dtick=1),
                    title=f'{sel_w} — 시간대별 출동 분포')
                st.plotly_chart(fig_heat, use_container_width=True)

            # ── 개인별 콜수 집계 테이블 ─────────────────
            st.divider()
            st.markdown(f"##### 📋 개인별 콜수 집계 — {sel_w}")
            st.caption(f"선택 인원 **{sel_w}** 의 설비별/요일별 출동 상세")
            if not wdf3.empty:
                # 시간대 구분 추가
                _wdf3_detail = wdf3.copy()
                _wdf3_detail['시간대'] = _wdf3_detail['발생일시'].dt.hour.apply(
                    lambda h: '주간(06~17시)' if 6<=h<=17 else '야간(18~05시)')
                _wdf3_detail['요일'] = _wdf3_detail['발생일시'].dt.day_name().map(
                    {'Monday':'월','Tuesday':'화','Wednesday':'수','Thursday':'목',
                     'Friday':'금','Saturday':'토','Sunday':'일'})
                # 요일별 콜수
                _day_call = (_wdf3_detail.groupby('요일')['소요시간']
                             .agg(콜수='count', 총소요시간_분='sum')
                             .reset_index())
                _day_call['평균소요_분'] = (_day_call['총소요시간_분']/_day_call['콜수']).round(1)
                _day_order_kr = ['월','화','수','목','금','토','일']
                _day_call['요일'] = pd.Categorical(
                    _day_call['요일'], categories=_day_order_kr, ordered=True)
                _day_call = _day_call.sort_values('요일')
                # 주간/야간 콜수
                _shift_call = (_wdf3_detail.groupby('시간대')['소요시간']
                               .agg(콜수='count', 총소요시간_분='sum').reset_index())
                _dc1, _dc2 = st.columns(2)
                with _dc1:
                    st.markdown("**요일별 콜수**")
                    st.dataframe(_day_call[['요일','콜수','총소요시간_분','평균소요_분']]
                                 .rename(columns={'총소요시간_분':'총소요(분)','평균소요_분':'평균소요(분)'}),
                                 use_container_width=True, hide_index=True)
                with _dc2:
                    st.markdown("**주간/야간 콜수**")
                    st.dataframe(_shift_call[['시간대','콜수','총소요시간_분']]
                                 .rename(columns={'총소요시간_분':'총소요(분)'}),
                                 use_container_width=True, hide_index=True)
                # 설비별 콜수 Top10
                st.markdown("**설비별 콜수 Top 10**")
                _equip_call = (_wdf3_detail.groupby('설비_KEY')['소요시간']
                               .agg(콜수='count', 총소요시간_분='sum')
                               .reset_index().sort_values('콜수', ascending=False).head(10))
                _equip_call['평균소요_분'] = (_equip_call['총소요시간_분']/_equip_call['콜수']).round(1)
                st.dataframe(_equip_call[['설비_KEY','콜수','총소요시간_분','평균소요_분']]
                             .rename(columns={'총소요시간_분':'총소요(분)','평균소요_분':'평균소요(분)'}),
                             use_container_width=True, hide_index=True)

            # ── 상세 집계 테이블 ───────────────────────
            with st.expander("📋 인원별 상세 집계 (단독/협업 포함)"):
                show_person = person_agg[[
                    '조치자','출동건수','단독콜','협업콜','협업비율(%)',
                    '총소요시간_시','평균소요시간_분'
                ]].copy()
                st.dataframe(show_person, use_container_width=True)


# ══════════════════════════════════════════════════════
# TAB 5 — 설비 위험도
# ══════════════════════════════════════════════════════
with tab5:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 🏆 설비 위험도 Ranking")
        st.caption("위험도 = 고장빈도 × 가중치 + 총정지시간 × 가중치 + 평균MTTR × 가중치")
        rf1,rf2,rf3 = st.columns([3,3,2])
        with rf1:
            eq_r = ['전체']+sorted(df['설비유형'].dropna().unique().tolist(),key=str)
            sel_eq_r = st.selectbox("설비유형",eq_r,key='r_eq')
        with rf3:
            top_n_r = st.slider("Top N 표시",10,50,20,key='r_top')
        rw1,rw2,rw3 = st.columns(3)
        with rw1: w_freq = st.slider("빈도 가중치(%)",0,100,40,5,key='r_wf')
        with rw2: w_time = st.slider("정지시간 가중치(%)",0,100,40,5,key='r_wt')
        with rw3: w_mttr = st.slider("MTTR 가중치(%)",0,100,20,5,key='r_wm')

        rdf = apply_global_filter(df)
        if sel_eq_r != '전체': rdf = rdf[rdf['설비유형']==sel_eq_r]
        risk = (rdf.groupby(['라인_차종','고장설비','설비유형'])
                .agg(건수=('소요시간','count'),총정지시간=('소요시간','sum'),평균MTTR=('소요시간','mean'))
                .reset_index())
        risk = risk[risk['건수']>=2].copy()
        risk['총정지시간'] = risk['총정지시간'].fillna(0)
        risk['평균MTTR'] = risk['평균MTTR'].fillna(0)
        mx_f = risk['건수'].max() or 1
        mx_t = risk['총정지시간'].max() or 1
        mx_m = risk['평균MTTR'].max() or 1
        risk['위험도점수'] = (risk['건수']/mx_f*w_freq + risk['총정지시간']/mx_t*w_time + risk['평균MTTR']/mx_m*w_mttr).round(1)
        p80 = risk['위험도점수'].quantile(0.80)
        p60 = risk['위험도점수'].quantile(0.60)
        risk['등급'] = risk['위험도점수'].apply(lambda s: '🔴 위험' if s>=p80 else ('🟠 주의' if s>=p60 else '🟢 양호'))
        risk = risk.sort_values('위험도점수',ascending=False).reset_index(drop=True)
        risk.index += 1
        risk_top = risk.head(top_n_r).copy()

        k1,k2,k3,k4 = st.columns(4)
        k1.metric("분석 설비 수",f"{len(risk):,}개")
        k2.metric("🔴 위험",f"{(risk['등급']=='🔴 위험').sum()}개")
        k3.metric("🟠 주의",f"{(risk['등급']=='🟠 주의').sum()}개")
        k4.metric("총 정지시간",f"{risk['총정지시간'].sum():,.0f}분")
        st.divider()

        risk_top['설비KEY'] = risk_top['라인_차종'].astype(str) + ' | ' + risk_top['고장설비'].astype(str)

        def _tf_risk(row, src_df):
            sub = src_df[(src_df['라인_차종']==row['라인_차종']) & (src_df['고장설비']==row['고장설비'])]
            cols = [c for c in ['고장부위','현상'] if c in sub.columns]
            if not cols: return '-'
            if len(cols) == 2:
                fk = sub['고장부위'].fillna('').str.strip() + ' / ' + sub['현상'].fillna('').str.strip()
                fk = fk[(sub['고장부위'].notna() & (sub['고장부위'].str.strip()!='')) |
                        (sub['현상'].notna() & (sub['현상'].str.strip()!=''))]
                fk = fk.str.strip(' /').str.strip()
            else:
                fk = sub[cols[0]].dropna()
                fk = fk[fk.str.strip() != '']
            if fk.empty: return '-'
            top = fk.value_counts().head(3)
            return '<br>'.join([f'  {i+1}. {str(k)[:30]} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

        def _te_risk(row, src_df):
            sub = src_df[(src_df['라인_차종']==row['라인_차종']) & (src_df['고장설비']==row['고장설비'])]
            if '조치자' not in sub.columns: return '-'
            col = sub['조치자'].dropna()
            col = col[col.astype(str).str.strip() != '']
            if col.empty: return '-'
            top = col.value_counts().head(3)
            return '<br>'.join([f'  {i+1}. {str(k)} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

        risk_top = risk_top.copy()
        risk_top['_tf'] = risk_top.apply(lambda r: _tf_risk(r, rdf), axis=1)
        risk_top['_te'] = risk_top.apply(lambda r: _te_risk(r, rdf), axis=1)
        color_map = {'🔴 위험':'#e74c3c','🟠 주의':'#e67e22','🟢 양호':'#27ae60'}
        fig_risk = px.bar(risk_top.sort_values('위험도점수'),x='위험도점수',y='설비KEY',orientation='h',
                          color='등급',color_discrete_map=color_map,
                          custom_data=['건수','총정지시간','평균MTTR','설비유형','_tf','_te'])
        fig_risk.update_traces(
            hovertemplate=('<b>%{y}</b><br>위험도: %{x:.1f}점<br>설비유형: %{customdata[3]}<br>'
                           '건수: %{customdata[0]}건<br>총정지: %{customdata[1]:,.0f}분<br>'
            '<b>▶ 반복 고장 Top 3 (부위/현상)</b><br>%{customdata[4]}<br>'
            '<b>▶ 주요 조치자 Top 3</b><br>%{customdata[5]}'
            '<extra></extra>'))
        fig_risk.update_layout(height=max(480,top_n_r*26),margin=dict(t=50,b=20,l=10,r=20),
                                legend=dict(orientation='h',y=1.0,x=0,yanchor='bottom',xanchor='left'))
        st.plotly_chart(fig_risk,use_container_width=True)

        with st.expander("📋 위험도 전체 순위표"):
            show_risk = risk[['등급','위험도점수','라인_차종','고장설비','설비유형','건수','총정지시간','평균MTTR']].head(100).copy()
            st.dataframe(show_risk,use_container_width=True)


# ══════════════════════════════════════════════════════
# TAB 6 — 유실시간 분석
# ══════════════════════════════════════════════════════
with tab6:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## ⏱️ 유실시간 분석")
        lf1,lf2,lf3 = st.columns(3)
        with lf2:
            eq_l = ['전체']+sorted(df['설비유형'].dropna().unique().tolist(),key=str)
            sel_eq_l = st.selectbox("설비유형",eq_l,key='l_eq')
        with lf3:
            lc_l = ['전체']+sorted(df['라인_차종'].dropna().unique().tolist(),key=str)
            sel_lc_l = st.selectbox("차종·라인",lc_l,key='l_lc')

        ldf = apply_global_filter(df)
        if sel_eq_l != '전체': ldf = ldf[ldf['설비유형']==sel_eq_l]
        if sel_lc_l != '전체': ldf = ldf[ldf['라인_차종']==sel_lc_l]
        ldf_v = ldf[ldf['소요시간'].notna()].copy()
        if len(ldf_v)==0:
            st.warning("필터 조건에 해당하는 소요시간 데이터가 없습니다.")
        else:
            총유실 = ldf_v['소요시간'].sum()
            lk1,lk2,lk3,lk4 = st.columns(4)
            lk1.metric("총 유실시간",f"{총유실:,.0f}분",f"({총유실/60:.1f}h)")
            lk2.metric("총 고장건수",f"{len(ldf_v):,}건")
            lk3.metric("건당 평균 유실",f"{ldf_v['소요시간'].mean():.1f}분")
            lk4.metric("단건 최대 유실",f"{ldf_v['소요시간'].max():.0f}분")
            st.divider()

            ldf_v['년월'] = ldf_v['발생일시'].dt.to_period('M').astype(str)
            monthly_eq = ldf_v.groupby(['년월','설비유형'])['소요시간'].sum().reset_index(name='유실시간')
            monthly_cnt = ldf_v.groupby('년월').size().reset_index(name='건수').sort_values('년월')

            fig_tr = make_subplots(specs=[[{"secondary_y":True}]])
            colors = px.colors.qualitative.Set2
            for i,eq in enumerate(ldf_v['설비유형'].dropna().unique()):
                sub = monthly_eq[monthly_eq['설비유형']==eq]
                fig_tr.add_trace(go.Bar(x=sub['년월'],y=sub['유실시간'],name=eq,
                                        marker_color=colors[i%len(colors)],
                                        hovertemplate='%{x}<br>'+eq+': %{y:,.0f}분<extra></extra>'),secondary_y=False)
            fig_tr.add_trace(go.Scatter(x=monthly_cnt['년월'],y=monthly_cnt['건수'],name='건수',
                                         line=dict(color='#1e3a5f',width=2),mode='lines+markers',
                                         hovertemplate='%{x}<br>건수: %{y}건<extra></extra>'),secondary_y=True)
            fig_tr.update_layout(barmode='stack',height=380,margin=dict(t=20,b=50,l=10,r=60),
                                  xaxis_tickangle=-45,legend=dict(orientation='h',y=1.08))
            fig_tr.update_yaxes(title_text='유실시간(분)',secondary_y=False)
            fig_tr.update_yaxes(title_text='건수',secondary_y=True)
            st.plotly_chart(fig_tr,use_container_width=True)


# ══════════════════════════════════════════════════════
# TAB 7 — 예방정비 추천
# ══════════════════════════════════════════════════════
with tab7:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 🔧 예방정비 추천")
        pf1,pf2,pf3 = st.columns([2,2,1])
        with pf2:
            eq_p = ['전체']+sorted(df['설비유형'].dropna().unique().tolist(),key=str)
            sel_eq_p = st.selectbox("설비유형",eq_p,key='t7eq')
        with pf3:
            min_cnt_p = st.number_input("최소 고장건수",min_value=1,value=2,key='t7mn')

        pdf = apply_global_filter(df)
        if sel_eq_p != '전체': pdf = pdf[pdf['설비유형']==sel_eq_p]
        pdf_v = pdf[pdf['소요시간'].notna() & pdf['발생일시'].notna()].copy()
        if pdf_v.empty:
            st.warning("데이터가 없습니다.")
        else:
            pm = (pdf_v.groupby(['라인_차종','고장설비','설비유형','고장부위'])
                  .agg(건수=('소요시간','count'),총정지시간=('소요시간','sum'),
                       평균MTTR=('소요시간','mean'),
                       최초발생=('발생일시','min'),최근발생=('발생일시','max'))
                  .reset_index())
            pm = pm[pm['건수']>=min_cnt_p].copy()
            pm['평균MTTR'] = pm['평균MTTR'].round(1)
            pm['기간_일'] = (pm['최근발생']-pm['최초발생']).dt.days
            pm['고장주기_일'] = (pm['기간_일']/(pm['건수']-1)).round(1).where(pm['건수']>1)

            mx_c = pm['건수'].max() or 1
            mx_t = pm['총정지시간'].max() or 1
            mx_m = pm['평균MTTR'].max() or 1
            pm['우선순위점수'] = (pm['건수']/mx_c*40+pm['총정지시간']/mx_t*40+pm['평균MTTR']/mx_m*20).round(1)
            p80 = pm['우선순위점수'].quantile(0.80)
            p60 = pm['우선순위점수'].quantile(0.60)
            pm['우선순위'] = pm['우선순위점수'].apply(lambda s: '🔴 즉시조치' if s>=p80 else ('🟠 조기예방' if s>=p60 else '🟢 정기점검'))

            def get_action(row):
                acts = []
                cyc = row['고장주기_일']
                if pd.notna(cyc):
                    if cyc<30: acts.append(f"⚡ 점검주기 단축 (평균 {cyc:.0f}일)")
                    elif cyc<90: acts.append(f"📅 월간 집중점검 ({cyc:.0f}일 주기)")
                    else: acts.append(f"📋 분기점검 유지 ({cyc:.0f}일 주기)")
                if row['평균MTTR']>=60: acts.append(f"🔩 예비부품 필수확보 (수리 {row['평균MTTR']:.0f}분)")
                elif row['평균MTTR']>=30: acts.append(f"🔩 핵심부품 재고점검 ({row['평균MTTR']:.0f}분)")
                if row['건수']>=10: acts.append(f"📊 근본원인 분석 필요 ({row['건수']}회 반복)")
                return ' | '.join(acts) if acts else '✅ 현행 정기점검 유지'
            pm['추천조치'] = pm.apply(get_action, axis=1)
            pm = pm.sort_values('우선순위점수',ascending=False).reset_index(drop=True)
            pm.index += 1

            pk1,pk2,pk3,pk4 = st.columns(4)
            pk1.metric("추천 대상 설비",f"{len(pm)}개")
            pk2.metric("🔴 즉시조치",f"{(pm['우선순위']=='🔴 즉시조치').sum()}개")
            pk3.metric("🟠 조기예방",f"{(pm['우선순위']=='🟠 조기예방').sum()}개")
            pk4.metric("평균 고장주기",f"{pm['고장주기_일'].mean():.0f}일")
            st.divider()

            # ── 현상/원인 Top3 조회 helper ─────────────────────
            def _fault_detail_html(row, src_df):
                mask = ((src_df['라인_차종']==row['라인_차종']) &
                        (src_df['고장설비']==row['고장설비']))
                if '고장부위' in src_df.columns:
                    mask = mask & (src_df['고장부위']==row['고장부위'])
                sub = src_df[mask]
                def _top3(col):
                    if col not in sub.columns: return '데이터 없음'
                    c = sub[col].dropna(); c = c[c.astype(str).str.strip()!='']
                    if c.empty: return '데이터 없음'
                    top = c.value_counts().head(3)
                    return '<br>'.join([f'&nbsp;&nbsp;{i+1}. {str(k)[:35]} ({v}건)' for i,(k,v) in enumerate(top.items())])
                ph = _top3('현상'); ca = _top3('원인')
                return (f'<b>🔍 고장 현상 Top 3</b><br>{ph}<br>'
                        f'<b>🔍 주요 원인 Top 3</b><br>{ca}')

            urgent = pm[pm['우선순위']=='🔴 즉시조치'].head(10)
            if len(urgent):
                st.markdown("##### 🔴 즉시조치 대상")
                for idx,row in urgent.iterrows():
                    detail_html = _fault_detail_html(row, pdf_v)
                    st.markdown(
                        f'<div class="card-red pm-card">'
                        f'<b>#{idx} {row["라인_차종"]} | {row["고장설비"]}</b>'
                        f' <span style="color:#888;font-size:12px">{row["설비유형"]}</span>'
                        f' &nbsp;<span style="color:#2471a3;font-size:11px">(마우스 올리면 상세보기)</span><br>'
                        f'고장 <b>{row["건수"]}건</b> / 총정지 <b>{row["총정지시간"]:.0f}분</b> / '
                        f'평균MTTR <b>{row["평균MTTR"]}분</b> / 고장주기 <b>{row["고장주기_일"]}일</b><br>'
                        f'<span style="color:#c0392b"><b>📌 추천:</b> {row["추천조치"]}</span>'
                        f'<div class="pm-tooltip">{detail_html}</div>'
                        f'</div>',
                        unsafe_allow_html=True)

            with st.expander("📋 예방정비 추천 전체 목록"):
                show_pm = pm[['우선순위','우선순위점수','라인_차종','고장설비','설비유형',
                               '고장부위','건수','총정지시간','평균MTTR','고장주기_일','추천조치']].copy()
                st.dataframe(show_pm,use_container_width=True)

            if st.button("📥 예방정비 추천 Excel",key='pm_xl'):
                out = to_excel({'예방정비추천':show_pm})
                st.download_button("⬇️ 다운로드",data=out,
                                   file_name=f"예방정비추천_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',key='pm_dl')


# ══════════════════════════════════════════════════════
# TAB 8 — 월별 트렌드 ★신규★
# ══════════════════════════════════════════════════════
with tab8:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 📈 월별 고장 트렌드")
        tf1,tf2 = st.columns([3,1])
        with tf2:
            eq_t = ['전체']+sorted(df['설비유형'].dropna().unique().tolist(),key=str)
            sel_eq_t = st.selectbox("설비유형",eq_t,key='t8eq')

        tdf = apply_global_filter(df)
        if sel_eq_t != '전체': tdf = tdf[tdf['설비유형']==sel_eq_t]
        tdf = tdf[tdf['발생일시'].notna()].copy()
        tdf['년월'] = tdf['발생일시'].dt.to_period('M').astype(str)

        if tdf.empty:
            st.warning("데이터가 없습니다.")
        else:
            monthly = (tdf.groupby('년월')
                       .agg(건수=('소요시간','count'),총정지시간=('소요시간','sum'))
                       .reset_index().sort_values('년월'))
            monthly['이동평균3M'] = monthly['건수'].rolling(3,min_periods=1).mean().round(1)

            st.markdown("##### ① 월별 고장건수 · 정지시간 복합 추이")
            fig_tr = make_subplots(specs=[[{"secondary_y":True}]])
            fig_tr.add_trace(go.Bar(x=monthly['년월'],y=monthly['건수'],name='고장건수',
                                     marker_color='#1e3a5f',opacity=0.8,
                                     hovertemplate='%{x}<br>건수: %{y}건<extra></extra>'),secondary_y=False)
            fig_tr.add_trace(go.Scatter(x=monthly['년월'],y=monthly['이동평균3M'],name='3개월 이동평균',
                                         line=dict(color='#e74c3c',width=2,dash='dot'),mode='lines',
                                         hovertemplate='%{x}<br>이동평균: %{y:.1f}건<extra></extra>'),secondary_y=False)
            fig_tr.add_trace(go.Scatter(x=monthly['년월'],y=monthly['총정지시간'],name='총정지시간(분)',
                                         line=dict(color='#f39c12',width=2),mode='lines+markers',
                                         hovertemplate='%{x}<br>정지시간: %{y:,.0f}분<extra></extra>'),secondary_y=True)
            fig_tr.update_yaxes(title_text='고장건수',secondary_y=False)
            fig_tr.update_yaxes(title_text='총정지시간(분)',secondary_y=True)
            fig_tr.update_layout(height=380,margin=dict(t=20,b=40,l=10,r=60),xaxis_tickangle=-45,
                                  legend=dict(orientation='h',y=1.08))
            st.plotly_chart(fig_tr,use_container_width=True)

            st.markdown("##### ② 설비유형별 월별 고장건수 (스택)")
            monthly_eq = tdf.groupby(['년월','설비유형']).size().reset_index(name='건수').sort_values('년월')
            fig_stk = px.bar(monthly_eq,x='년월',y='건수',color='설비유형',barmode='stack',
                             color_discrete_sequence=px.colors.qualitative.Set2)
            fig_stk.update_layout(height=340,margin=dict(t=20,b=50),xaxis_tickangle=-45,
                                   legend=dict(orientation='h',y=1.05))
            st.plotly_chart(fig_stk,use_container_width=True)

            st.markdown("##### ③ 라인별 × 월별 히트맵")
            heat_ln = tdf.groupby(['라인','년월']).size().reset_index(name='건수')
            pivot_ln = heat_ln.pivot(index='라인',columns='년월',values='건수').fillna(0).astype(int)
            fig_hm = px.imshow(pivot_ln,color_continuous_scale='YlOrRd',text_auto=True,
                               labels=dict(x='월',y='라인',color='건수'),aspect='auto')
            fig_hm.update_layout(height=max(280,len(pivot_ln)*28+80),margin=dict(t=20,b=20))
            st.plotly_chart(fig_hm,use_container_width=True)

            _t8_yrs = tdf['년'].dropna().unique() if tdf is not None and not tdf.empty else []
            if len(_t8_yrs)>=2:
                st.markdown("##### ④ 연도별 월별 비교")
                tdf['월'] = tdf['발생일시'].dt.month
                cmp = tdf.groupby(['년','월']).size().reset_index(name='건수')
                cmp['년'] = cmp['년'].astype(str)
                fig_cmp = px.line(cmp,x='월',y='건수',color='년',markers=True,
                                  color_discrete_sequence=px.colors.qualitative.D3)
                fig_cmp.update_xaxes(tickvals=list(range(1,13)),ticktext=[f'{m}월' for m in range(1,13)])
                fig_cmp.update_layout(height=320,margin=dict(t=20,b=20))
                st.plotly_chart(fig_cmp,use_container_width=True)


# ══════════════════════════════════════════════════════
# TAB 9 — KPI 목표관리 ★연/월 선택 + 전월비교 + 목표저장★
# ══════════════════════════════════════════════════════
with tab9:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 🎯 KPI 목표 대비 실적")

        # ── 목표값 설정 (저장/불러오기 포함) ───────────────
        with st.expander("⚙️ KPI 목표값 설정 (클릭하여 수정)", expanded=True):
            tg1,tg2,tg3,tg4 = st.columns(4)
            with tg1:
                st.session_state.kpi_target_cnt = st.number_input(
                    "고장건수 목표(건)", min_value=1,
                    value=st.session_state.kpi_target_cnt, step=5, key='_kpi_cnt')
            with tg2:
                st.session_state.kpi_target_mttr = st.number_input(
                    "MTTR 목표(분)", min_value=1,
                    value=st.session_state.kpi_target_mttr, step=1, key='_kpi_mttr')
            with tg3:
                st.session_state.kpi_target_stop = st.number_input(
                    "총정지시간 목표(분)", min_value=10,
                    value=st.session_state.kpi_target_stop, step=100, key='_kpi_stop')
            with tg4:
                st.session_state.kpi_target_mtbf = st.number_input(
                    "MTBF 목표(시간)", min_value=1,
                    value=st.session_state.kpi_target_mtbf, step=10, key='_kpi_mtbf')
            st.caption("💾 목표값은 세션 동안 유지됩니다. 앱 재시작 시 초기화됩니다.")

        target_cnt  = st.session_state.kpi_target_cnt
        target_mttr = st.session_state.kpi_target_mttr
        target_stop = st.session_state.kpi_target_stop
        target_mtbf = st.session_state.kpi_target_mtbf

        st.divider()

        # ── 기간 선택: 연간 / 월간 ───────────────────────
        kpi_mode = st.radio("분석 기간", ["📅 연간", "📆 월간"], horizontal=True, key='kpi_mode')

        all_yrs = sorted(df['년'].dropna().unique().astype(int), reverse=True)
        df_kpi_base = df[df['발생일시'].notna() & df['소요시간'].notna()].copy()
        df_kpi_base['년'] = df_kpi_base['발생일시'].dt.year
        df_kpi_base['월'] = df_kpi_base['발생일시'].dt.month
        df_kpi_base['년월'] = df_kpi_base['발생일시'].dt.to_period('M').astype(str)

        # ── 연간 모드 ──────────────────────────────────
        if kpi_mode == "📅 연간":
            kp1, kp2 = st.columns([2, 1])
            with kp1:
                sel_kpi_yr = st.selectbox("기준 연도", all_yrs, key='kpi_yr_ann')
            with kp2:
                cmp_kpi_yr = st.selectbox("비교 연도 (전년)", [y for y in all_yrs if y != sel_kpi_yr] or all_yrs,
                                           key='kpi_yr_cmp')

            kdf  = df_kpi_base[df_kpi_base['년'] == sel_kpi_yr]
            kdf_p = df_kpi_base[df_kpi_base['년'] == cmp_kpi_yr]

            def _calc_kpi_metrics(d):
                if d.empty:
                    return 0, 0, 0, 0
                mo = d.groupby('년월').agg(건수=('소요시간','count'), 정지=('소요시간','sum')).reset_index()
                avg_cnt  = mo['건수'].mean()
                avg_stop = mo['정지'].mean()
                avg_mttr = d['소요시간'].mean()
                # 방법B 근사 MTBF
                mttr_r, _ = calc_mttr_mtbf(d, cluster_min=60)
                avg_mtbf = mttr_r['MTBF_근사(시간)'].dropna().mean() if not mttr_r.empty else 0
                return avg_cnt, avg_stop, avg_mttr, avg_mtbf

            cur_cnt, cur_stop, cur_mttr, cur_mtbf   = _calc_kpi_metrics(kdf)
            prev_cnt, prev_stop, prev_mttr, prev_mtbf = _calc_kpi_metrics(kdf_p)

            period_label = f"{sel_kpi_yr}년"
            x_label = "월평균"
            compare_label = f"전년({cmp_kpi_yr}) 대비"

            monthly_kdf = (kdf.groupby('년월')
                           .agg(건수=('소요시간','count'), 총정지시간=('소요시간','sum'))
                           .reset_index().sort_values('년월'))

        # ── 월간 모드 ──────────────────────────────────
        else:
            all_yms = sorted(df_kpi_base['년월'].unique(), reverse=True)
            km1, km2 = st.columns([2, 1])
            with km1:
                sel_kpi_ym = st.selectbox("기준 월", all_yms, key='kpi_ym_sel')
            with km2:
                # 자동 전월 계산
                from datetime import date
                _yr = int(sel_kpi_ym[:4]); _mo = int(sel_kpi_ym[5:7])
                _prev_yr = _yr if _mo > 1 else _yr - 1
                _prev_mo = _mo - 1 if _mo > 1 else 12
                _prev_ym = f"{_prev_yr}-{_prev_mo:02d}"
                cmp_kpi_ym = st.selectbox(
                    "비교 월 (자동=전월)",
                    [_prev_ym] + [y for y in all_yms if y != sel_kpi_ym and y != _prev_ym],
                    key='kpi_ym_cmp')

            kdf   = df_kpi_base[df_kpi_base['년월'] == sel_kpi_ym]
            kdf_p = df_kpi_base[df_kpi_base['년월'] == cmp_kpi_ym]

            def _calc_kpi_metrics_mo(d):
                if d.empty: return 0, 0, 0, 0
                cnt  = len(d)
                stop = d['소요시간'].sum()
                mttr = d['소요시간'].mean()
                # 방법B 근사 MTBF
                mttr_r, _ = calc_mttr_mtbf(d, cluster_min=60)
                mtbf = mttr_r['MTBF_근사(시간)'].dropna().mean() if not mttr_r.empty else 0
                return cnt, stop, mttr, mtbf

            cur_cnt,  cur_stop,  cur_mttr,  cur_mtbf  = _calc_kpi_metrics_mo(kdf)
            prev_cnt, prev_stop, prev_mttr, prev_mtbf = _calc_kpi_metrics_mo(kdf_p)

            period_label = f"{sel_kpi_ym} 월"
            x_label = "당월 합계"
            compare_label = f"전월({cmp_kpi_ym}) 대비"

            # 월간: 일별 추이
            kdf2 = kdf.copy()
            kdf2['일'] = kdf2['발생일시'].dt.day
            monthly_kdf = (kdf2.groupby('일')
                           .agg(건수=('소요시간','count'), 총정지시간=('소요시간','sum'))
                           .reset_index().rename(columns={'일':'년월'}))
            monthly_kdf['년월'] = monthly_kdf['년월'].astype(str) + '일'

        # ── 전월/전년 delta 계산 ──────────────────────
        def _delta_str(cur, prev, unit='', higher_better=False):
            if prev == 0: return None
            diff = cur - prev
            pct  = diff / prev * 100
            sign = '+' if diff >= 0 else ''
            bad  = (diff > 0 and not higher_better) or (diff < 0 and higher_better)
            arrow = '▲' if diff >= 0 else '▼'
            color = 'red' if bad else 'green'
            return f"{arrow} {sign}{diff:.1f}{unit} ({sign}{pct:.1f}%)"

        # ── KPI 카드 4개 (전월비교 delta 포함) ──────────
        st.markdown(f"#### {period_label} KPI 실적")
        kk1,kk2,kk3,kk4 = st.columns(4)
        with kk1:
            st.markdown(kpi_card_html(
                f"고장건수 ({x_label})", cur_cnt, target_cnt, "건",
                prev_val=prev_cnt, compare_label=compare_label), unsafe_allow_html=True)
        with kk2:
            st.markdown(kpi_card_html(
                "평균 MTTR", cur_mttr, target_mttr, "분",
                prev_val=prev_mttr, compare_label=compare_label), unsafe_allow_html=True)
        with kk3:
            st.markdown(kpi_card_html(
                f"총정지시간 ({x_label})", cur_stop, target_stop, "분",
                prev_val=prev_stop, compare_label=compare_label), unsafe_allow_html=True)
        with kk4:
            st.markdown(kpi_card_html(
                "평균 MTBF", cur_mtbf, target_mtbf, "시간", higher_is_better=True,
                prev_val=prev_mtbf, compare_label=compare_label), unsafe_allow_html=True)

        st.divider()

        # ── 추이 차트 ────────────────────────────────
        x_col = '년월'
        st.markdown(f"##### 실적 추이 vs 목표선")
        fig_kpi = make_subplots(rows=2, cols=1, shared_xaxes=True,
                                 subplot_titles=('고장건수 vs 목표', '총정지시간 vs 목표'),
                                 vertical_spacing=0.12)
        fig_kpi.add_trace(go.Bar(
            x=monthly_kdf[x_col], y=monthly_kdf['건수'],
            name='고장건수', marker_color='#1e3a5f',
            hovertemplate='%{x}<br>건수: %{y}건<extra></extra>'), row=1, col=1)
        fig_kpi.add_hline(y=target_cnt, line_dash='dash', line_color='#e74c3c',
                           annotation_text=f'목표 {target_cnt}건', row=1, col=1)
        fig_kpi.add_trace(go.Bar(
            x=monthly_kdf[x_col], y=monthly_kdf['총정지시간'],
            name='총정지시간', marker_color='#e67e22',
            hovertemplate='%{x}<br>정지시간: %{y:,.0f}분<extra></extra>'), row=2, col=1)
        fig_kpi.add_hline(y=target_stop, line_dash='dash', line_color='#e74c3c',
                           annotation_text=f'목표 {target_stop}분', row=2, col=1)
        fig_kpi.update_layout(height=500, margin=dict(t=40, b=40, l=10, r=80),
                               showlegend=False)
        fig_kpi.update_xaxes(tickangle=-45)
        st.plotly_chart(fig_kpi, use_container_width=True)

        # ── Gauge ─────────────────────────────────────
        def make_gauge(title, value, target, higher_is_better=False):
            if higher_is_better:
                pct = min(value/target*100, 150) if target else 0
            else:
                pct = min(target/value*100, 150) if value else 0
            color = '#27ae60' if pct>=100 else ('#e67e22' if pct>=80 else '#e74c3c')
            return go.Indicator(
                mode='gauge+number+delta', value=pct,
                title={'text': title, 'font': {'size': 13}},
                number={'suffix': '%', 'font': {'size': 22}},
                delta={'reference': 100, 'suffix': '%'},
                gauge={'axis': {'range': [0, 150]}, 'bar': {'color': color},
                       'steps': [{'range':[0,80],'color':'#fdecea'},
                                  {'range':[80,100],'color':'#fff3e0'},
                                  {'range':[100,150],'color':'#e8f5e9'}],
                       'threshold': {'line':{'color':'black','width':3},
                                     'thickness':0.75, 'value':100}})
        gf = make_subplots(rows=1, cols=4, specs=[[{'type':'indicator'}]*4])
        gf.add_trace(make_gauge("고장건수 달성률", cur_cnt, target_cnt), row=1, col=1)
        gf.add_trace(make_gauge("MTTR 달성률",   cur_mttr, target_mttr), row=1, col=2)
        gf.add_trace(make_gauge("정지시간 달성률", cur_stop, target_stop), row=1, col=3)
        gf.add_trace(make_gauge("MTBF 달성률",   cur_mtbf, target_mtbf, True), row=1, col=4)
        gf.update_layout(height=260, margin=dict(t=30, b=10))
        st.plotly_chart(gf, use_container_width=True)

        # ── 설비유형별 KPI 상세 ───────────────────────
        st.markdown("##### 설비유형별 KPI 상세 비교")
        eq_kpi = (kdf.groupby('설비유형')
                  .agg(건수=('소요시간','count'), 총정지시간=('소요시간','sum'),
                       평균MTTR=('소요시간','mean'))
                  .reset_index().sort_values('건수', ascending=False))
        eq_kpi['평균MTTR'] = eq_kpi['평균MTTR'].round(1)
        eq_kpi['목표달성_MTTR'] = eq_kpi['평균MTTR'].apply(
            lambda v: '✅' if v <= target_mttr else '❌')
        st.dataframe(eq_kpi, use_container_width=True, hide_index=True)

        # ── 자동 코멘트 (룰 기반) ────────────────────
        st.divider()
        st.markdown("##### 📝 자동 분석 코멘트")
        comment_lines = []
        # 건수 판정
        if cur_cnt <= target_cnt:
            comment_lines.append(f"✅ 고장건수 {cur_cnt:.0f}건 — 목표 {target_cnt}건 **달성** ({period_label})")
        else:
            over = cur_cnt - target_cnt
            comment_lines.append(f"❌ 고장건수 {cur_cnt:.0f}건 — 목표 대비 **{over:.0f}건 초과** ({period_label})")
        # 전기 계통 반복 판정
        if '고장계통코드' in kdf.columns and not kdf.empty:
            top_sys = kdf['고장계통코드'].value_counts().head(1)
            if not top_sys.empty:
                sys_pct = top_sys.iloc[0] / len(kdf) * 100
                if sys_pct >= 40:
                    comment_lines.append(
                        f"🔍 **{top_sys.index[0]}** 계통 고장이 {sys_pct:.0f}% 집중 — "
                        f"해당 계통 집중 점검 권고")
        # 원인코드 반복
        if '원인코드' in kdf.columns and not kdf.empty:
            top_ca = kdf['원인코드'].value_counts().head(1)
            if not top_ca.empty and top_ca.iloc[0] >= 3:
                comment_lines.append(
                    f"🔩 원인 **'{top_ca.index[0]}'** {top_ca.iloc[0]}건 반복 — "
                    f"설계적 대책 또는 예방보전 주기 단축 검토")
        # MTTR 판정
        if cur_mttr > target_mttr:
            comment_lines.append(
                f"⏱️ 평균 MTTR {cur_mttr:.0f}분 — 목표 {target_mttr}분 초과. "
                f"예비품 재고 및 정비 매뉴얼 점검 권고")
        # 재발 판정
        if '재발여부' in kdf.columns and not kdf.empty:
            r_pct = kdf['재발여부'].mean() * 100
            if r_pct >= 25:
                comment_lines.append(
                    f"🔁 {period_label} 재발 고장 비율 {r_pct:.0f}% — "
                    f"재발 억제 대책 수립 필요 (恒久대책 미수립 의심)")
        # 전기계 비교
        if prev_cnt > 0:
            chg = (cur_cnt - prev_cnt) / prev_cnt * 100
            direction = "증가" if chg > 0 else "감소"
            comment_lines.append(
                f"📊 {compare_label} 고장건수 **{abs(chg):.1f}% {direction}** "
                f"({prev_cnt:.0f}건 → {cur_cnt:.0f}건)")

        for line in comment_lines:
            st.markdown(line)


# ══════════════════════════════════════════════════════
# TAB 10 — BM/PM 분석 ★신규★
# ══════════════════════════════════════════════════════
with tab10:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 🔄 돌발(BM) vs 예방(PM) 분석")
        bf1,bf2 = st.columns([3,1])
        with bf2:
            eq_b = ['전체']+sorted(df['설비유형'].dropna().unique().tolist(),key=str)
            sel_eq_b = st.selectbox("설비유형",eq_b,key='t10eq')

        bdf = apply_global_filter(df)
        if sel_eq_b != '전체': bdf = bdf[bdf['설비유형']==sel_eq_b]

        if bdf.empty:
            st.warning("데이터가 없습니다.")
        else:
            total_b = len(bdf)
            bm_cnt = (bdf['보전구분']=='BM(돌발)').sum()
            pm_cnt = (bdf['보전구분']=='PM(계획)').sum()
            bm_pct = bm_cnt/total_b*100 if total_b else 0
            pm_pct = pm_cnt/total_b*100 if total_b else 0

            bk1,bk2,bk3,bk4 = st.columns(4)
            bk1.metric("전체 건수",f"{total_b:,}건")
            bk2.metric("BM(돌발)",f"{bm_cnt:,}건",f"{bm_pct:.1f}%")
            bk3.metric("PM(계획)",f"{pm_cnt:,}건",f"{pm_pct:.1f}%")
            bk4.metric("PM비율 목표 30%","✅ 달성" if pm_pct>=30 else "❌ 미달")
            st.divider()

            bc1,bc2 = st.columns(2)
            with bc1:
                pie_data = bdf['보전구분'].value_counts().reset_index()
                pie_data.columns = ['구분','건수']
                fig_pie = px.pie(pie_data,values='건수',names='구분',
                                 color='구분',color_discrete_map={'BM(돌발)':'#e74c3c','PM(계획)':'#27ae60'},hole=0.4)
                fig_pie.update_traces(textinfo='label+percent')
                fig_pie.update_layout(height=300,margin=dict(t=20,b=10),showlegend=False,title="전체 BM/PM 비율")
                st.plotly_chart(fig_pie,use_container_width=True)
            with bc2:
                eq_bm = bdf.groupby(['설비유형','보전구분']).size().reset_index(name='건수')
                fig_eq = px.bar(eq_bm,x='설비유형',y='건수',color='보전구분',barmode='stack',
                                color_discrete_map={'BM(돌발)':'#e74c3c','PM(계획)':'#27ae60'},
                                title="설비유형별 BM/PM")
                fig_eq.update_layout(height=300,margin=dict(t=40,b=60),xaxis_tickangle=-30,
                                      legend=dict(orientation='h',y=1.05))
                st.plotly_chart(fig_eq,use_container_width=True)

            st.markdown("##### 월별 BM/PM 비율 추이")
            bdf2 = bdf[bdf['발생일시'].notna()].copy()
            bdf2['년월'] = bdf2['발생일시'].dt.to_period('M').astype(str)
            monthly_bm = bdf2.groupby(['년월','보전구분']).size().reset_index(name='건수').sort_values('년월')
            total_mo = bdf2.groupby('년월').size().reset_index(name='합계')
            pm_mo    = bdf2[bdf2['보전구분']=='PM(계획)'].groupby('년월').size().reset_index(name='PM')
            ratio_mo = total_mo.merge(pm_mo,on='년월',how='left').fillna(0)
            ratio_mo['PM비율'] = (ratio_mo['PM']/ratio_mo['합계']*100).round(1)

            fig_bm = make_subplots(specs=[[{"secondary_y":True}]])
            for btype,col in [('BM(돌발)','#e74c3c'),('PM(계획)','#27ae60')]:
                sub = monthly_bm[monthly_bm['보전구분']==btype]
                fig_bm.add_trace(go.Bar(x=sub['년월'],y=sub['건수'],name=btype,
                                         marker_color=col,opacity=0.85),secondary_y=False)
            fig_bm.add_trace(go.Scatter(x=ratio_mo['년월'],y=ratio_mo['PM비율'],name='PM비율(%)',
                                         line=dict(color='#2ecc71',width=2),mode='lines+markers'),secondary_y=True)
            fig_bm.add_hline(y=30,line_dash='dash',line_color='green',annotation_text='PM목표 30%',secondary_y=True)
            fig_bm.update_yaxes(title_text='건수',secondary_y=False)
            fig_bm.update_yaxes(title_text='PM비율(%)',range=[0,100],secondary_y=True)
            fig_bm.update_layout(height=360,margin=dict(t=20,b=50),xaxis_tickangle=-45,
                                  barmode='stack',legend=dict(orientation='h',y=1.08))
            st.plotly_chart(fig_bm,use_container_width=True)

            st.markdown("##### BM 야간(22시~06시) 비율 — 설비유형별")
            bm_night = bdf[bdf['발생일시'].notna() & (bdf['보전구분']=='BM(돌발)')].copy()
            bm_night['야간'] = bm_night['발생일시'].dt.hour.apply(lambda h: '야간(22~06시)' if h>=22 or h<6 else '주간(06~22시)')
            ng = bm_night.groupby(['설비유형','야간']).size().reset_index(name='건수')
            fig_ng = px.bar(ng,x='설비유형',y='건수',color='야간',barmode='group',
                            color_discrete_map={'야간(22~06시)':'#2c3e50','주간(06~22시)':'#85c1e9'})
            fig_ng.update_layout(height=300,margin=dict(t=20,b=60),xaxis_tickangle=-30,
                                  legend=dict(orientation='h',y=1.05))
            st.plotly_chart(fig_ng,use_container_width=True)


# ══════════════════════════════════════════════════════
# TAB 11 — 재발 고장 전용 ★신규★
# ══════════════════════════════════════════════════════
with tab11:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 🔁 재발 고장 전용 분석")
        st.caption("수리 완료가 아니라 **재발 억제** 관점에서 관리합니다")

        rc1,rc2,rc3,rc4 = st.columns([2,2,1,1])
        with rc2:
            eq_rc = ['전체']+sorted(df['설비유형'].dropna().unique().tolist(),key=str)
            sel_eq_rc = st.selectbox("설비유형",eq_rc,key='t11eq')
        with rc3:
            window_rc = st.selectbox("재발 판정 윈도우",['30일','60일','90일','180일'],index=2,key='t11win')
        with rc4:
            min_recur_cnt = st.number_input("최소 발생건수",min_value=1,max_value=20,value=3,step=1,key='t11mn',
                                             help="설비별 재발률 차트에 포함할 최소 고장건수 기준")
        win_days = int(window_rc.replace('일',''))

        rdf = apply_global_filter(df)
        if sel_eq_rc != '전체': rdf = rdf[rdf['설비유형']==sel_eq_rc]
        rdf = rdf[rdf['발생일시'].notna()].sort_values('발생일시').reset_index(drop=True)

        # 재발 재계산 (윈도우 변경 반영)
        rdf['재발여부'] = calc_recurrence(rdf, win_days)

        total_rc = len(rdf)
        recur_cnt = rdf['재발여부'].sum()
        recur_pct = recur_cnt/total_rc*100 if total_rc else 0

        rr1,rr2,rr3,rr4 = st.columns(4)
        rr1.metric("전체 고장건수",f"{total_rc:,}건")
        rr2.metric(f"재발 고장 ({win_days}일 기준)",f"{recur_cnt:,}건")
        rr3.metric("재발률",f"{recur_pct:.1f}%")
        rr4.metric("비재발 고장",f"{total_rc-recur_cnt:,}건")
        st.divider()

        # ── 설비별 재발률 ──
        st.markdown(f"##### 설비별 재발률 ({min_recur_cnt}건 이상 발생 설비)")
        eq_recur = (rdf.groupby(['설비_KEY','설비유형'])
                    .agg(전체건수=('재발여부','count'), 재발건수=('재발여부','sum'))
                    .reset_index())
        eq_recur = eq_recur[eq_recur['전체건수']>=min_recur_cnt].copy()
        eq_recur['재발률(%)'] = (eq_recur['재발건수']/eq_recur['전체건수']*100).round(1)
        eq_recur = eq_recur.sort_values('재발률(%)',ascending=False).reset_index(drop=True)
        eq_recur.index += 1

        def recur_grade(p):
            if p>=50: return '🔴 매우위험'
            if p>=30: return '🟠 위험'
            if p>=15: return '🟡 주의'
            return '🟢 양호'
        eq_recur['등급'] = eq_recur['재발률(%)'].apply(recur_grade)

        top_recur = eq_recur.head(25)
        color_map_r = {'🔴 매우위험':'#e74c3c','🟠 위험':'#e67e22','🟡 주의':'#f1c40f','🟢 양호':'#27ae60'}
        # ── hover helper (TAB11 재발) ───────────────────────
        def _tf_rr(key, src_df):
            sub = src_df[src_df['설비_KEY'] == key] if '설비_KEY' in src_df.columns else src_df.iloc[0:0]
            cols = [c for c in ['고장부위','현상'] if c in sub.columns]
            if not cols: return '-'
            if len(cols) == 2:
                fk = sub['고장부위'].fillna('').str.strip() + ' / ' + sub['현상'].fillna('').str.strip()
                fk = fk[(sub['고장부위'].notna() & (sub['고장부위'].str.strip()!='')) |
                        (sub['현상'].notna() & (sub['현상'].str.strip()!=''))]
                fk = fk.str.strip(' /').str.strip()
            else:
                fk = sub[cols[0]].dropna()
                fk = fk[fk.str.strip() != '']
            if fk.empty: return '-'
            top = fk.value_counts().head(3)
            return '<br>'.join([f'  {i+1}. {str(k)[:30]} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

        def _te_rr(key, src_df):
            sub = src_df[src_df['설비_KEY'] == key] if '설비_KEY' in src_df.columns else src_df.iloc[0:0]
            if '조치자' not in sub.columns: return '-'
            col = sub['조치자'].dropna()
            col = col[col.astype(str).str.strip() != '']
            if col.empty: return '-'
            top = col.value_counts().head(3)
            return '<br>'.join([f'  {i+1}. {str(k)} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'

        top_recur = top_recur.copy()
        top_recur['_tf'] = top_recur['설비_KEY'].apply(lambda k: _tf_rr(k, rdf))
        top_recur['_te'] = top_recur['설비_KEY'].apply(lambda k: _te_rr(k, rdf))
        fig_rr = px.bar(top_recur.sort_values('재발률(%)'),
                        x='재발률(%)',y='설비_KEY',orientation='h',color='등급',
                        color_discrete_map=color_map_r,
                        custom_data=['전체건수','재발건수','설비유형','_tf','_te'])
        fig_rr.update_traces(
            hovertemplate='<b>%{y}</b><br>재발률: %{x:.1f}%<br>전체: %{customdata[0]}건 / 재발: %{customdata[1]}건<br>'
            '<b>▶ 반복 고장 Top 3 (부위/현상)</b><br>%{customdata[3]}<br>'
            '<b>▶ 주요 조치자 Top 3</b><br>%{customdata[4]}'
            '<extra></extra>')
        fig_rr.add_vline(x=30,line_dash='dash',line_color='#e67e22',annotation_text='경고 30%')
        fig_rr.update_layout(height=max(400,len(top_recur)*26),margin=dict(t=50,b=20),yaxis_title='',
                             legend=dict(orientation='h',y=1.0,x=0,yanchor='bottom',xanchor='left'))
        st.plotly_chart(fig_rr,use_container_width=True)

        # ── 재발 간격 상세 분석 ──
        st.markdown("##### 동일 설비·동일 부위 재발 간격 분석")
        rdf['재발KEY'] = rdf['설비_KEY'].astype(str)+'||'+rdf['고장부위'].fillna('').astype(str)
        gap_rows = []
        for key,grp in rdf.groupby('재발KEY'):
            grp = grp.sort_values('발생일시')
            n = len(grp)
            if n < 2: continue
            gaps = grp['발생일시'].diff().dropna().dt.days
            gaps = gaps[gaps>0]
            if len(gaps)==0: continue
            parts = key.split('||')
            gap_rows.append({
                '설비_KEY': parts[0],
                '고장부위': parts[1] if len(parts)>1 else '',
                '재발횟수': n,
                '평균재발간격_일': round(gaps.mean(),1),
                '최단재발간격_일': int(gaps.min()),
                '최장재발간격_일': int(gaps.max()),
                '총정지시간': round(grp['소요시간'].sum(),0) if grp['소요시간'].notna().any() else 0,
                '설비유형': grp['설비유형'].mode()[0] if not grp['설비유형'].isna().all() else '',
                '최근발생': grp['발생일시'].max().strftime('%Y-%m-%d'),
                '주요현상': (
                    grp['현상'].dropna().astype(str)
                    .pipe(lambda s: s[s.str.strip()!=''].value_counts().index[0]
                          if (s.str.strip()!='').any() else '-')
                    if '현상' in grp.columns else '-'),
            })
        if gap_rows:
            gap_df = pd.DataFrame(gap_rows).sort_values('평균재발간격_일')
            gap_df['위험등급'] = gap_df['평균재발간격_일'].apply(
                lambda d: '🔴 2주 이내' if d<14 else ('🟠 1달 이내' if d<30 else ('🟡 3달 이내' if d<90 else '🟢 관찰')))

            # 위험 카드
            danger = gap_df[gap_df['위험등급'].str.startswith('🔴')].head(8)
            if not danger.empty:
                st.markdown("**🔴 평균 재발간격 2주 이내 — 즉시 근본원인 분석 필요**")
                for _,row in danger.iterrows():
                    st.markdown(f'<div class="card-red">'
                                f'<b>{row["설비_KEY"]}</b> | 부위: <b>{row["고장부위"] or "미상"}</b>' f' | 현상: <b>{str(row.get("주요현상","-")) or "-"}</b>'
                                f' <span style="color:#888;font-size:12px">({row["설비유형"]})</span><br>'
                                f'재발 <b>{row["재발횟수"]}회</b> / 평균간격 <b>{row["평균재발간격_일"]}일</b>'
                                f' / 최단 <b>{row["최단재발간격_일"]}일</b>'
                                f' / 총정지 <b>{row["총정지시간"]:.0f}분</b> / 최근: {row["최근발생"]}<br>'
                                f'<span style="color:#c0392b;font-size:12px">'
                                f'📌 恒久대책 수립 — 계획정비 전환 검토 필요</span>'
                                f'</div>', unsafe_allow_html=True)

            with st.expander("📋 전체 재발 간격 분석 테이블"):
                st.dataframe(gap_df.reset_index(drop=True),use_container_width=True)

        # ── 최근 30일 / 90일 재발 현황 ──
        st.markdown("##### 최근 기간별 재발 현황")
        now = rdf['발생일시'].max()
        for days,label in [(30,'최근 30일'),(90,'최근 90일')]:
            cutoff = now - timedelta(days=days)
            recent = rdf[rdf['발생일시']>=cutoff]
            r_cnt = recent['재발여부'].sum()
            r_total = len(recent)
            r_pct = r_cnt/r_total*100 if r_total else 0
            badge_color = '#e74c3c' if r_pct>=30 else ('#e67e22' if r_pct>=15 else '#27ae60')
            st.markdown(f'<div style="border-left:4px solid {badge_color};background:#f8f9fa;'
                        f'padding:10px 16px;border-radius:0 6px 6px 0;margin-bottom:6px;">'
                        f'<b>{label}</b>: 전체 {r_total}건 중 재발 {r_cnt}건 '
                        f'<b style="color:{badge_color};">({r_pct:.1f}%)</b>'
                        f'</div>', unsafe_allow_html=True)

        if st.button("📥 재발분석 Excel 다운로드",key='recur_xl'):
            if gap_rows:
                out = to_excel({'재발간격분석':gap_df,'설비별재발률':eq_recur})
                st.download_button("⬇️ 다운로드",data=out,
                                   file_name=f"재발분석_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                   key='recur_dl')


# ══════════════════════════════════════════════════════
# TAB 12 — 표준코드 분석 ★신규★
# ══════════════════════════════════════════════════════
with tab12:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 🏷️ 표준코드 분석")
        st.caption("고장계통코드 · 원인코드 · 조치코드 기반 정밀 Pareto 분석")

        st.info("💡 표준코드는 현상/원인/조치 텍스트를 자동 분류한 값입니다. "
                "차후 원본 데이터에 코드 컬럼을 직접 추가필요.")

        sf1,sf2 = st.columns([3,1])
        with sf2:
            eq_s = ['전체']+sorted(df['설비유형'].dropna().unique().tolist(),key=str)
            sel_eq_s = st.selectbox("설비유형",eq_s,key='t12eq')

        sdf = apply_global_filter(df)
        if sel_eq_s != '전체': sdf = sdf[sdf['설비유형']==sel_eq_s]

        if sdf.empty:
            st.warning("데이터가 없습니다.")
        else:
            sc1,sc2,sc3 = st.columns(3)

            with sc1:
                st.markdown("##### 고장계통코드 Pareto")
                if '고장계통코드' in sdf.columns:
                    def _te_sys(code, src):
                        sub = src[src['고장계통코드']==code]
                        col = '설비_KEY' if '설비_KEY' in sub.columns else ('고장설비' if '고장설비' in sub.columns else None)
                        if col is None: return '-'
                        c = sub[col].dropna(); c = c[c.astype(str).str.strip()!='']
                        if c.empty: return '-'
                        top = c.value_counts().head(3)
                        return '<br>'.join([f'  {i+1}. {k} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'
                    def _tf_sys(code, src):
                        sub = src[src['고장계통코드']==code]
                        if '현상' not in sub.columns: return '-'
                        c = sub['현상'].dropna(); c = c[c.str.strip()!='']
                        if c.empty: return '-'
                        top = c.value_counts().head(3)
                        return '<br>'.join([f'  {i+1}. {str(k)[:30]} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'
                    sys_grp = sdf['고장계통코드'].value_counts().reset_index()
                    sys_grp.columns = ['계통','건수']
                    sys_grp['누적%'] = (sys_grp['건수'].cumsum()/sys_grp['건수'].sum()*100).round(1)
                    sys_grp['_te'] = sys_grp['계통'].apply(lambda c: _te_sys(c, sdf))
                    sys_grp['_tf'] = sys_grp['계통'].apply(lambda c: _tf_sys(c, sdf))
                    fig_sys = make_subplots(specs=[[{"secondary_y":True}]])
                    fig_sys.add_trace(go.Bar(x=sys_grp['계통'],y=sys_grp['건수'],
                                             marker_color='#1e3a5f',name='건수',
                                             customdata=np.stack([sys_grp['누적%'],sys_grp['_tf'],sys_grp['_te']],axis=-1),
                                             hovertemplate='<b>%{x}</b><br>건수: %{y}건<br>누적: %{customdata[0]:.1f}%<br><b>▶ 반복 고장 Top 3</b><br>%{customdata[1]}<br><b>▶ 고장 다발 설비 Top 3</b><br>%{customdata[2]}<extra></extra>'),secondary_y=False)
                    fig_sys.add_trace(go.Scatter(x=sys_grp['계통'],y=sys_grp['누적%'],
                                                  mode='lines+markers',line=dict(color='#e74c3c',width=2),
                                                  name='누적%'),secondary_y=True)
                    fig_sys.add_hline(y=80,line_dash='dash',line_color='orange',secondary_y=True)
                    fig_sys.update_yaxes(title_text='건수',secondary_y=False)
                    fig_sys.update_yaxes(title_text='누적%',range=[0,105],secondary_y=True)
                    fig_sys.update_layout(height=340,margin=dict(t=10,b=40,l=10,r=50),showlegend=False)
                    st.plotly_chart(fig_sys,use_container_width=True)

            with sc2:
                st.markdown("##### 원인코드 Pareto")
                if '원인코드' in sdf.columns:
                    def _te_ca(code, src):
                        sub = src[src['원인코드']==code]
                        col = '설비_KEY' if '설비_KEY' in sub.columns else ('고장설비' if '고장설비' in sub.columns else None)
                        if col is None: return '-'
                        c = sub[col].dropna(); c = c[c.astype(str).str.strip()!='']
                        if c.empty: return '-'
                        top = c.value_counts().head(3)
                        return '<br>'.join([f'  {i+1}. {k} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'
                    ca_grp = sdf['원인코드'].value_counts().reset_index()
                    ca_grp.columns = ['원인','건수']
                    ca_grp['누적%'] = (ca_grp['건수'].cumsum()/ca_grp['건수'].sum()*100).round(1)
                    ca_grp['_te'] = ca_grp['원인'].apply(lambda c: _te_ca(c, sdf))
                    fig_ca = make_subplots(specs=[[{"secondary_y":True}]])
                    fig_ca.add_trace(go.Bar(x=ca_grp['원인'],y=ca_grp['건수'],
                                            marker_color='#8e44ad',name='건수',
                                            customdata=np.stack([ca_grp['누적%'],ca_grp['_te']],axis=-1),
                                            hovertemplate='<b>%{x}</b><br>건수: %{y}건<br>누적: %{customdata[0]:.1f}%<br><b>▶ 고장 다발 설비 Top 3</b><br>%{customdata[1]}<extra></extra>'),secondary_y=False)
                    fig_ca.add_trace(go.Scatter(x=ca_grp['원인'],y=ca_grp['누적%'],
                                                 mode='lines+markers',line=dict(color='#e74c3c',width=2),
                                                 name='누적%'),secondary_y=True)
                    fig_ca.add_hline(y=80,line_dash='dash',line_color='orange',secondary_y=True)
                    fig_ca.update_yaxes(title_text='건수',secondary_y=False)
                    fig_ca.update_yaxes(title_text='누적%',range=[0,105],secondary_y=True)
                    fig_ca.update_layout(height=340,margin=dict(t=10,b=40,l=10,r=50),showlegend=False)
                    st.plotly_chart(fig_ca,use_container_width=True)

            with sc3:
                st.markdown("##### 조치코드 Pareto")
                if '조치코드' in sdf.columns:
                    def _te_ac(code, src):
                        sub = src[src['조치코드']==code]
                        col = '설비_KEY' if '설비_KEY' in sub.columns else ('고장설비' if '고장설비' in sub.columns else None)
                        if col is None: return '-'
                        c = sub[col].dropna(); c = c[c.astype(str).str.strip()!='']
                        if c.empty: return '-'
                        top = c.value_counts().head(3)
                        return '<br>'.join([f'  {i+1}. {k} ({v}건)' for i,(k,v) in enumerate(top.items())]) or '-'
                    ac_grp = sdf['조치코드'].value_counts().reset_index()
                    ac_grp.columns = ['조치','건수']
                    ac_grp['누적%'] = (ac_grp['건수'].cumsum()/ac_grp['건수'].sum()*100).round(1)
                    ac_grp['_te'] = ac_grp['조치'].apply(lambda c: _te_ac(c, sdf))
                    fig_ac = make_subplots(specs=[[{"secondary_y":True}]])
                    fig_ac.add_trace(go.Bar(x=ac_grp['조치'],y=ac_grp['건수'],
                                            marker_color='#27ae60',name='건수',
                                            customdata=np.stack([ac_grp['누적%'],ac_grp['_te']],axis=-1),
                                            hovertemplate='<b>%{x}</b><br>건수: %{y}건<br>누적: %{customdata[0]:.1f}%<br><b>▶ 고장 다발 설비 Top 3</b><br>%{customdata[1]}<extra></extra>'),secondary_y=False)
                    fig_ac.add_trace(go.Scatter(x=ac_grp['조치'],y=ac_grp['누적%'],
                                                 mode='lines+markers',line=dict(color='#e74c3c',width=2),
                                                 name='누적%'),secondary_y=True)
                    fig_ac.add_hline(y=80,line_dash='dash',line_color='orange',secondary_y=True)
                    fig_ac.update_yaxes(title_text='건수',secondary_y=False)
                    fig_ac.update_yaxes(title_text='누적%',range=[0,105],secondary_y=True)
                    fig_ac.update_layout(height=340,margin=dict(t=10,b=40,l=10,r=50),showlegend=False)
                    st.plotly_chart(fig_ac,use_container_width=True)

            st.divider()

            # ── 계통 × 원인 크로스 히트맵 ──
            st.markdown("##### 고장계통 × 원인코드 크로스 분석")
            if '고장계통코드' in sdf.columns and '원인코드' in sdf.columns:
                cross = sdf.groupby(['고장계통코드','원인코드']).size().reset_index(name='건수')
                pivot_c = cross.pivot(index='고장계통코드',columns='원인코드',values='건수').fillna(0).astype(int)
                fig_cross = px.imshow(pivot_c,color_continuous_scale='Blues',text_auto=True,
                                      labels=dict(x='원인코드',y='고장계통코드',color='건수'),aspect='auto')
                fig_cross.update_layout(height=max(280,len(pivot_c)*32+80),margin=dict(t=20,b=20))
                st.plotly_chart(fig_cross,use_container_width=True)

            # ── 설비유형 × 고장계통 ──
            st.markdown("##### 설비유형 × 고장계통코드")
            if '고장계통코드' in sdf.columns:
                eq_sys = sdf.groupby(['설비유형','고장계통코드']).size().reset_index(name='건수')
                fig_eqs = px.bar(eq_sys,x='설비유형',y='건수',color='고장계통코드',barmode='stack',
                                 color_discrete_sequence=px.colors.qualitative.Set3)
                fig_eqs.update_layout(height=340,margin=dict(t=20,b=60),xaxis_tickangle=-30,
                                       legend=dict(orientation='h',y=1.05))
                st.plotly_chart(fig_eqs,use_container_width=True)

            # ── 부품교체 주기 추정 ──
            st.markdown("##### 🔩 부품교체 주기 추정")
            st.caption("조치코드='교체' 기록 기준으로 동일 설비·부위 교체 간격을 계산합니다")
            if '조치코드' in sdf.columns:
                parts_df = sdf[sdf['조치코드']=='교체'].copy()
                parts_df = parts_df[parts_df['발생일시'].notna()].sort_values('발생일시')
                parts_df['교체KEY'] = parts_df['설비_KEY'].astype(str)+'||'+parts_df['고장부위'].fillna('').astype(str)
                parts_rows = []
                for key,grp in parts_df.groupby('교체KEY'):
                    grp = grp.sort_values('발생일시')
                    n = len(grp)
                    if n < 2: continue
                    gaps = grp['발생일시'].diff().dropna().dt.days
                    gaps = gaps[gaps>0]
                    if len(gaps)==0: continue
                    parts2 = key.split('||')
                    avg_gap = gaps.mean()
                    # 다음 교체 예정일
                    next_date = (grp['발생일시'].max() + timedelta(days=avg_gap)).strftime('%Y-%m-%d')
                    parts_rows.append({
                        '설비_KEY': parts2[0],
                        '부품/부위': parts2[1] if len(parts2)>1 else '',
                        '교체횟수': n,
                        '평균교체주기_일': round(avg_gap,1),
                        '최단주기_일': int(gaps.min()),
                        '최근교체일': grp['발생일시'].max().strftime('%Y-%m-%d'),
                        '다음교체예정': next_date,
                        '설비유형': grp['설비유형'].mode()[0] if not grp['설비유형'].isna().all() else '',
                    })
                if parts_rows:
                    parts_result = pd.DataFrame(parts_rows).sort_values('평균교체주기_일')
                    st.dataframe(parts_result.reset_index(drop=True),use_container_width=True)
                    if st.button("📥 부품교체주기 Excel",key='parts_xl'):
                        out = to_excel({'부품교체주기':parts_result})
                        st.download_button("⬇️ 다운로드",data=out,
                                           file_name=f"부품교체주기_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                           key='parts_dl')
                else:
                    st.info("부품교체 기록이 2건 이상인 설비·부위가 없습니다.")

            # ── 예비품 사용 현황 ──
            st.markdown("##### 예비품 사용 현황")
            if '예비품사용여부' in sdf.columns:
                spare_grp = (sdf.groupby(['설비유형','예비품사용여부']).size()
                             .reset_index(name='건수'))
                fig_spare = px.bar(spare_grp,x='설비유형',y='건수',color='예비품사용여부',
                                   barmode='stack',
                                   color_discrete_map={'Y':'#e74c3c','N':'#95a5a6'},
                                   labels={'예비품사용여부':'예비품 사용'})
                fig_spare.update_layout(height=280,margin=dict(t=20,b=60),xaxis_tickangle=-30,
                                         legend=dict(orientation='h',y=1.05))
                st.plotly_chart(fig_spare,use_container_width=True)

            # ── 계획정비전환대상 목록 ──
            st.markdown("##### 계획정비 전환 대상 설비")
            if '계획정비전환대상' in sdf.columns:
                pm_target = (sdf[sdf['계획정비전환대상']=='Y']
                             .groupby(['설비_KEY','설비유형'])
                             .agg(건수=('소요시간','count'),총정지시간=('소요시간','sum'),재발수=('재발여부','sum'))
                             .reset_index().sort_values('총정지시간',ascending=False))
                if not pm_target.empty:
                    st.markdown(f"**{len(pm_target)}개 설비 해당** (재발 3회 이상 또는 MTTR 60분 이상)")
                    st.dataframe(pm_target.reset_index(drop=True),use_container_width=True)
                else:
                    st.info("계획정비 전환 대상 없음")


# ══════════════════════════════════════════════════════
# TAB 13 — 월보·주보 자동 작성 ★신규★
# ══════════════════════════════════════════════════════
with tab13:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.markdown("## 📝 월보 · 주보 자동 작성")
        st.caption("기간을 선택하면 보고서용 요약 문구를 자동 생성합니다")

        rep_mode = st.radio("보고서 종류", ["📅 월보", "📆 주보"], horizontal=True, key='rep_mode')

        df_rep = df[df['발생일시'].notna()].copy()
        df_rep['년']  = df_rep['발생일시'].dt.year
        df_rep['월']  = df_rep['발생일시'].dt.month
        df_rep['주차'] = df_rep['발생일시'].dt.isocalendar().week.astype(int)
        df_rep['년월'] = df_rep['발생일시'].dt.to_period('M').astype(str)

        if rep_mode == "📅 월보":
            all_yms = sorted(df_rep['년월'].unique(), reverse=True)
            rm1, rm2 = st.columns([2, 2])
            with rm1:
                sel_ym = st.selectbox("대상 월", all_yms, key='rep_ym')
            with rm2:
                # 자동 전월
                _ryr = int(sel_ym[:4]); _rmo = int(sel_ym[5:7])
                _rpyr = _ryr if _rmo>1 else _ryr-1
                _rpmo = _rmo-1 if _rmo>1 else 12
                _prev_ym_rep = f"{_rpyr}-{_rpmo:02d}"
                cmp_ym = st.selectbox("비교 월(전월)",
                    [_prev_ym_rep]+[y for y in all_yms if y!=sel_ym and y!=_prev_ym_rep],
                    key='rep_ym_cmp')

            cur_df  = df_rep[df_rep['년월']==sel_ym]
            prev_df = df_rep[df_rep['년월']==cmp_ym]
            label_cur  = f"{sel_ym} 월보"
            label_prev = cmp_ym

        else:  # 주보
            all_yrs_rep = sorted(df_rep['년'].dropna().unique().astype(int), reverse=True)
            rw1,rw2,rw3 = st.columns(3)
            with rw1: sel_yr_rep = st.selectbox("연도", all_yrs_rep, key='rep_yr')
            all_wks = sorted(df_rep[df_rep['년']==sel_yr_rep]['주차'].unique(), reverse=True)
            with rw2: sel_wk = st.selectbox("주차", all_wks, key='rep_wk')
            prev_wk = sel_wk-1 if sel_wk>1 else 52
            with rw3: cmp_wk = st.selectbox("비교 주차(전주)",
                [prev_wk]+[w for w in all_wks if w!=sel_wk and w!=prev_wk],
                key='rep_wk_cmp')
            cur_df  = df_rep[(df_rep['년']==sel_yr_rep) & (df_rep['주차']==sel_wk)]
            prev_df = df_rep[(df_rep['년']==sel_yr_rep) & (df_rep['주차']==cmp_wk)]
            label_cur  = f"{sel_yr_rep}년 {sel_wk}주차 주보"
            label_prev = f"{cmp_wk}주차"

        st.divider()

        # ── KPI 요약 수치 ──
        def _period_stats(d):
            if d.empty: return 0, 0, 0, 0
            cnt  = len(d)
            stop = d['소요시간'].sum() if d['소요시간'].notna().any() else 0
            mttr = d['소요시간'].mean() if d['소요시간'].notna().any() else 0
            recur= d['재발여부'].sum() if '재발여부' in d.columns else 0
            return cnt, stop, mttr, recur

        cc, cs, cm, cr = _period_stats(cur_df)
        pc, ps, pm_s, pr = _period_stats(prev_df)

        rk1,rk2,rk3,rk4 = st.columns(4)
        rk1.metric("고장건수", f"{cc}건", f"{cc-pc:+d}건 vs {label_prev}")
        rk2.metric("총정지시간", f"{cs:.0f}분", f"{cs-ps:+.0f}분 vs {label_prev}")
        rk3.metric("평균 MTTR", f"{cm:.0f}분", f"{cm-pm_s:+.0f}분 vs {label_prev}")
        rk4.metric("재발 고장", f"{cr}건", f"{cr-pr:+d}건 vs {label_prev}")

        st.divider()

        # ── 자동 보고서 문구 생성 ──────────────────────
        st.markdown(f"##### 📄 {label_cur} 자동 생성 문구")

        report_lines = []

        # ── 1. 총괄 ──────────────────────────────────────────
        report_lines.append(f"【총괄】")
        if prev_df.empty or pc == 0:
            report_lines.append(f"  - {label_cur} 기간 중 고장 발생 총 {cc}건, "
                                 f"총 정지시간 {cs:.0f}분(약 {cs/60:.1f}시간), 평균 MTTR {cm:.0f}분.")
        else:
            chg_cnt  = (cc-pc)/pc*100
            chg_stop = (cs-ps)/ps*100 if ps>0 else 0
            chg_mttr = (cm-pm_s)/pm_s*100 if pm_s>0 else 0
            dir_cnt  = "증가" if chg_cnt>0 else "감소"
            dir_stop = "증가" if chg_stop>0 else "감소"
            dir_mttr = "악화" if chg_mttr>0 else "개선"
            report_lines.append(f"  - 고장건수: {cc}건 — 전기({label_prev}) 대비 "
                                 f"{abs(chg_cnt):.0f}% {dir_cnt} ({pc}건 → {cc}건).")
            report_lines.append(f"  - 총 정지시간: {cs:.0f}분({cs/60:.1f}시간) — 전기 대비 "
                                 f"{abs(chg_stop):.0f}% {dir_stop} ({ps:.0f}분 → {cs:.0f}분).")
            report_lines.append(f"  - 평균 MTTR: {cm:.0f}분 — 전기({pm_s:.0f}분) 대비 "
                                 f"{abs(chg_mttr):.0f}% {dir_mttr}.")

        # ── 2. BM/PM 비율 ─────────────────────────────────────
        if '보전구분' in cur_df.columns and not cur_df.empty:
            bm_cnt = (cur_df['보전구분'] == 'BM(돌발)').sum()
            pm_cnt = (cur_df['보전구분'] == 'PM(계획)').sum()
            bm_pct = bm_cnt/cc*100 if cc else 0
            pm_pct = pm_cnt/cc*100 if cc else 0
            report_lines.append(f"\n【BM/PM 현황】")
            report_lines.append(f"  - BM(돌발): {bm_cnt}건({bm_pct:.0f}%) / "
                                 f"PM(계획): {pm_cnt}건({pm_pct:.0f}%)")
            if bm_pct >= 80:
                report_lines.append(f"  ⚠ BM 비율 {bm_pct:.0f}% — PM 전환 검토 필요")
            if not prev_df.empty and '보전구분' in prev_df.columns and pc > 0:
                prev_bm_pct = (prev_df['보전구분'] == 'BM(돌발)').sum() / pc * 100
                diff_bm = bm_pct - prev_bm_pct
                arrow = "▲" if diff_bm > 0 else "▼"
                report_lines.append(f"  - 전기 BM 비율: {prev_bm_pct:.0f}% → 금기: {bm_pct:.0f}% "
                                     f"({arrow} {abs(diff_bm):.0f}%p)")

        # ── 3. 설비유형별 현황 ────────────────────────────────
        if not cur_df.empty and '설비유형' in cur_df.columns:
            eq_rep = cur_df['설비유형'].value_counts().head(5)
            report_lines.append(f"\n【설비유형별 고장 현황】")
            for eq, cnt_eq in eq_rep.items():
                pct_eq  = cnt_eq/cc*100
                stop_eq = cur_df[cur_df['설비유형']==eq]['소요시간'].sum()
                mttr_eq = cur_df[cur_df['설비유형']==eq]['소요시간'].mean()
                report_lines.append(
                    f"  - {eq}: {cnt_eq}건({pct_eq:.0f}%), "
                    f"정지시간 {stop_eq:.0f}분, 평균MTTR {mttr_eq:.0f}분")

        # ── 4. 라인별 현황 Top 5 ─────────────────────────────
        if not cur_df.empty and '라인_차종' in cur_df.columns:
            line_rep = (cur_df.groupby('라인_차종')
                        .agg(건수=('소요시간','count'), 정지=('소요시간','sum'))
                        .sort_values('건수', ascending=False).head(5))
            report_lines.append(f"\n【라인별 고장 현황 Top 5】")
            for ln, row_ln in line_rep.iterrows():
                report_lines.append(
                    f"  - {ln}: {int(row_ln['건수'])}건, 정지 {row_ln['정지']:.0f}분")

        # ── 5. 최장 정지 단건 Top 5 ──────────────────────────
        if not cur_df.empty and '소요시간' in cur_df.columns:
            _top_cols = [c for c in ['발생일시','설비_KEY','현상','소요시간','조치자']
                         if c in cur_df.columns]
            top_stop = cur_df[_top_cols].dropna(subset=['소요시간']).nlargest(5, '소요시간')
            report_lines.append(f"\n【최장 정지 단건 Top 5】")
            for _, rs in top_stop.iterrows():
                dt_str  = rs['발생일시'].strftime('%m/%d %H:%M') if '발생일시' in rs.index and pd.notna(rs['발생일시']) else '-'
                key_str = rs.get('설비_KEY', '-')
                phenom  = str(rs.get('현상') or '-')[:25]
                worker  = str(rs.get('조치자') or '-')
                report_lines.append(
                    f"  - [{dt_str}] {key_str} | {phenom} | "
                    f"{rs['소요시간']:.0f}분 | 조치자: {worker}")

        # ── 6. 고장 계통 분析 ─────────────────────────────────
        if '고장계통코드' in cur_df.columns and not cur_df.empty:
            sys_rep = cur_df['고장계통코드'].value_counts().head(5)
            report_lines.append(f"\n【고장 계통 분析】")
            for sys_cd, cnt_s in sys_rep.items():
                stop_s = cur_df[cur_df['고장계통코드']==sys_cd]['소요시간'].sum()
                report_lines.append(
                    f"  - {sys_cd}: {cnt_s}건({cnt_s/cc*100:.0f}%), 정지 {stop_s:.0f}분")

        # ── 7. 주요 원인 Top 3 ────────────────────────────────
        if '원인코드' in cur_df.columns and not cur_df.empty:
            cause_rep = cur_df['원인코드'].value_counts().head(3)
            report_lines.append(f"\n【주요 고장 원인 Top 3】")
            for cause, cnt_c in cause_rep.items():
                report_lines.append(f"  - {cause}: {cnt_c}건({cnt_c/cc*100:.0f}%)")

        # ── 8. 재발 고장 ──────────────────────────────────────
        if '재발여부' in cur_df.columns and cr > 0:
            r_pct = cr/cc*100 if cc else 0
            report_lines.append(f"\n【재발 고장】")
            report_lines.append(f"  - 재발 고장 {cr}건(재발률 {r_pct:.0f}%) 발생.")
            if '설비_KEY' in cur_df.columns:
                top_rec = (cur_df[cur_df['재발여부']==True]
                           .groupby('설비_KEY')
                           .agg(재발건수=('재발여부','count'), 정지시간=('소요시간','sum'))
                           .sort_values('재발건수', ascending=False).head(5))
                for eq_k, row_r in top_rec.iterrows():
                    report_lines.append(
                        f"  - {eq_k}: {int(row_r['재발건수'])}건 재발, "
                        f"정지 {row_r['정지시간']:.0f}분 → 근본원인 분析 필요")

        # ── 9. 야간 돌발 현황 ─────────────────────────────────
        if '보전구분' in cur_df.columns and '발생일시' in cur_df.columns:
            bm_d = cur_df[cur_df['보전구분']=='BM(돌발)']
            if not bm_d.empty:
                night_d = bm_d[bm_d['발생일시'].dt.hour.apply(lambda h: h>=22 or h<6)]
                night_pct = len(night_d)/len(bm_d)*100
                report_lines.append(f"\n【야간 돌발 현황】")
                report_lines.append(
                    f"  - 야간(22시~06시) 돌발 고장: {len(night_d)}건 "
                    f"(BM 대비 {night_pct:.0f}%)")
                if night_pct >= 25:
                    report_lines.append(f"  ⚠ 야간 비율 높음 — 야간 단독 작업 위험도 점검 필요")

        # ── 10. 조치자별 출동 현황 ────────────────────────────
        if '조치자' in cur_df.columns and not cur_df.empty:
            from collections import Counter
            all_workers_rep = []
            for v in cur_df['조치자'].dropna():
                all_workers_rep.extend(parse_workers(str(v)))
            if all_workers_rep:
                worker_cnt_rep = Counter(all_workers_rep)
                top_workers_rep = worker_cnt_rep.most_common(5)
                report_lines.append(f"\n【조치자 출동 현황 Top 5】")
                for w_name, w_cnt in top_workers_rep:
                    report_lines.append(f"  - {w_name}: {w_cnt}건")

        # ── 11. 예비품 교체 현황 ──────────────────────────────
        if '조치코드' in cur_df.columns and not cur_df.empty:
            spare_cnt = (cur_df['조치코드'] == '교체').sum()
            spare_pct  = spare_cnt/cc*100 if cc else 0
            report_lines.append(f"\n【예비품 교체 현황】")
            report_lines.append(
                f"  - 부품 교체 {spare_cnt}건({spare_pct:.0f}%) — "
                f"예비품 재고 현황 확인 필요")

        # ── 12. MTTR 개선/악화 설비 알림 ─────────────────────
        if (not cur_df.empty and not prev_df.empty
                and '설비_KEY' in cur_df.columns
                and '소요시간' in cur_df.columns):
            cur_mttr_map  = cur_df.groupby('설비_KEY')['소요시간'].mean()
            prev_mttr_map = (prev_df.groupby('설비_KEY')['소요시간'].mean()
                             if '소요시간' in prev_df.columns else pd.Series(dtype=float))
            common_keys   = cur_mttr_map.index.intersection(prev_mttr_map.index)
            if len(common_keys) > 0:
                delta_mttr = (cur_mttr_map[common_keys] - prev_mttr_map[common_keys]).sort_values(ascending=False)
                worst3 = delta_mttr.head(3)
                best3  = delta_mttr.tail(3)
                if not worst3.empty and worst3.iloc[0] > 0:
                    report_lines.append(f"\n【MTTR 악화 설비 (전기 대비)】")
                    for eq_k, diff_v in worst3.items():
                        if diff_v > 0:
                            report_lines.append(
                                f"  - {eq_k}: +{diff_v:.0f}분 악화 "
                                f"({prev_mttr_map[eq_k]:.0f}분 → {cur_mttr_map[eq_k]:.0f}분)")
                if not best3.empty and best3.iloc[-1] < 0:
                    report_lines.append(f"\n【MTTR 개선 설비 (전기 대비)】")
                    for eq_k, diff_v in best3.items():
                        if diff_v < 0:
                            report_lines.append(
                                f"  - {eq_k}: {diff_v:.0f}분 개선 "
                                f"({prev_mttr_map[eq_k]:.0f}분 → {cur_mttr_map[eq_k]:.0f}분)")

        # ── 13. 계획정비 전환 대상 ────────────────────────────
        if '계획정비전환대상' in cur_df.columns and not cur_df.empty:
            _pm_tgt = (cur_df[cur_df['계획정비전환대상']=='Y']['설비_KEY'].unique()
                       if '설비_KEY' in cur_df.columns else [])
            if len(_pm_tgt) > 0:
                report_lines.append(f"\n【계획정비 전환 대상】")
                report_lines.append(
                    f"  - 해당 설비 {len(_pm_tgt)}개 — 재발 3회 이상 또는 MTTR 60분 이상")
                for eq_k in _pm_tgt[:5]:
                    report_lines.append(f"  · {eq_k}")
                if len(_pm_tgt) > 5:
                    report_lines.append(f"  · 외 {len(_pm_tgt)-5}개")

        # ── 14. MTTR 경고 ─────────────────────────────────────
        if cm >= st.session_state.kpi_target_mttr:
            report_lines.append(f"\n【MTTR 경고】")
            report_lines.append(
                f"  - 평균 MTTR {cm:.0f}분 — 목표 "
                f"{st.session_state.kpi_target_mttr}분 초과.")
            report_lines.append(
                f"  - 수리 장기화 설비 예비품 확보 및 수리 매뉴얼 점검 필요.")

        # ── 15. 차기 조치 사항 ────────────────────────────────
        report_lines.append(f"\n【차기 조치 사항】")
        if cc > pc:
            report_lines.append(
                f"  ① 고장건수 증가 추세({pc}건→{cc}건) — "
                f"예방보전 점검주기 단축 검토 필요.")
        if cr >= 3:
            report_lines.append(
                f"  ② 재발 고장 {cr}건 — 근본원인 분析 및 恒久대책 수립.")
        if '보전구분' in cur_df.columns:
            _bm_pct_final = (cur_df['보전구분']=='BM(돌발)').mean()*100
            if _bm_pct_final >= 80:
                report_lines.append(
                    f"  ③ BM 비율 {_bm_pct_final:.0f}% — 계획정비 전환 대상 선정 추진.")
        report_lines.append(f"  - 위험도 상위 설비 집중 관리 지속.")
        report_lines.append(f"  - 예비품 재고 현황 점검 및 긴급 발주 대상 확인.")
        report_text = "\n".join(report_lines)
        st.text_area("보고서 문구 (복사하여 사용)", value=report_text, height=420, key='rep_text')

        # 다운로드

        # ── HTML 인쇄용 다운로드 ─────────────────────────────
        def _make_html_report(text, title):
            import html as _ht
            safe = _ht.escape(text)
            return ('<!DOCTYPE html><html lang="ko"><head>'
                    '<meta charset="UTF-8">'
                    f'<title>{title}</title>'
                    '<style>'
                    'body{font-family:"Malgun Gothic","맑은 고딕",sans-serif;'
                    '     margin:0;padding:20mm 20mm 20mm 25mm;'
                    '     line-height:2.2;font-size:11pt;color:#222;}'
                    'h1{font-size:15pt;color:#1e3a5f;'
                    '   border-bottom:2px solid #1e3a5f;padding-bottom:8px;margin-bottom:20px;}'
                    'pre{white-space:pre-wrap;word-wrap:break-word;font-family:inherit;font-size:10.5pt;}'
                    '@page{size:A4;margin:18mm;}'
                    '@media print{body{padding:0;}}'
                    '</style></head><body>'
                    f'<h1>⚙ {title}</h1><pre>{safe}</pre>'
                    '</body></html>')

        html_report = _make_html_report(report_text, f'보전팀 {label_cur} 고장분석 보고서')
        col_dl1, col_dl2, col_dl3 = st.columns(3)
        with col_dl1:
            st.download_button("📋 TXT 다운로드", data=report_text.encode('utf-8-sig'),
                               file_name=f"{label_cur.replace(' ','_')}.txt",
                               mime='text/plain', use_container_width=True)

        with col_dl3:
            st.download_button(
                '🖨 HTML 인쇄용 다운로드',
                data=html_report.encode('utf-8'),
                file_name=f"{label_cur.replace(' ','_')}_보고서.html",
                mime='text/html',
                use_container_width=True,
                help='다운로드 후 브라우저에서 열어 Ctrl+P → PDF 저장 (잘림 없음)')
        with col_dl2:
            # Excel 요약표
            if not cur_df.empty:
                summary_rows = {
                    '항목': ['고장건수','총정지시간(분)','평균MTTR(분)','재발건수'],
                    f'{label_cur}': [cc, cs, cm, cr],
                    f'{label_prev}': [pc, ps, pm_s, pr],
                    '증감': [cc-pc, cs-ps, cm-pm_s, cr-pr],
                }
                out = to_excel({'보고서요약': pd.DataFrame(summary_rows)})
                st.download_button("📊 Excel 요약 다운로드", data=out,
                                   file_name=f"{label_cur.replace(' ','_')}.xlsx",
                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                   use_container_width=True)


# ══════════════════════════════════════════════════════
# TAB 14 — 출력
# ══════════════════════════════════════════════════════
with tab14:
    df = st.session_state.merged_df
    if df is None:
        st.info("Tab1에서 데이터를 불러오고 '통합 실행'을 눌러주세요.")
    else:
        st.subheader("분석 결과 출력")
        oc1,oc2 = st.columns(2)
        with oc1:
            st.markdown("#### 📊 Excel 분석결과 다운로드")
            if st.button("Excel 파일 생성",use_container_width=True,key='ex1'):
                with st.spinner("생성 중..."):
                    try:
                        mttr_r, quality_r = calc_mttr_mtbf(df)
                        wdf_all = get_worker_df(df)
                        person_r = (wdf_all.groupby('조치자')
                                    .agg(출동건수=('소요시간','count'),
                                         단독콜=('출동유형', lambda x: (x=='단독').sum()),
                                         협업콜=('출동유형', lambda x: (x=='협업').sum()),
                                         총소요시간_분=('소요시간','sum'))
                                    .reset_index().sort_values('출동건수',ascending=False))
                        person_r['협업비율(%)'] = (
                            person_r['협업콜'] / person_r['출동건수'] * 100).round(1)
                        pareto_r = (df.groupby(['라인','설비유형'])
                                    .agg(건수=('소요시간','count'),총정지_분=('소요시간','sum'))
                                    .reset_index().sort_values('건수',ascending=False))
                        # 표준코드 집계
                        code_r = (df.groupby(['고장계통코드','원인코드','조치코드'])
                                  .agg(건수=('소요시간','count'),총정지_분=('소요시간','sum'))
                                  .reset_index().sort_values('건수',ascending=False)
                                  ) if '고장계통코드' in df.columns else pd.DataFrame()
                        # 재발 집계
                        recur_r = (df.groupby('설비_KEY')
                                   .agg(전체건수=('재발여부','count'),재발건수=('재발여부','sum'))
                                   .reset_index()) if '재발여부' in df.columns else pd.DataFrame()
                        if not recur_r.empty:
                            recur_r['재발률(%)'] = (recur_r['재발건수']/recur_r['전체건수']*100).round(1)
                            recur_r = recur_r[recur_r['전체건수']>=3].sort_values('재발률(%)',ascending=False)

                        export_cols = ['발생일시','라인','설비_KEY','설비유형','고장설비','고장부위',
                                       '현상','원인','조치내역','조치자','소요시간','조치유형','고장분류',
                                       '고장계통코드','원인코드','조치코드','보전구분','재발여부',
                                       '계획정비전환대상','파일출처']
                        sheets = {
                            '통합데이터': df[[c for c in export_cols if c in df.columns]],
                            'Pareto_라인별설비': pareto_r,
                            'MTTR_MTBF': mttr_r,
                            '인원별_부하': person_r,
                        }
                        if not code_r.empty: sheets['표준코드집계'] = code_r
                        if not recur_r.empty: sheets['설비별재발률'] = recur_r
                        if not quality_r.empty: sheets['데이터품질이슈'] = quality_r
                        excel_data = to_excel(sheets)
                        st.download_button("⬇️ Excel 다운로드",data=excel_data,
                                           file_name=f"보전팀_분석_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                           use_container_width=True)
                    except Exception as e:
                        st.error(f"Excel 생성 오류: {e}")
        with oc2:
            st.markdown("#### 📋 통합 원본 CSV 다운로드")
            if st.button("CSV 생성",use_container_width=True,key='csv1'):
                csv = df.to_csv(index=False,encoding='utf-8-sig')
                st.download_button("⬇️ CSV 다운로드",data=csv.encode('utf-8-sig'),
                                   file_name=f"보전팀_통합_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                                   mime='text/csv',use_container_width=True)
        st.divider()
        st.markdown("#### 🖨 PDF 출력 방법")
        st.markdown(
            '<div style="background:#eaf4fb;border-left:4px solid #2471a3;'
            'padding:16px 20px;border-radius:6px;font-size:14px;line-height:2.2;">'
            '<b>분석 화면을 PDF로 저장하는 방법</b><br>'
            '1단계: 저장하려는 탭으로 이동<br>'
            '2단계: 키보드 <code>Ctrl+P</code> (Mac: <code>Cmd+P</code>) 입력<br>'
            '3단계: 프린터 선택 → <b>PDF로 저장</b><br>'
            '4단계: 용지 → <b>A3 가로</b> 권장 / 배율 → <b>맞춤</b><br>'
            '5단계: <b>저장</b> 클릭</div>', unsafe_allow_html=True)

        st.divider()
        st.markdown('#### 📄 HTML 보고서 직접 다운로드 (PDF 잘림 없음)')
        st.caption('다운로드된 HTML 파일을 브라우저에서 열고 Ctrl+P → PDF 저장하면 잘림 없이 인쇄됩니다.')
        if st.button('HTML 보고서 생성', key='html_rpt_btn', use_container_width=True):
            def _make_full_html(src_df):
                import html as _ht
                sections = []
                if '설비유형' in src_df.columns:
                    top_eq = src_df['설비유형'].value_counts().head(10)
                    rows_eq = ''.join(f'<tr><td>{k}</td><td>{v}건</td></tr>' for k,v in top_eq.items())
                    sections.append(f'<h2>설비유형별 고장건수</h2><table class="t"><tr><th>설비유형</th><th>건수</th></tr>{rows_eq}</table>')
                if '라인_차종' in src_df.columns:
                    top_ln = src_df['라인_차종'].value_counts().head(15)
                    rows_ln = ''.join(f'<tr><td>{k}</td><td>{v}건</td></tr>' for k,v in top_ln.items())
                    sections.append(f'<h2>라인별 고장건수 Top 15</h2><table class="t"><tr><th>라인</th><th>건수</th></tr>{rows_ln}</table>')
                body = ''.join(sections)
                return ('<!DOCTYPE html><html lang="ko"><head>'
                        '<meta charset="UTF-8"><title>보전팀 분석 보고서</title>'
                        '<style>'
                        'body{font-family:"Malgun Gothic",sans-serif;margin:0;padding:20mm;font-size:10pt;line-height:1.8;}'
                        'h1{font-size:16pt;color:#1e3a5f;border-bottom:2px solid #1e3a5f;padding-bottom:6px;}'
                        'h2{font-size:12pt;color:#1e3a5f;margin-top:16px;page-break-after:avoid;}'
                        '.t{border-collapse:collapse;width:100%;margin-bottom:16px;page-break-inside:avoid;}'
                        '.t th{background:#1e3a5f;color:#fff;padding:6px 10px;text-align:left;font-size:9pt;}'
                        '.t td{border:1px solid #ddd;padding:5px 10px;font-size:9pt;}'
                        '.t tr:nth-child(even){background:#f5f8ff;}'
                        '@page{size:A4;margin:15mm;}'
                        '</style></head><body>'
                        f'<h1>⚙ 보전팀 분析 보고서 — {datetime.now().strftime("%Y-%m-%d %H:%M")} 출력</h1>'
                        f'{body}</body></html>')
            html_full = _make_full_html(df)
            st.download_button(
                '⬇️ HTML 보고서 다운로드',
                data=html_full.encode('utf-8'),
                file_name=f"보전팀_분析보고서_{datetime.now().strftime('%Y%m%d_%H%M')}.html",
                mime='text/html',
                use_container_width=True,
                key='html_full_dl')


# ══════════════════════════════════════════════════════
# TAB 15 — POP양식 변환
# ══════════════════════════════════════════════════════
def _pop_split_line(line_raw):
    """라인명 → 차종 / 라인 분리 (첫 공백 기준)"""
    if not line_raw:
        return '', ''
    s = str(line_raw).strip()
    parts = s.split(' ', 1)
    return parts[0], parts[1].strip() if len(parts) > 1 else ''

def _pop_calc_출동시각(정지시각, 출동시간_분):
    """정지시각 + 출동시간(분) → 출동시각 역산"""
    if not isinstance(정지시각, datetime):
        return None
    try:
        mins = float(출동시간_분)
        if mins < 0 or mins > 1440:
            return 정지시각
        return 정지시각 + timedelta(minutes=mins)
    except (TypeError, ValueError):
        return 정지시각

def _pop_make_비고(비고_원본, 구분):
    """비고 | 공장구분 병기"""
    b = str(비고_원본).strip() if 비고_원본 and str(비고_원본).strip() not in ('None','nan','') else ''
    g = str(구분).strip()      if 구분      and str(구분).strip()      not in ('None','nan','') else ''
    if b and g:  return f"{b} | {g}"
    if g:        return g
    if b:        return b
    return None

def convert_pop_to_excel(df_src):
    """
    POP 고장이력 DataFrame → 차체 설비보전 형식 DataFrame 변환
    입력 컬럼: 구분, 라인명, 정지시각, 출동시간, 완료시각, 소요시각,
               설비유형, 고장설비, 고장부위, 현상, 원인, 조치, 비고, 조치자
    """
    OUTPUT_COLS = ['년','월','일','주','라인','라인_KEY','차종','설비유형',
                   '고장설비','고장부위','고장부위_STD','현상','원인','조치',
                   '소요시간','정지시각','출동시각','완료시각','조치자','비고','NO']

    records = []
    skip_cnt = 0

    for no_idx, row in enumerate(df_src.itertuples(index=False), 1):
        구분      = getattr(row, '구분',     None)
        라인명_raw = getattr(row, '라인명',   None)
        정지시각_v = getattr(row, '정지시각', None)
        출동시간_v = getattr(row, '출동시간', None)
        완료시각_v = getattr(row, '완료시각', None)
        소요시각_v = getattr(row, '소요시각', None)
        설비유형_v = getattr(row, '설비유형', None)
        고장설비_v = getattr(row, '고장설비', None)
        고장부위_v = getattr(row, '고장부위', None)
        현상_v     = getattr(row, '현상',     None)
        원인_v     = getattr(row, '원인',     None)
        조치_v     = getattr(row, '조치',     None)
        비고_v     = getattr(row, '비고',     None)
        조치자_v   = getattr(row, '조치자',   None)

        # datetime 변환
        정지시각 = parse_dt(정지시각_v)
        완료시각 = parse_dt(완료시각_v)
        정지시각 = sanitize_dt(정지시각)
        완료시각 = sanitize_dt(완료시각)

        기준dt = 정지시각 or 완료시각
        if 기준dt is None:
            skip_cnt += 1
            continue

        출동시각 = _pop_calc_출동시각(정지시각, 출동시간_v)
        차종, 라인 = _pop_split_line(라인명_raw)

        try:
            주 = 기준dt.isocalendar()[1]
        except Exception:
            주 = None

        records.append({
            '년':          기준dt.year,
            '월':          기준dt.month,
            '일':          기준dt.day,
            '주':          주,
            '라인':        라인,
            '라인_KEY':    라인,
            '차종':        차종,
            '설비유형':    설비유형_v,
            '고장설비':    고장설비_v,
            '고장부위':    고장부위_v,
            '고장부위_STD': 고장부위_v,
            '현상':        현상_v,
            '원인':        원인_v,
            '조치':        조치_v,
            '소요시간':    소요시각_v,
            '정지시각':    정지시각,
            '출동시각':    출동시각,
            '완료시각':    완료시각,
            '조치자':      조치자_v,
            '비고':        _pop_make_비고(비고_v, 구분),
            'NO':          no_idx,
        })

    df_out = pd.DataFrame(records, columns=OUTPUT_COLS)
    return df_out, skip_cnt

def pop_df_to_excel_bytes(df):
    """변환된 DataFrame → xlsx bytes (다운로드용)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    hdr_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill   = PatternFill('solid', start_color='1E3A5F')
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side  = Side(style='thin', color='AAAAAA')
    bdr        = Border(left=thin_side, right=thin_side,
                        top=thin_side,  bottom=thin_side)
    data_font  = Font(name='Arial', size=9)
    data_align = Alignment(vertical='center')
    alt_fill   = PatternFill('solid', start_color='F4F7FB')

    COL_WIDTHS = [6,5,5,5,22,22,10,10,12,18,18,40,40,40,10,18,18,18,16,30,7]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df)+1}"

    # 헤더
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = bdr

    # 데이터
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        is_alt = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row, 1):
            v = None if (isinstance(val, float) and pd.isna(val)) else val
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = data_font; cell.border = bdr
            cell.alignment = data_align
            if is_alt: cell.fill = alt_fill
            if c_idx in (16,17,18) and isinstance(v, datetime):
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            if c_idx == 15 and v is not None:
                cell.number_format = '#,##0.0'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

with tab15:
    st.markdown('<div class="main-title">🔄 POP양식 변환</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">POP 고장이력(광주) → 차체 설비보전 형식으로 변환</div>',
                unsafe_allow_html=True)

    # ── 입력 컬럼 안내 ──────────────────────────────
    with st.expander("📋 입력 파일 컬럼 형식 안내", expanded=False):
        st.markdown("""
| 순서 | 컬럼명 | 설명 |
|---:|---|---|
| 1 | 구분 | 공장/동 구분 (예: 하남공장 7동) |
| 2 | 라인명 | 차종+라인 포함 (예: NQ5 RR FLR COMPL) |
| 3 | 정지시각 | datetime (YYYY-MM-DD HH:MM:SS) |
| 4 | 출동시간 | 숫자(분) — 정지시각 기준 역산 |
| 5 | 완료시각 | datetime |
| 6 | 소요시각 | 숫자(분) |
| 7 | 설비유형 | 로봇 / 지그 / 생산 등 |
| 8 | 고장설비 | R03, A01 등 |
| 9 | 고장부위 | 일시정지, 에러 등 |
| 10 | 현상 | 고장 현상 설명 |
| 11 | 원인 | 고장 원인 |
| 12 | 조치 | 조치 내용 |
| 13 | 비고 | 비고 (없으면 공백) |
| 14 | 조치자 | 담당자 이름 |
        """)

    st.divider()

    # ── 파일 업로드 ──────────────────────────────────
    pop_file = st.file_uploader(
        "📂 POP 고장이력 파일 업로드 (.xlsx)",
        type=['xlsx'],
        key='pop_uploader',
        help="광주공장 POP 시스템 고장이력 엑셀 파일을 업로드하세요."
    )

    if pop_file is not None:
        try:
            with st.spinner("파일 읽는 중..."):
                df_pop_raw = pd.read_excel(pop_file, sheet_name=0, header=0)

            # 컬럼명 강제 지정 (순서 기반)
            POP_COLS = ['구분','라인명','정지시각','출동시간','완료시각','소요시각',
                        '설비유형','고장설비','고장부위','현상','원인','조치','비고','조치자']
            if len(df_pop_raw.columns) >= len(POP_COLS):
                df_pop_raw.columns = POP_COLS + list(df_pop_raw.columns[len(POP_COLS):])
            else:
                st.markdown(
                    f'<div class="err-box">❌ 컬럼 수 부족 — '
                    f'필요: {len(POP_COLS)}개 / 실제: {len(df_pop_raw.columns)}개</div>',
                    unsafe_allow_html=True)
                st.stop()

            total_raw = len(df_pop_raw)

            # ── 원본 미리보기 ──────────────────────
            st.markdown("#### 📄 원본 데이터 미리보기 (상위 5행)")
            st.dataframe(df_pop_raw.head(5), use_container_width=True)

            st.markdown(
                f'<div class="ok-box">✅ 파일 로드 완료 — 총 <b>{total_raw:,}행</b></div>',
                unsafe_allow_html=True)

            st.divider()

            # ── 변환 실행 ──────────────────────────
            if st.button("⚡ 변환 실행", use_container_width=True, key='pop_convert_btn',
                         type='primary'):
                with st.spinner(f"변환 중... ({total_raw:,}행)"):
                    df_converted, skip_cnt = convert_pop_to_excel(df_pop_raw)

                st.session_state['pop_converted_df']  = df_converted
                st.session_state['pop_skip_cnt']      = skip_cnt
                st.session_state['pop_total_raw']     = total_raw
                st.session_state['pop_source_fname']  = pop_file.name

        except Exception as e:
            st.markdown(f'<div class="err-box">❌ 파일 읽기 오류: {e}</div>',
                        unsafe_allow_html=True)

    # ── 변환 결과 표시 ───────────────────────────────
    if st.session_state.get('pop_converted_df') is not None:
        df_conv   = st.session_state['pop_converted_df']
        skip_cnt  = st.session_state['pop_skip_cnt']
        total_raw = st.session_state['pop_total_raw']
        fname     = st.session_state.get('pop_source_fname', '')

        conv_cnt  = len(df_conv)

        st.divider()
        st.markdown("#### 📊 변환 결과 통계")

        sc1, sc2, sc3 = st.columns(3)
        sc1.metric("원본 행수",   f"{total_raw:,}행")
        sc2.metric("변환 성공",   f"{conv_cnt:,}행",
                   f"-{skip_cnt}행 (날짜없음 제외)" if skip_cnt > 0 else "전량 변환")
        sc3.metric("Skip(날짜없음)", f"{skip_cnt}행")

        # 공장별 건수
        if '비고' in df_conv.columns:
            st.markdown("##### 공장별 건수")
            def _extract_plant(b):
                if not b: return '(미상)'
                parts = str(b).split('|')
                return parts[-1].strip() if parts else '(미상)'
            plant_cnt = df_conv['비고'].apply(_extract_plant).value_counts().reset_index()
            plant_cnt.columns = ['공장구분','건수']
            st.dataframe(plant_cnt, use_container_width=True, hide_index=True)

        # 설비유형별 건수
        if '설비유형' in df_conv.columns:
            st.markdown("##### 설비유형별 건수")
            eq_cnt = df_conv['설비유형'].value_counts().reset_index()
            eq_cnt.columns = ['설비유형','건수']
            st.dataframe(eq_cnt, use_container_width=True, hide_index=True)

        st.divider()

        # ── 변환 결과 미리보기 ──────────────────────
        st.markdown("#### 🔍 변환 결과 미리보기 (상위 20행)")
        st.dataframe(df_conv.head(20), use_container_width=True)

        st.divider()

        # ── 다운로드 ────────────────────────────────
        st.markdown("#### ⬇️ 변환 파일 다운로드")

        with st.spinner("Excel 파일 생성 중..."):
            excel_bytes = pop_df_to_excel_bytes(df_conv)

        out_fname = fname.replace('.xlsx','').replace('.XLSX','')
        out_fname = f"{out_fname}_차체형식변환_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        st.download_button(
            label=f"📥 변환 파일 다운로드 ({conv_cnt:,}행)",
            data=excel_bytes,
            file_name=out_fname,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
            type='primary',
        )

        st.markdown(
            '<div class="ok-box">'
            '✅ 변환 완료 — 다운로드 후 <b>보전팀 통합 분석 시스템 Tab1</b>에서 '
            '"로봇/지그" 파일로 업로드하여 분석하세요.'
            '</div>',
            unsafe_allow_html=True)

    elif pop_file is None:
        st.markdown(
            '<div class="warn-box">'
            '⬆️ POP 고장이력 xlsx 파일을 업로드하면 변환을 시작할 수 있습니다.'
            '</div>',
            unsafe_allow_html=True)
